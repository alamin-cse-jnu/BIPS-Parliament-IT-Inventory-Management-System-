"""
Views for Devices app in PIMS (Parliament IT Inventory Management System)
Bangladesh Parliament Secretariat

This module defines views for device management including:
- Device CRUD operations with JSON specifications support
- Category and subcategory management
- Device filtering, search, and status-based views
- QR code generation and management
- Warranty tracking and management
- Bulk operations and exports
- Dashboard and reporting views
- AJAX endpoints for dynamic functionality

Features:
- Modern class-based views with Bootstrap 5.3 styling
- Comprehensive device management with flexible JSON specifications
- Interactive search and filtering capabilities
- QR code generation and bulk operations
- Export functionality (Excel, PDF, CSV)
- Real-time AJAX updates
"""

from urllib import request
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.db.models import Count, Sum, Avg, Q, F, Case, When, DecimalField, Max
from django.http import JsonResponse, HttpResponse, Http404
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.utils.decorators import method_decorator
from django.db import transaction
from datetime import date, timedelta, datetime
import json
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import qrcode
from PIL import Image
import zipfile

from .models import DeviceCategory, DeviceSubcategory, Device, Warranty, QRCode
from .forms import (
    DeviceCategoryForm, DeviceSubcategoryForm, DeviceForm,
    DeviceFilterForm, DeviceSearchForm, WarrantyForm,
    DeviceBulkUpdateForm
)
from pims.utils.qr_code import (
    create_device_qr_code, 
    bulk_generate_device_qr_codes,
    get_qr_code_for_device
)
from vendors.models import Vendor
from locations.models import Location
from users.models import CustomUser
from django.views.decorators.cache import cache_page
import calendar


# Mixins
class DevicePermissionMixin(PermissionRequiredMixin):
    """Base permission mixin for device views."""
    permission_required = 'devices.view_device'


class DeviceDashboardView(LoginRequiredMixin, TemplateView):
    """
    Main dashboard for device management with real-time statistics.
    """
    template_name = 'devices/dashboard.html'
    
    @method_decorator(cache_page(300))  # Cache for 5 minutes
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Basic statistics
        total_devices = Device.objects.filter(is_active=True).count()
        
        status_stats = Device.objects.filter(is_active=True).aggregate(
            available=Count(Case(When(status='AVAILABLE', then=1))),
            assigned=Count(Case(When(status='ASSIGNED', then=1))),
            maintenance=Count(Case(When(status='MAINTENANCE', then=1))),
            retired=Count(Case(When(status='RETIRED', then=1)))
        )
        
        # Category distribution
        category_stats = (
            Device.objects.filter(is_active=True)
            .values('subcategory__category__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )
        
        # Recent activities
        recent_devices = Device.objects.filter(is_active=True).order_by('-created_at')[:5]
        
        # Value statistics
        value_stats = Device.objects.filter(is_active=True).aggregate(
            total_value=Sum('purchase_price', output_field=DecimalField()),
            avg_value=Avg('purchase_price', output_field=DecimalField())
        )
        
        # Warranty alerts
        today = datetime.now().date()
        warranty_alerts = Warranty.objects.filter(
            device__is_active=True,
            end_date__lte=today + timedelta(days=30),
            end_date__gte=today
        ).count()
        
        # Monthly trends (last 6 months)
        months_data = []
        for i in range(6):
            month_date = datetime.now().date().replace(day=1) - timedelta(days=30*i)
            month_devices = Device.objects.filter(
                created_at__year=month_date.year,
                created_at__month=month_date.month,
                is_active=True
            ).count()
            months_data.append({
                'month': calendar.month_name[month_date.month],
                'count': month_devices
            })
        
        context.update({
            'total_devices': total_devices,
            'status_stats': status_stats,
            'category_stats': category_stats,
            'recent_devices': recent_devices,
            'value_stats': value_stats,
            'warranty_alerts': warranty_alerts,
            'months_data': list(reversed(months_data)),
            'page_title': 'Device Management Dashboard',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context

# Device CRUD Views
class DeviceListView(LoginRequiredMixin, ListView):
    """List view for devices with filtering and search."""
    model = Device
    template_name = 'devices/list.html'
    context_object_name = 'devices'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Device.objects.filter(is_active=True).select_related(
            'subcategory__category', 'vendor', 'current_location'
        ).prefetch_related('warranties')
        
        # Apply search
        search_form = DeviceSearchForm(self.request.GET)
        if search_form.is_valid():
            queryset = search_form.search(queryset)
        
        # Apply filters
        filter_form = DeviceFilterForm(self.request.GET)
        if filter_form.is_valid():
            queryset = self.apply_filters(queryset, filter_form)
        
        return queryset.order_by('-created_at')
    
    def apply_filters(self, queryset, form):
        """Apply filter form to queryset."""
        if form.cleaned_data.get('category'):
            queryset = queryset.filter(subcategory__category=form.cleaned_data['category'])
        
        if form.cleaned_data.get('subcategory'):
            queryset = queryset.filter(subcategory=form.cleaned_data['subcategory'])
        
        if form.cleaned_data.get('device_type'):
            queryset = queryset.filter(device_type=form.cleaned_data['device_type'])
        
        if form.cleaned_data.get('status'):
            queryset = queryset.filter(status=form.cleaned_data['status'])
        
        if form.cleaned_data.get('condition'):
            queryset = queryset.filter(condition=form.cleaned_data['condition'])
        
        if form.cleaned_data.get('vendor'):
            queryset = queryset.filter(vendor=form.cleaned_data['vendor'])
        
        if form.cleaned_data.get('location'):
            queryset = queryset.filter(current_location=form.cleaned_data['location'])
        
        if form.cleaned_data.get('purchase_date_from'):
            queryset = queryset.filter(purchase_date__gte=form.cleaned_data['purchase_date_from'])
        
        if form.cleaned_data.get('purchase_date_to'):
            queryset = queryset.filter(purchase_date__lte=form.cleaned_data['purchase_date_to'])
        
        if form.cleaned_data.get('is_assignable') == 'true':
            queryset = queryset.filter(is_assignable=True)
        elif form.cleaned_data.get('is_assignable') == 'false':
            queryset = queryset.filter(is_assignable=False)
        
        # Warranty status filtering
        warranty_status = form.cleaned_data.get('warranty_status')
        if warranty_status == 'under_warranty':
            queryset = queryset.filter(
                warranties__start_date__lte=date.today(),
                warranties__end_date__gte=date.today(),
                warranties__is_active=True
            )
        elif warranty_status == 'expires_soon':
            thirty_days_later = date.today() + timedelta(days=30)
            queryset = queryset.filter(
                warranties__end_date__range=[date.today(), thirty_days_later],
                warranties__is_active=True
            )
        elif warranty_status == 'no_warranty':
            queryset = queryset.exclude(
                warranties__start_date__lte=date.today(),
                warranties__end_date__gte=date.today(),
                warranties__is_active=True
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = DeviceSearchForm(self.request.GET)
        context['filter_form'] = DeviceFilterForm(self.request.GET)
        context['title'] = 'All Devices'
        return context


class DeviceDetailView(LoginRequiredMixin, DetailView):
    """Detail view for devices."""
    model = Device
    template_name = 'devices/detail.html'
    context_object_name = 'device'
    
    def get_queryset(self):
        return Device.objects.select_related(
            'subcategory__category', 'vendor', 'current_location', 'parent_device', 'created_by'
        ).prefetch_related(
            'warranties__provider', 'qr_codes', 'components'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device = self.object
        
        # Get active warranties
        active_warranties = device.warranties.filter(is_active=True).order_by('-end_date')
        
        # Get QR codes
        qr_codes = device.qr_codes.filter(is_active=True)
        
        # Get components if this is a complete device
        components = None
        if device.device_type == 'COMPLETE':
            components = device.components.filter(is_active=True).select_related(
                'subcategory__category', 'vendor'
            )
        
        # Get current assignment
        current_assignment = device.get_current_assignment()
        
        # Get maintenance history
        maintenance_history = device.get_maintenance_history()[:5]
        
        context.update({
            'active_warranties': active_warranties,
            'qr_codes': qr_codes,
            'components': components,
            'current_assignment': current_assignment,
            'maintenance_history': maintenance_history,
        })
        
        return context


import logging

logger = logging.getLogger(__name__)

class DeviceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create view for devices."""
    model = Device
    form_class = DeviceForm
    template_name = 'devices/create.html'
    permission_required = 'devices.add_device'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Device'
        context['form_action'] = 'Create'
        context['breadcrumb_title'] = 'Create Device'
        context['location'] = 'Bangladesh Parliament Secretariat, Dhaka'
        return context
    
    def form_valid(self, form):
        """Handle successful form submission."""
        try:
            # Set created_by field
            form.instance.created_by = self.request.user
            
            # Save the device using transaction for data integrity
            with transaction.atomic():
                # Call parent form_valid which saves the object
                response = super().form_valid(form)
                
                # Add success message
                messages.success(
                    self.request,
                    f'Device "{self.object.brand} {self.object.model}" created successfully! '
                    f'Device ID: {self.object.device_id}'
                )
                
                logger.info(f"Device created successfully: {self.object.device_id} by user {self.request.user}")
                
                return response
            
        except Exception as e:
            # Log the error with full details
            logger.error(f"Error creating device: {str(e)}", exc_info=True)
            
            # Add error message for user
            messages.error(
                self.request,
                f'Error creating device: {str(e)}. Please try again.'
            )
            
            # Return to form with errors
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        """Handle form validation errors."""
        # Log form errors for debugging
        logger.warning(f"Device creation form errors: {form.errors}")
        
        # Add user-friendly error message
        messages.error(
            self.request, 
            'Please correct the errors below and try again.'
        )
        
        return super().form_invalid(form)
    
    def get_success_url(self):
        """Redirect to device detail page after successful creation."""
        return reverse('devices:detail', kwargs={'pk': self.object.pk})

class DeviceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update view for devices."""
    model = Device
    form_class = DeviceForm
    template_name = 'devices/edit.html'
    permission_required = 'devices.change_device'
    
    def get_success_url(self):
        return reverse('devices:detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(
            self.request,
            f'Device {form.instance.device_id} updated successfully!'
        )
        return super().form_valid(form)


class DeviceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete view for devices."""
    model = Device
    template_name = 'devices/delete.html'
    permission_required = 'devices.delete_device'
    success_url = reverse_lazy('devices:list')
    
    def delete(self, request, *args, **kwargs):
        device = self.get_object()
        device_id = device.device_id
        
        # Check if device can be deleted
        if device.status == 'ASSIGNED':
            messages.error(request, 'Cannot delete assigned devices. Please return the device first.')
            return redirect('devices:detail', pk=device.pk)
        
        messages.success(request, f'Device {device_id} deleted successfully!')
        return super().delete(request, *args, **kwargs)

class DeviceSummaryView(LoginRequiredMixin, TemplateView):
    """
    Device summary with detailed breakdowns and analytics.
    """
    template_name = 'devices/summary.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Detailed status breakdown
        devices = Device.objects.filter(is_active=True)
        
        status_breakdown = {}
        for status, label in Device.STATUS_CHOICES:
            count = devices.filter(status=status).count()
            status_breakdown[status] = {
                'label': label,
                'count': count,
                'percentage': round((count / devices.count() * 100) if devices.count() > 0 else 0, 1)
            }
        
        # Condition breakdown
        condition_breakdown = {}
        for condition, label in Device.CONDITION_CHOICES:
            count = devices.filter(condition=condition).count()
            condition_breakdown[condition] = {
                'label': label,
                'count': count,
                'percentage': round((count / devices.count() * 100) if devices.count() > 0 else 0, 1)
            }
        
        # Category-wise analysis
        category_analysis = (
            devices.values('subcategory__category__name')
            .annotate(
                total=Count('id'),
                available=Count(Case(When(status='AVAILABLE', then=1))),
                assigned=Count(Case(When(status='ASSIGNED', then=1))),
                maintenance=Count(Case(When(status='MAINTENANCE', then=1))),
                total_value=Sum('purchase_price', output_field=DecimalField()),
                avg_value=Avg('purchase_price', output_field=DecimalField())
            )
            .order_by('-total')
        )
        
        # Vendor analysis
        vendor_analysis = (
            devices.values('vendor__name')
            .annotate(
                total_devices=Count('id'),
                total_value=Sum('purchase_price', output_field=DecimalField()),
                maintenance_count=Count(Case(When(status='MAINTENANCE', then=1)))
            )
            .order_by('-total_devices')[:10]
        )
        
        # Age analysis
        age_ranges = {
            'new': devices.filter(purchase_date__gte=datetime.now().date() - timedelta(days=365)).count(),
            'recent': devices.filter(
                purchase_date__gte=datetime.now().date() - timedelta(days=365*2),
                purchase_date__lt=datetime.now().date() - timedelta(days=365)
            ).count(),
            'mature': devices.filter(
                purchase_date__gte=datetime.now().date() - timedelta(days=365*5),
                purchase_date__lt=datetime.now().date() - timedelta(days=365*2)
            ).count(),
            'old': devices.filter(purchase_date__lt=datetime.now().date() - timedelta(days=365*5)).count()
        }
        
        context.update({
            'status_breakdown': status_breakdown,
            'condition_breakdown': condition_breakdown,
            'category_analysis': category_analysis,
            'vendor_analysis': vendor_analysis,
            'age_ranges': age_ranges,
            'total_devices': devices.count(),
            'page_title': 'Device Summary Report',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context


class DeviceAlertsView(LoginRequiredMixin, TemplateView):
    """
    Device alerts and notifications dashboard for proactive management.
    """
    template_name = 'devices/alerts.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        today = datetime.now().date()
        
        # Warranty alerts
        warranty_expiring_30 = Warranty.objects.filter(
            device__is_active=True,
            end_date__lte=today + timedelta(days=30),
            end_date__gte=today
        ).select_related('device')
        
        warranty_expired = Warranty.objects.filter(
            device__is_active=True,
            end_date__lt=today
        ).select_related('device')
        
        # Maintenance alerts
        overdue_maintenance = Device.objects.filter(
            is_active=True,
            status='MAINTENANCE',
            updated_at__lt=timezone.now() - timedelta(days=30)
        )
        
        # Device age alerts (devices older than 5 years)
        aging_devices = Device.objects.filter(
            is_active=True,
            purchase_date__lt=today - timedelta(days=365*5)
        )
        
        # Unassigned valuable devices (price > 50,000 BDT)
        unassigned_valuable = Device.objects.filter(
            is_active=True,
            status='AVAILABLE',
            purchase_price__gt=50000
        )
        
        # Devices without location
        devices_no_location = Device.objects.filter(
            is_active=True,
            current_location__isnull=True
        )
        
        # Devices with incomplete data
        incomplete_devices = Device.objects.filter(
            is_active=True
        ).filter(
            Q(brand__isnull=True) | Q(brand='') |
            Q(model__isnull=True) | Q(model='') |
            Q(serial_number__isnull=True) | Q(serial_number='') |
            Q(purchase_date__isnull=True) |
            Q(purchase_price__isnull=True)
        )
        
        # Alert summary
        alert_counts = {
            'warranty_expiring': warranty_expiring_30.count(),
            'warranty_expired': warranty_expired.count(),
            'overdue_maintenance': overdue_maintenance.count(),
            'aging_devices': aging_devices.count(),
            'unassigned_valuable': unassigned_valuable.count(),
            'no_location': devices_no_location.count(),
            'incomplete_data': incomplete_devices.count()
        }
        
        total_alerts = sum(alert_counts.values())
        
        # Priority alerts (high impact)
        priority_alerts = []
        
        if warranty_expired.count() > 0:
            priority_alerts.append({
                'type': 'danger',
                'title': 'Expired Warranties',
                'count': warranty_expired.count(),
                'message': 'Devices with expired warranties need immediate attention',
                'url': 'devices:warranty_expired'
            })
        
        if warranty_expiring_30.count() > 0:
            priority_alerts.append({
                'type': 'warning', 
                'title': 'Warranties Expiring Soon',
                'count': warranty_expiring_30.count(),
                'message': 'Warranties expiring within 30 days',
                'url': 'devices:warranty_expiring'
            })
        
        if unassigned_valuable.count() > 0:
            priority_alerts.append({
                'type': 'info',
                'title': 'Unassigned Valuable Devices',
                'count': unassigned_valuable.count(),
                'message': 'High-value devices not currently assigned',
                'url': 'devices:available'
            })
        
        context.update({
            'warranty_expiring_30': warranty_expiring_30,
            'warranty_expired': warranty_expired,
            'overdue_maintenance': overdue_maintenance,
            'aging_devices': aging_devices,
            'unassigned_valuable': unassigned_valuable,
            'devices_no_location': devices_no_location,
            'incomplete_devices': incomplete_devices,
            'alert_counts': alert_counts,
            'total_alerts': total_alerts,
            'priority_alerts': priority_alerts,
            'page_title': 'Device Alerts & Notifications',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context

# Status-based Device Views
class AvailableDevicesView(DeviceListView):
    """View for available devices."""
    template_name = 'devices/available.html'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(status='AVAILABLE')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Available Devices'
        return context


class AssignedDevicesView(DeviceListView):
    """View for assigned devices."""
    template_name = 'devices/assigned.html'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(status='ASSIGNED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Assigned Devices'
        return context


class MaintenanceDevicesView(DeviceListView):
    """View for devices in maintenance."""
    template_name = 'devices/maintenance.html'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(status='MAINTENANCE')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Devices in Maintenance'
        return context


class RetiredDevicesView(DeviceListView):
    """View for retired devices."""
    template_name = 'devices/retired.html'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(status='RETIRED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Retired Devices'
        return context


# Search and Filter Views
class DeviceSearchView(DeviceListView):
    """Search view for devices."""
    template_name = 'devices/search_results.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('query', '')
        context['query'] = query
        context['title'] = f'Search Results for "{query}"'
        return context


class DeviceFilterView(DeviceListView):
    """Filter view for devices."""
    template_name = 'devices/filter_results.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Filtered Devices'
        return context


# Category Management Views
class CategoryListView(LoginRequiredMixin, ListView):
    """List view for device categories."""
    model = DeviceCategory
    template_name = 'devices/categories/list.html'
    context_object_name = 'categories'
    ordering = ['sort_order', 'name']
    
    def get_queryset(self):
        return DeviceCategory.objects.annotate(
            device_count=Count('subcategories__devices'),
            subcategory_count=Count('subcategories')
        )


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create view for device categories."""
    model = DeviceCategory
    form_class = DeviceCategoryForm
    template_name = 'devices/categories/create.html'
    permission_required = 'devices.add_devicecategory'
    success_url = reverse_lazy('devices:category_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Category "{form.instance.name}" created successfully!')
        return super().form_valid(form)


class CategoryDetailView(LoginRequiredMixin, DetailView):
    """Detail view for device categories."""
    model = DeviceCategory
    template_name = 'devices/categories/detail.html'
    context_object_name = 'category'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = self.object
        
        # Get subcategories
        subcategories = category.subcategories.filter(is_active=True).annotate(
            device_count=Count('devices')
        )
        
        # Get recent devices in this category
        recent_devices = Device.objects.filter(
            subcategory__category=category,
            is_active=True
        ).order_by('-created_at')[:10]
        
        context.update({
            'subcategories': subcategories,
            'recent_devices': recent_devices,
        })
        
        return context


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update view for device categories."""
    model = DeviceCategory
    form_class = DeviceCategoryForm
    template_name = 'devices/categories/edit.html'
    permission_required = 'devices.change_devicecategory'
    success_url = reverse_lazy('devices:category_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Category "{form.instance.name}" updated successfully!')
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete view for device categories."""
    model = DeviceCategory
    template_name = 'devices/categories/delete.html'
    permission_required = 'devices.delete_devicecategory'
    success_url = reverse_lazy('devices:category_list')
    
    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        
        # Check if category has devices
        device_count = Device.objects.filter(subcategory__category=category).count()
        if device_count > 0:
            messages.error(
                request,
                f'Cannot delete category "{category.name}". It has {device_count} associated devices.'
            )
            return redirect('devices:category_detail', pk=category.pk)
        
        messages.success(request, f'Category "{category.name}" deleted successfully!')
        return super().delete(request, *args, **kwargs)


# Subcategory Management Views
class SubcategoryListView(LoginRequiredMixin, ListView):
    """List view for device subcategories."""
    model = DeviceSubcategory
    template_name = 'devices/subcategories/list.html'
    context_object_name = 'subcategories'
    ordering = ['category', 'sort_order', 'name']
    
    def get_queryset(self):
        return DeviceSubcategory.objects.select_related('category').annotate(
            device_count=Count('devices')
        )


class SubcategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create view for device subcategories."""
    model = DeviceSubcategory
    form_class = DeviceSubcategoryForm
    template_name = 'devices/subcategories/create.html'
    permission_required = 'devices.add_devicesubcategory'
    success_url = reverse_lazy('devices:subcategory_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Subcategory "{form.instance.name}" created successfully!')
        return super().form_valid(form)


class SubcategoryDetailView(LoginRequiredMixin, DetailView):
    """Detail view for device subcategories."""
    model = DeviceSubcategory
    template_name = 'devices/subcategories/detail.html'
    context_object_name = 'subcategory'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        subcategory = self.object
        
        # Get devices in this subcategory
        devices = Device.objects.filter(
            subcategory=subcategory,
            is_active=True
        ).select_related('vendor', 'current_location').order_by('-created_at')
        
        context['devices'] = devices
        return context


class SubcategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update view for device subcategories."""
    model = DeviceSubcategory
    form_class = DeviceSubcategoryForm
    template_name = 'devices/subcategories/edit.html'
    permission_required = 'devices.change_devicesubcategory'
    success_url = reverse_lazy('devices:subcategory_list')
    
    def form_valid(self, form):
        messages.success(self.request, f'Subcategory "{form.instance.name}" updated successfully!')
        return super().form_valid(form)


class SubcategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete view for device subcategories."""
    model = DeviceSubcategory
    template_name = 'devices/subcategories/delete.html'
    permission_required = 'devices.delete_devicesubcategory'
    success_url = reverse_lazy('devices:subcategory_list')
    
    def delete(self, request, *args, **kwargs):
        subcategory = self.get_object()
        
        # Check if subcategory has devices
        device_count = subcategory.devices.count()
        if device_count > 0:
            messages.error(
                request,
                f'Cannot delete subcategory "{subcategory.name}". It has {device_count} associated devices.'
            )
            return redirect('devices:subcategory_detail', pk=subcategory.pk)
        
        messages.success(request, f'Subcategory "{subcategory.name}" deleted successfully!')
        return super().delete(request, *args, **kwargs)


# Component Management Views
class DeviceComponentsView(LoginRequiredMixin, DetailView):
    """View for managing device components."""
    model = Device
    template_name = 'devices/components.html'
    context_object_name = 'device'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device = self.object
        
        if device.device_type != 'COMPLETE':
            raise Http404("Only complete devices can have components")
        
        components = device.components.filter(is_active=True).select_related(
            'subcategory__category', 'vendor'
        ).order_by('subcategory__name', 'brand', 'model')
        
        context['components'] = components
        return context


class AddComponentView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """View for adding components to devices."""
    model = Device
    form_class = DeviceForm
    template_name = 'devices/add_component.html'
    permission_required = 'devices.add_device'
    
    def dispatch(self, request, *args, **kwargs):
        self.parent_device = get_object_or_404(Device, pk=kwargs['pk'])
        if self.parent_device.device_type != 'COMPLETE':
            messages.error(request, 'Components can only be added to complete devices.')
            return redirect('devices:detail', pk=self.parent_device.pk)
        return super().dispatch(request, *args, **kwargs)
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Set device type to component and parent device
        form.fields['device_type'].initial = 'COMPONENT'
        form.fields['parent_device'].initial = self.parent_device
        form.fields['device_type'].widget.attrs['readonly'] = True
        form.fields['parent_device'].widget.attrs['readonly'] = True
        return form
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.device_type = 'COMPONENT'
        form.instance.parent_device = self.parent_device
        messages.success(
            self.request,
            f'Component {form.instance.brand} {form.instance.model} added to {self.parent_device.device_id}!'
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('devices:components', kwargs={'pk': self.parent_device.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['parent_device'] = self.parent_device
        return context


# Warranty Management Views
class WarrantyListView(LoginRequiredMixin, ListView):
    """List view for warranties."""
    model = Warranty
    template_name = 'devices/warranties/list.html'
    context_object_name = 'warranties'
    paginate_by = 25
    ordering = ['-end_date']
    
    def get_queryset(self):
        return Warranty.objects.select_related(
            'device', 'provider'
        ).filter(is_active=True)


class WarrantyCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create view for warranties."""
    model = Warranty
    form_class = WarrantyForm
    template_name = 'devices/warranties/create.html'
    permission_required = 'devices.add_warranty'
    success_url = reverse_lazy('devices:warranty_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Warranty created successfully!')
        return super().form_valid(form)


class WarrantyDetailView(LoginRequiredMixin, DetailView):
    """Detail view for warranties."""
    model = Warranty
    template_name = 'devices/warranties/detail.html'
    context_object_name = 'warranty'
    
    def get_queryset(self):
        return Warranty.objects.select_related(
            'device', 'provider'
        )


class WarrantyUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update view for warranties."""
    model = Warranty
    form_class = WarrantyForm
    template_name = 'devices/warranties/edit.html'
    permission_required = 'devices.change_warranty'
    
    def get_success_url(self):
        return reverse('devices:warranty_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Warranty updated successfully!')
        return super().form_valid(form)


class WarrantyDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete view for warranties."""
    model = Warranty
    template_name = 'devices/warranties/delete.html'
    permission_required = 'devices.delete_warranty'
    success_url = reverse_lazy('devices:warranty_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Warranty deleted successfully!')
        return super().delete(request, *args, **kwargs)


class ExpiringWarrantiesView(LoginRequiredMixin, ListView):
    """View for warranties expiring soon."""
    model = Warranty
    template_name = 'devices/warranties/expiring.html'
    context_object_name = 'warranties'
    
    def get_queryset(self):
        thirty_days_later = date.today() + timedelta(days=30)
        return Warranty.objects.filter(
            end_date__range=[date.today(), thirty_days_later],
            is_active=True
        ).select_related('device', 'provider').order_by('end_date')


class ExpiredWarrantiesView(LoginRequiredMixin, ListView):
    """View for expired warranties."""
    model = Warranty
    template_name = 'devices/warranties/expired.html'
    context_object_name = 'warranties'
    
    def get_queryset(self):
        return Warranty.objects.filter(
            end_date__lt=date.today(),
            is_active=True
        ).select_related('device', 'provider').order_by('-end_date')


# QR Code Management Views
class QRCodeListView(LoginRequiredMixin, ListView):
    """List view for QR codes."""
    model = QRCode
    template_name = 'devices/qr/list.html'
    context_object_name = 'qr_codes'
    paginate_by = 25
    ordering = ['-created_at']
    
    def get_queryset(self):
        return QRCode.objects.select_related('device').filter(is_active=True)


class GenerateQRCodeView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for generating QR codes for devices."""
    permission_required = 'devices.add_qrcode'
    
    def post(self, request, pk):
        """Generate QR code for a specific device."""
        device = get_object_or_404(Device, pk=pk, is_active=True)
        
        try:
            # Use centralized QR generation
            from pims.utils.qr_code import create_device_qr_code
            
            qr_code_obj = create_device_qr_code(device, request)
            
            if qr_code_obj:
                messages.success(
                    request, 
                    f'QR code generated successfully for device "{device.device_id}"!'
                )
                return redirect('devices:qr_view', pk=qr_code_obj.pk)
            else:
                messages.error(request, 'Error generating QR code. Please try again.')
                return redirect('devices:detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error generating QR code: {str(e)}')
            return redirect('devices:detail', pk=pk)


class ViewQRCodeView(LoginRequiredMixin, DetailView):
    """View for displaying QR codes."""
    model = QRCode
    template_name = 'devices/qr/view.html'
    context_object_name = 'qr_code'


class DownloadQRCodeView(LoginRequiredMixin, DetailView):
    """View for downloading QR codes."""
    model = QRCode
    
    def get(self, request, *args, **kwargs):
        qr_code = self.get_object()
        
        if not qr_code.qr_code:
            raise Http404("QR code file not found")
        
        response = HttpResponse(qr_code.qr_code.read(), content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="device_{qr_code.device.device_id}_qr.png"'
        return response


class BulkGenerateQRView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """View for bulk generating QR codes for devices."""
    template_name = 'devices/qr/bulk_generate.html'
    permission_required = 'devices.add_qrcode'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all active devices
        all_devices = Device.objects.filter(is_active=True).select_related(
            'subcategory__category'
        )
        
        # Get devices without QR codes (devices that don't have any active QR codes)
        devices_without_qr = all_devices.exclude(
            qr_codes__is_active=True
        ).distinct()
        
        # Get devices with QR codes
        devices_with_qr = all_devices.filter(
            qr_codes__is_active=True
        ).distinct()
        
        # Calculate statistics
        total_devices = all_devices.count()
        devices_with_qr_count = devices_with_qr.count()
        devices_without_qr_count = devices_without_qr.count()
        
        # Get categories for filtering
        categories = DeviceCategory.objects.filter(
            is_active=True,
            subcategories__devices__is_active=True
        ).distinct().order_by('name')
        
        # Add context data
        context.update({
            'total_devices': total_devices,
            'devices_with_qr': devices_with_qr_count,
            'devices_without_qr': devices_without_qr_count,
            'devices_without_qr_list': devices_without_qr,
            'categories': categories,
            'page_title': 'Bulk Generate QR Codes',
            'breadcrumb_title': 'Bulk Generate QR Codes',
        })
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Process bulk QR code generation using centralized functions."""
        selected_devices = request.POST.getlist('devices')
        regenerate_existing = request.POST.get('regenerate_existing') == 'on'
        auto_download = request.POST.get('auto_download') == 'on'
        
        if not selected_devices:
            messages.error(request, 'Please select at least one device for QR code generation.')
            return self.get(request, *args, **kwargs)
        
        # Get selected devices
        devices = Device.objects.filter(
            id__in=selected_devices, 
            is_active=True
        )
        
        # Use centralized bulk generation
        from pims.utils.qr_code import bulk_generate_device_qr_codes
        
        results = bulk_generate_device_qr_codes(
            devices, 
            request, 
            regenerate_existing=regenerate_existing
        )
        
        # Prepare success message
        message_parts = []
        if results['generated'] > 0:
            message_parts.append(f'Generated {results["generated"]} new QR codes')
        if results['updated'] > 0:
            message_parts.append(f'Updated {results["updated"]} existing QR codes')
        if results['skipped'] > 0:
            message_parts.append(f'Skipped {results["skipped"]} existing QR codes')
        if results['errors'] > 0:
            message_parts.append(f'{results["errors"]} errors occurred')
        
        if message_parts:
            if results['errors'] == 0:
                messages.success(request, '. '.join(message_parts) + '!')
            else:
                messages.warning(request, '. '.join(message_parts) + '.')
                # Show specific error devices
                if results['error_devices']:
                    for error in results['error_devices'][:5]:  # Show first 5 errors
                        messages.error(request, f'Error: {error}')
        
        # Handle auto-download
        if auto_download and (results['generated'] > 0 or results['updated'] > 0):
            return redirect('devices:qr_bulk_download')
        
        return redirect('devices:qr_list')


class BulkDownloadQRView(LoginRequiredMixin, TemplateView):
    """View for bulk downloading QR codes."""
    template_name = 'devices/qr/bulk_download.html'
    
    def post(self, request, *args, **kwargs):
        selected_qr_codes = request.POST.getlist('qr_codes')
        
        if not selected_qr_codes:
            messages.error(request, 'Please select at least one QR code.')
            return self.get(request, *args, **kwargs)
        
        # Create ZIP file
        response = HttpResponse(content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="qr_codes_{timezone.now().strftime("%Y%m%d")}.zip"'
        
        with zipfile.ZipFile(response, 'w') as zip_file:
            for qr_id in selected_qr_codes:
                try:
                    qr_code = QRCode.objects.get(id=qr_id, is_active=True)
                    if qr_code.qr_code:
                        filename = f'device_{qr_code.device.device_id}_qr.png'
                        zip_file.writestr(filename, qr_code.qr_code.read())
                except QRCode.DoesNotExist:
                    continue
        
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['qr_codes'] = QRCode.objects.filter(is_active=True).select_related('device')
        return context


# Bulk Operations
class DeviceBulkUpdateView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """View for bulk updating devices."""
    template_name = 'devices/bulk_update.html'
    permission_required = 'devices.change_device'
    
    def post(self, request, *args, **kwargs):
        form = DeviceBulkUpdateForm(request.POST)
        selected_devices = request.POST.getlist('devices')
        
        if not selected_devices:
            messages.error(request, 'Please select at least one device.')
            return self.get(request, *args, **kwargs)
        
        if not form.is_valid():
            return self.get(request, *args, **kwargs)
        
        action = form.cleaned_data['action']
        updated_count = 0
        
        devices = Device.objects.filter(id__in=selected_devices)
        
        with transaction.atomic():
            if action == 'update_status':
                new_status = form.cleaned_data['new_status']
                updated_count = devices.update(status=new_status)
                
            elif action == 'update_condition':
                new_condition = form.cleaned_data['new_condition']
                updated_count = devices.update(condition=new_condition)
                
            elif action == 'update_location':
                new_location = form.cleaned_data['new_location']
                updated_count = devices.update(current_location=new_location)
                
            elif action == 'update_assignability':
                new_assignability = form.cleaned_data['new_assignability']
                updated_count = devices.update(is_assignable=new_assignability)
        
        messages.success(request, f'Updated {updated_count} devices successfully!')
        return redirect('devices:list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = DeviceBulkUpdateForm()
        context['devices'] = Device.objects.filter(is_active=True).select_related(
            'subcategory__category'
        )
        return context


class DeviceBulkExportView(LoginRequiredMixin, TemplateView):
    """View for bulk exporting devices."""
    template_name = 'devices/bulk_export.html'
    
    def post(self, request, *args, **kwargs):
        selected_devices = request.POST.getlist('devices')
        export_format = request.POST.get('format', 'csv')
        
        if not selected_devices:
            messages.error(request, 'Please select at least one device.')
            return self.get(request, *args, **kwargs)
        
        devices = Device.objects.filter(id__in=selected_devices).select_related(
            'subcategory__category', 'vendor', 'current_location'
        )
        
        if export_format == 'csv':
            return self.export_csv(devices)
        elif export_format == 'excel':
            return self.export_excel(devices)
        elif export_format == 'pdf':
            return self.export_pdf(devices)
    
    def export_csv(self, devices):
        """Export devices to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="devices_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Device ID', 'Category', 'Subcategory', 'Brand', 'Model',
            'Serial Number', 'Status', 'Condition', 'Purchase Date',
            'Purchase Price', 'Vendor', 'Location', 'Specifications'
        ])
        
        for device in devices:
            writer.writerow([
                device.device_id,
                device.subcategory.category.name,
                device.subcategory.name,
                device.brand,
                device.model,
                device.serial_number,
                device.get_status_display(),
                device.get_condition_display(),
                device.purchase_date,
                device.purchase_price,
                device.vendor.name,
                device.current_location.name if device.current_location else '',
                json.dumps(device.specifications) if device.specifications else ''
            ])
        
        return response
    
    def export_excel(self, devices):
        """Export devices to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Devices"
            
            # Headers
            headers = [
                'Device ID', 'Category', 'Subcategory', 'Brand', 'Model',
                'Serial Number', 'Status', 'Condition', 'Purchase Date',
                'Purchase Price', 'Vendor', 'Location', 'Specifications'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            for row, device in enumerate(devices, 2):
                ws.cell(row=row, column=1, value=device.device_id)
                ws.cell(row=row, column=2, value=device.subcategory.category.name)
                ws.cell(row=row, column=3, value=device.subcategory.name)
                ws.cell(row=row, column=4, value=device.brand)
                ws.cell(row=row, column=5, value=device.model)
                ws.cell(row=row, column=6, value=device.serial_number)
                ws.cell(row=row, column=7, value=device.get_status_display())
                ws.cell(row=row, column=8, value=device.get_condition_display())
                ws.cell(row=row, column=9, value=device.purchase_date)
                ws.cell(row=row, column=10, value=float(device.purchase_price))
                ws.cell(row=row, column=11, value=device.vendor.name)
                ws.cell(row=row, column=12, value=device.current_location.name if device.current_location else '')
                ws.cell(row=row, column=13, value=json.dumps(device.specifications) if device.specifications else '')
            
            # Save to response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="devices_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            messages.error(self.request, 'Excel export is not available. openpyxl package required.')
            return redirect('devices:list')
    
    def export_pdf(self, devices):
        """Export devices to PDF."""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="devices_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("Device Inventory Report", styles['Title'])
        elements.append(title)
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        # Table data
        data = [['Device ID', 'Brand', 'Model', 'Status', 'Condition', 'Purchase Date']]
        
        for device in devices:
            data.append([
                device.device_id,
                device.brand,
                device.model,
                device.get_status_display(),
                device.get_condition_display(),
                device.purchase_date.strftime('%Y-%m-%d')
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['devices'] = Device.objects.filter(is_active=True).select_related(
            'subcategory__category'
        )
        return context


# Report Views
class DeviceReportsView(LoginRequiredMixin, TemplateView):
    """Main reports dashboard."""
    template_name = 'devices/reports/index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Summary statistics
        total_devices = Device.objects.filter(is_active=True).count()
        total_value = Device.objects.filter(is_active=True).aggregate(
            total=Sum('purchase_price')
        )['total'] or 0
        
        # Status breakdown
        status_stats = Device.objects.filter(is_active=True).values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        # Category breakdown
        category_stats = DeviceCategory.objects.annotate(
            device_count=Count('subcategories__devices'),
            total_value=Sum('subcategories__devices__purchase_price')
        ).filter(device_count__gt=0)
        
        context.update({
            'total_devices': total_devices,
            'total_value': total_value,
            'status_stats': status_stats,
            'category_stats': category_stats,
        })
        
        return context


class InventoryReportView(LoginRequiredMixin, TemplateView):
    """Inventory summary report."""
    template_name = 'devices/reports/inventory.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Inventory by category
        inventory_by_category = DeviceCategory.objects.annotate(
            available_count=Count('subcategories__devices', filter=Q(subcategories__devices__status='AVAILABLE')),
            assigned_count=Count('subcategories__devices', filter=Q(subcategories__devices__status='ASSIGNED')),
            maintenance_count=Count('subcategories__devices', filter=Q(subcategories__devices__status='MAINTENANCE')),
            total_count=Count('subcategories__devices'),
            total_value=Sum('subcategories__devices__purchase_price')
        ).filter(total_count__gt=0)
        
        context['inventory_by_category'] = inventory_by_category
        return context


class CategoryReportView(LoginRequiredMixin, DetailView):
    """Category-specific report."""
    model = DeviceCategory
    template_name = 'devices/reports/category.html'
    context_object_name = 'category'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = self.object
        
        # Device breakdown by subcategory
        subcategory_stats = category.subcategories.annotate(
            device_count=Count('devices'),
            available_count=Count('devices', filter=Q(devices__status='AVAILABLE')),
            assigned_count=Count('devices', filter=Q(devices__status='ASSIGNED')),
            total_value=Sum('devices__purchase_price')
        ).filter(device_count__gt=0)
        
        context['subcategory_stats'] = subcategory_stats
        return context

class LocationReportView(LoginRequiredMixin, TemplateView):
    """
    Device report by location for Bangladesh Parliament Secretariat.
    Shows device distribution across different parliamentary buildings and offices.
    """
    template_name = 'devices/reports/location_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get device counts by location
        location_data = (
            Device.objects.filter(is_active=True)
            .values('current_location__name', 'current_location__building__name')
            .annotate(
                total_devices=Count('id'),
                available_devices=Count(Case(When(status='AVAILABLE', then=1))),
                assigned_devices=Count(Case(When(status='ASSIGNED', then=1))),
                maintenance_devices=Count(Case(When(status='MAINTENANCE', then=1))),
                total_value=Sum('purchase_price', output_field=DecimalField())
            )
            .order_by('current_location__building__name', 'current_location__name')
        )
        
        # Summary statistics
        summary = {
            'total_locations': location_data.count(),
            'total_devices': sum(loc['total_devices'] for loc in location_data),
            'total_value': sum(loc['total_value'] or 0 for loc in location_data),
            'avg_devices_per_location': round(
                sum(loc['total_devices'] for loc in location_data) / max(location_data.count(), 1), 2
            )
        }
        
        # Group by building for better visualization
        buildings = {}
        for loc in location_data:
            building = loc['current_location__building__name'] or 'Unassigned'
            if building not in buildings:
                buildings[building] = {
                    'name': building,
                    'locations': [],
                    'total_devices': 0,
                    'total_value': 0
                }
            buildings[building]['locations'].append(loc)
            buildings[building]['total_devices'] += loc['total_devices']
            buildings[building]['total_value'] += loc['total_value'] or 0
        
        context.update({
            'location_data': location_data,
            'buildings': buildings.values(),
            'summary': summary,
            'page_title': 'Devices by Location Report',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context

class VendorReportView(LoginRequiredMixin, TemplateView):
    """
    Device report by vendor showing procurement patterns and vendor performance.
    """
    template_name = 'devices/reports/vendor_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get device counts and values by vendor
        vendor_data = (
            Device.objects.filter(is_active=True)
            .values('vendor__name', 'vendor__contact_person', 'vendor__phone')
            .annotate(
                total_devices=Count('id'),
                total_value=Sum('purchase_price', output_field=DecimalField()),
                avg_device_price=Avg('purchase_price'),
                available_devices=Count(Case(When(status='AVAILABLE', then=1))),
                assigned_devices=Count(Case(When(status='ASSIGNED', then=1))),
                maintenance_devices=Count(Case(When(status='MAINTENANCE', then=1))),
                latest_purchase=Max('purchase_date'),
                device_categories=Count('subcategory__category', distinct=True)
            )
            .order_by('-total_value')
        )
        
        # Summary statistics
        summary = {
            'total_vendors': vendor_data.count(),
            'total_procurement_value': sum(v['total_value'] or 0 for v in vendor_data),
            'avg_order_value': round(
                sum(v['total_value'] or 0 for v in vendor_data) / max(vendor_data.count(), 1), 2
            ),
            'top_vendor': vendor_data.first() if vendor_data else None
        }
        
        # Vendor performance metrics
        for vendor in vendor_data:
            if vendor['total_devices'] > 0:
                vendor['maintenance_rate'] = round(
                    (vendor['maintenance_devices'] / vendor['total_devices']) * 100, 1
                )
            else:
                vendor['maintenance_rate'] = 0
        
        context.update({
            'vendor_data': vendor_data,
            'summary': summary,
            'page_title': 'Devices by Vendor Report',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context


class DepreciationReportView(LoginRequiredMixin, TemplateView):
    """
    Depreciation report showing asset depreciation and current values.
    """
    template_name = 'devices/reports/depreciation_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate depreciation (assuming 5-year straight line for IT equipment)
        devices = Device.objects.filter(is_active=True).select_related(
            'subcategory__category', 'vendor'
        )
        
        depreciation_data = []
        total_original_value = 0
        total_current_value = 0
        
        for device in devices:
            if device.purchase_date and device.purchase_price:
                age_years = (datetime.now().date() - device.purchase_date).days / 365.25
                depreciation_years = min(age_years, 5)  # Max 5 years depreciation
                annual_depreciation = device.purchase_price / 5
                accumulated_depreciation = annual_depreciation * depreciation_years
                current_value = max(device.purchase_price - accumulated_depreciation, 0)
                
                depreciation_data.append({
                    'device': device,
                    'age_years': round(age_years, 1),
                    'original_value': device.purchase_price,
                    'annual_depreciation': annual_depreciation,
                    'accumulated_depreciation': accumulated_depreciation,
                    'current_value': current_value
                })
                
                total_original_value += device.purchase_price
                total_current_value += current_value
        
        # Summary
        summary = {
            'total_devices': len(depreciation_data),
            'total_original_value': total_original_value,
            'total_current_value': total_current_value,
        }
        
        context.update({
            'depreciation_data': depreciation_data,
            'summary': summary,
            'page_title': 'Asset Depreciation Report',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context


class WarrantyReportView(LoginRequiredMixin, TemplateView):
    """
    Warranty summary report showing warranty status across all devices.
    """
    template_name = 'devices/reports/warranty_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get warranty data
        today = datetime.now().date()
        thirty_days = today + timedelta(days=30)
        ninety_days = today + timedelta(days=90)
        
        warranties = Warranty.objects.filter(
            device__is_active=True
        ).select_related('device__subcategory__category', 'device__vendor')
        
        warranty_summary = {
            'total_warranties': warranties.count(),
            'active_warranties': warranties.filter(end_date__gte=today).count(),
            'expired_warranties': warranties.filter(end_date__lt=today).count(),
            'expiring_30_days': warranties.filter(
                end_date__gte=today, end_date__lte=thirty_days
            ).count(),
            'expiring_90_days': warranties.filter(
                end_date__gte=today, end_date__lte=ninety_days
            ).count()
        }
        
        # Warranty status breakdown
        warranty_data = []
        for warranty in warranties:
            days_remaining = (warranty.end_date - today).days
            
            if days_remaining < 0:
                status = 'expired'
                status_class = 'danger'
            elif days_remaining <= 30:
                status = 'expiring_soon'
                status_class = 'warning'
            elif days_remaining <= 90:
                status = 'expiring_90_days'
                status_class = 'info'
            else:
                status = 'active'
                status_class = 'success'
            
            warranty_data.append({
                'warranty': warranty,
                'days_remaining': days_remaining,
                'status': status,
                'status_class': status_class
            })
        
        # Group by vendor for analysis
        vendor_warranty_stats = (
            warranties.values('device__vendor__name')
            .annotate(
                total=Count('id'),
                active=Count(Case(When(end_date__gte=today, then=1))),
                expired=Count(Case(When(end_date__lt=today, then=1)))
            )
            .order_by('-total')
        )
        
        context.update({
            'warranty_summary': warranty_summary,
            'warranty_data': warranty_data,
            'vendor_warranty_stats': vendor_warranty_stats,
            'page_title': 'Warranty Summary Report',
            'organization': 'Bangladesh Parliament Secretariat',
            'location': 'Dhaka, Bangladesh'
        })
        
        return context

# Export Views
class ExportDevicesExcelView(LoginRequiredMixin, View):
    """
    Export devices to Excel format with comprehensive data.
    """
    
    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl package.')
            return redirect('devices:list')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devices Inventory"
        
        # Headers
        headers = [
            'Device ID', 'Category', 'Subcategory', 'Brand', 'Model',
            'Serial Number', 'Status', 'Condition', 'Purchase Date',
            'Purchase Price (BDT)', 'Vendor', 'Current Location', 'Building',
            'Warranty End Date', 'Specifications'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get devices data
        devices = Device.objects.filter(is_active=True).select_related(
            'subcategory__category', 'vendor', 'current_location__building'
        ).prefetch_related('warranties')
        
        # Add data rows
        for row, device in enumerate(devices, 2):
            # Get latest warranty
            latest_warranty = device.warranties.filter(is_active=True).order_by('-end_date').first()
            
            row_data = [
                device.device_id,
                device.subcategory.category.name,
                device.subcategory.name,
                device.brand,
                device.model,
                device.serial_number,
                device.get_status_display(),
                device.get_condition_display(),
                device.purchase_date.strftime('%Y-%m-%d') if device.purchase_date else '',
                device.purchase_price if device.purchase_price else '',
                device.vendor.name if device.vendor else '',
                device.current_location.name if device.current_location else '',
                device.current_location.building.name if device.current_location and device.current_location.building else '',
                latest_warranty.end_date.strftime('%Y-%m-%d') if latest_warranty else '',
                json.dumps(device.specifications) if device.specifications else ''
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Summary data
        summary_data = [
            ['Report Title', 'PIMS Device Inventory Report'],
            ['Organization', 'Bangladesh Parliament Secretariat'],
            ['Location', 'Dhaka, Bangladesh'],
            ['Generated Date', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Devices', devices.count()],
            ['', ''],
            ['Status Summary', ''],
            ['Available', devices.filter(status='AVAILABLE').count()],
            ['Assigned', devices.filter(status='ASSIGNED').count()],
            ['Maintenance', devices.filter(status='MAINTENANCE').count()],
            ['Retired', devices.filter(status='RETIRED').count()],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row, column=2, value=value)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'devices_inventory_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class ExportDevicesPDFView(LoginRequiredMixin, View):
    """
    Export devices to PDF format with parliamentary formatting.
    """
    
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        response = HttpResponse(content_type='application/pdf')
        filename = f'devices_inventory_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            alignment=1  # Center
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#6b7280'),
            spaceAfter=20,
            alignment=1  # Center
        )
        
        # Header
        elements.append(Paragraph("Bangladesh Parliament Secretariat", title_style))
        elements.append(Paragraph("PIMS - Device Inventory Report", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Location: Dhaka, Bangladesh", subtitle_style))
        elements.append(Spacer(1, 12))
        
        # Get devices data
        devices = Device.objects.filter(is_active=True).select_related(
            'subcategory__category', 'vendor', 'current_location'
        )[:100]  # Limit for PDF
        
        # Create table data
        table_data = [
            ['Device ID', 'Category', 'Brand/Model', 'Status', 'Location', 'Purchase Date', 'Price (BDT)']
        ]
        
        for device in devices:
            table_data.append([
                device.device_id,
                device.subcategory.category.name,
                f"{device.brand} {device.model}",
                device.get_status_display(),
                device.current_location.name if device.current_location else 'Unassigned',
                device.purchase_date.strftime('%Y-%m-%d') if device.purchase_date else 'N/A',
                f"{device.purchase_price:,.0f}" if device.purchase_price else 'N/A'
            ])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Add summary statistics
        summary_data = [
            ['Summary Statistics', ''],
            ['Total Devices', devices.count()],
            ['Available', devices.filter(status='AVAILABLE').count()],
            ['Assigned', devices.filter(status='ASSIGNED').count()],
            ['In Maintenance', devices.filter(status='MAINTENANCE').count()],
        ]
        
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Summary", styles['Heading2']))
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ]))
        
        elements.append(summary_table)
        
        # Build PDF
        doc.build(elements)
        return response


class ExportDevicesCSVView(LoginRequiredMixin, View):
    """
    Export devices to CSV format for data analysis.
    """
    
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        filename = f'devices_inventory_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'Device ID', 'Category', 'Subcategory', 'Brand', 'Model',
            'Serial Number', 'Status', 'Condition', 'Purchase Date',
            'Purchase Price (BDT)', 'Vendor', 'Current Location', 'Building',
            'Warranty End Date', 'Created Date', 'Last Updated', 'Specifications'
        ])
        
        # Get devices data
        devices = Device.objects.filter(is_active=True).select_related(
            'subcategory__category', 'vendor', 'current_location__building'
        ).prefetch_related('warranties')
        
        # Write data rows
        for device in devices:
            latest_warranty = device.warranties.filter(is_active=True).order_by('-end_date').first()
            
            writer.writerow([
                device.device_id,
                device.subcategory.category.name,
                device.subcategory.name,
                device.brand,
                device.model,
                device.serial_number,
                device.get_status_display(),
                device.get_condition_display(),
                device.purchase_date.strftime('%Y-%m-%d') if device.purchase_date else '',
                device.purchase_price if device.purchase_price else '',
                device.vendor.name if device.vendor else '',
                device.current_location.name if device.current_location else '',
                device.current_location.building.name if device.current_location and device.current_location.building else '',
                latest_warranty.end_date.strftime('%Y-%m-%d') if latest_warranty else '',
                device.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                device.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                json.dumps(device.specifications) if device.specifications else ''
            ])
        
        return response

# AJAX Endpoints
@login_required
@require_http_methods(["GET"])
def get_subcategories_ajax(request, category_id):
    """AJAX endpoint to get subcategories for a category."""
    try:
        subcategories = DeviceSubcategory.objects.filter(
            category_id=category_id,
            is_active=True
        ).values('id', 'name').order_by('name')
        
        return JsonResponse({
            'subcategories': list(subcategories)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_parent_devices_ajax(request, subcategory_id):
    """AJAX endpoint to get possible parent devices."""
    try:
        parent_devices = Device.objects.filter(
            device_type='COMPLETE',
            is_active=True
        ).values('id', 'device_id', 'brand', 'model').order_by('device_id')
        
        return JsonResponse({
            'parent_devices': list(parent_devices)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_device_info_ajax(request, device_id):
    """AJAX endpoint to get device information."""
    try:
        device = Device.objects.select_related(
            'subcategory__category', 'vendor'
        ).get(id=device_id)
        
        return JsonResponse({
            'device': {
                'id': device.id,
                'device_id': device.device_id,
                'brand': device.brand,
                'model': device.model,
                'serial_number': device.serial_number,
                'status': device.status,
                'condition': device.condition,
                'category': device.subcategory.category.name,
                'subcategory': device.subcategory.name,
                'vendor': device.vendor.name,
                'specifications': device.specifications
            }
        })
    except Device.DoesNotExist:
        return JsonResponse({'error': 'Device not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_specifications_template_ajax(request, subcategory_id):
    """AJAX endpoint to get specification template for subcategory."""
    try:
        subcategory = DeviceSubcategory.objects.get(id=subcategory_id)
        
        # Predefined templates based on subcategory
        templates = {
            'LAP': {  # Laptop
                'cpu': '',
                'ram': '',
                'storage': '',
                'screen_size': '',
                'graphics': ''
            },
            'DESK': {  # Desktop
                'cpu': '',
                'ram': '',
                'storage': '',
                'motherboard': '',
                'graphics': ''
            },
            'RAM': {  # RAM
                'size': '',
                'type': '',
                'speed': '',
                'form_factor': ''
            },
            'SSD': {  # SSD
                'capacity': '',
                'interface': '',
                'form_factor': '',
                'read_speed': '',
                'write_speed': ''
            },
            'RTR': {  # Router
                'ports': '',
                'wifi_standard': '',
                'throughput': '',
                'antenna': ''
            }
        }
        
        template = templates.get(subcategory.code, {})
        
        return JsonResponse({
            'template': template,
            'subcategory': subcategory.name
        })
    except DeviceSubcategory.DoesNotExist:
        return JsonResponse({'error': 'Subcategory not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_specifications_examples_ajax(request):
    """AJAX endpoint to get specification examples."""
    examples = {
        'Desktop/Laptop': {
            'cpu': 'Intel i5-12400',
            'ram': '16GB DDR4',
            'storage': '512GB SSD',
            'graphics': 'Intel UHD 770'
        },
        'Router': {
            'ports': '24 Gigabit Ethernet',
            'wifi': '802.11ac',
            'throughput': '100 Mbps'
        },
        'RAM Component': {
            'size': '8GB',
            'type': 'DDR4',
            'speed': '3200MHz',
            'form_factor': 'DIMM'
        },
        'Storage': {
            'capacity': '1TB',
            'type': 'NVMe SSD',
            'interface': 'M.2',
            'read_speed': '3500 MB/s'
        }
    }
    
    return JsonResponse({'examples': examples})


@login_required
@require_http_methods(["POST"])
def validate_device_id_ajax(request):
    """AJAX endpoint to validate device ID uniqueness."""
    device_id = request.POST.get('device_id', '').upper().strip()
    current_id = request.POST.get('current_id')  # For editing
    
    if not device_id:
        return JsonResponse({'valid': True})
    
    queryset = Device.objects.filter(device_id=device_id)
    if current_id:
        queryset = queryset.exclude(id=current_id)
    
    exists = queryset.exists()
    
    return JsonResponse({
        'valid': not exists,
        'message': f'Device ID "{device_id}" already exists.' if exists else 'Device ID is available.'
    })


@login_required
@require_http_methods(["POST"])
def validate_serial_number_ajax(request):
    """AJAX endpoint to validate serial number uniqueness."""
    serial_number = request.POST.get('serial_number', '').strip()
    current_id = request.POST.get('current_id')  # For editing
    
    if not serial_number:
        return JsonResponse({'valid': True})
    
    queryset = Device.objects.filter(serial_number=serial_number)
    if current_id:
        queryset = queryset.exclude(id=current_id)
    
    exists = queryset.exists()
    
    return JsonResponse({
        'valid': not exists,
        'message': f'Serial number "{serial_number}" already exists.' if exists else 'Serial number is available.'
    })


@login_required
@require_http_methods(["GET"])
def get_dashboard_stats_ajax(request):
    """AJAX endpoint for dashboard statistics."""
    try:
        stats = {
            'total_devices': Device.objects.filter(is_active=True).count(),
            'available_devices': Device.objects.filter(status='AVAILABLE').count(),
            'assigned_devices': Device.objects.filter(status='ASSIGNED').count(),
            'maintenance_devices': Device.objects.filter(status='MAINTENANCE').count(),
            'total_value': float(Device.objects.filter(is_active=True).aggregate(
                total=Sum('purchase_price')
            )['total'] or 0)
        }
        
        return JsonResponse({'stats': stats})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_category_stats_ajax(request, category_id):
    """AJAX endpoint for category statistics."""
    try:
        category = DeviceCategory.objects.get(id=category_id)
        
        stats = {
            'total_devices': Device.objects.filter(subcategory__category=category).count(),
            'available_devices': Device.objects.filter(
                subcategory__category=category,
                status='AVAILABLE'
            ).count(),
            'assigned_devices': Device.objects.filter(
                subcategory__category=category,
                status='ASSIGNED'
            ).count(),
            'maintenance_devices': Device.objects.filter(
                subcategory__category=category,
                status='MAINTENANCE'
            ).count()
        }
        
        return JsonResponse({'stats': stats})
    except DeviceCategory.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)