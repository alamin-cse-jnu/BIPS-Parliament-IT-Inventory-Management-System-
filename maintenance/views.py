"""
Views for Maintenance app in PIMS (Parliament IT Inventory Management System)
Bangladesh Parliament Secretariat

This module defines comprehensive views for maintenance management including:
- CRUD operations for maintenance records
- Status management and workflow transitions
- Scheduling and calendar functionality
- Reports and analytics
- Dashboard and statistics
- AJAX endpoints for dynamic functionality
- Export capabilities (PDF, Excel, CSV)

Location: maintenance/views.py
Author: PIMS Development Team
"""

# Standard library imports
import csv
import json
import calendar
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
import zipfile
import sys
# Django core imports
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import (
    Q, Count, Sum, Avg, Max, Min, Case, When, IntegerField,
    F, Value, CharField, DateField, DecimalField
)
from django.http import (
    HttpResponse, HttpResponseRedirect, JsonResponse, 
    Http404, HttpResponseBadRequest, FileResponse
)
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, 
    TemplateView, FormView
)

# Third-party imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image
)
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing

# Local app imports
from .models import Maintenance
from .forms import (
    MaintenanceCreateForm, MaintenanceUpdateForm, MaintenanceScheduleForm,
    MaintenanceQuickActionForm, MaintenanceFilterForm, MaintenanceApprovalForm,
    MaintenanceSearchForm, MaintenanceReportForm, MaintenanceBulkUpdateForm,
    create_maintenance_schedule, apply_bulk_maintenance_update
)
from devices.models import Device, DeviceCategory
from vendors.models import Vendor

# Additional imports for utility functions
import mimetypes
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.utils.html import strip_tags
from django.core.cache import cache
import tempfile
import os
from typing import Dict, List, Optional, Tuple

User = get_user_model()


# ============================================================================
# Core Maintenance CRUD Views
# ============================================================================

class MaintenanceListView(LoginRequiredMixin, ListView):
    """
    Display list of all maintenance records with advanced filtering and search.
    Location: maintenance/list.html
    """
    model = Maintenance
    template_name = 'maintenance/list.html'
    context_object_name = 'maintenance_records'
    paginate_by = 25
    
    def get_queryset(self):
        """Filter maintenance records based on search parameters."""
        queryset = Maintenance.objects.select_related(
            'device', 'vendor', 'created_by', 'approved_by'
        ).prefetch_related(
            'device__subcategory__category'
        )
        
        # Apply filters from GET parameters
        filter_form = MaintenanceFilterForm(self.request.GET)
        if filter_form.is_valid():
            if filter_form.cleaned_data.get('status'):
                queryset = queryset.filter(status=filter_form.cleaned_data['status'])
            if filter_form.cleaned_data.get('maintenance_type'):
                queryset = queryset.filter(maintenance_type=filter_form.cleaned_data['maintenance_type'])
            if filter_form.cleaned_data.get('priority'):
                queryset = queryset.filter(priority=filter_form.cleaned_data['priority'])
            if filter_form.cleaned_data.get('date_range'):
                date_range = filter_form.cleaned_data['date_range']
                today = timezone.now().date()
                
                if date_range == 'today':
                    queryset = queryset.filter(start_date=today)
                elif date_range == 'week':
                    week_start = today - timedelta(days=today.weekday())
                    queryset = queryset.filter(start_date__gte=week_start)
                elif date_range == 'month':
                    queryset = queryset.filter(
                        start_date__year=today.year,
                        start_date__month=today.month
                    )
                elif date_range == 'overdue':
                    queryset = queryset.filter(
                        expected_end_date__lt=today,
                        status__in=['SCHEDULED', 'IN_PROGRESS']
                    )
                elif date_range == 'due_soon':
                    soon_date = today + timedelta(days=7)
                    queryset = queryset.filter(
                        start_date__lte=soon_date,
                        status='SCHEDULED'
                    )
            
            if filter_form.cleaned_data.get('search'):
                search_term = filter_form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(maintenance_id__icontains=search_term) |
                    Q(title__icontains=search_term) |
                    Q(description__icontains=search_term) |
                    Q(device__device_id__icontains=search_term) |
                    Q(device__brand__icontains=search_term) |
                    Q(vendor__name__icontains=search_term)
                )
        
        return queryset.order_by('-start_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        """Add additional context data for the template."""
        context = super().get_context_data(**kwargs)
        
        # Add filter form
        context['filter_form'] = MaintenanceFilterForm(self.request.GET)
        
        # Add statistics
        all_maintenance = Maintenance.objects.all()
        context['total_maintenance'] = all_maintenance.count()
        context['in_progress_count'] = all_maintenance.filter(status='IN_PROGRESS').count()
        context['completed_count'] = all_maintenance.filter(status='COMPLETED').count()
        context['overdue_count'] = len([m for m in all_maintenance if m.is_overdue])
        
        # Add quick stats
        context['stats'] = {
            'scheduled': all_maintenance.filter(status='SCHEDULED').count(),
            'on_hold': all_maintenance.filter(status='ON_HOLD').count(),
            'cancelled': all_maintenance.filter(status='CANCELLED').count(),
            'failed': all_maintenance.filter(status='FAILED').count(),
        }
        
        return context


class MaintenanceDetailView(LoginRequiredMixin, DetailView):
    """
    Display detailed information about a maintenance record.
    Location: maintenance/detail.html
    """
    model = Maintenance
    template_name = 'maintenance/detail.html'
    context_object_name = 'maintenance'
    
    def get_object(self):
        """Get maintenance object with optimized queries."""
        return get_object_or_404(
            Maintenance.objects.select_related(
                'device', 'vendor', 'created_by', 'approved_by', 'updated_by'
            ).prefetch_related(
                'device__subcategory__category'
            ),
            pk=self.kwargs['pk']
        )
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        
        # Add related maintenance records for the same device
        context['related_maintenance'] = Maintenance.objects.filter(
            device=self.object.device
        ).exclude(pk=self.object.pk).order_by('-start_date')[:5]
        
        # Add device assignment history (if available)
        if hasattr(self.object.device, 'assignments'):
            context['device_assignments'] = self.object.device.assignments.filter(
                is_active=True
            ).select_related('assigned_to')[:3]
        
        # Check user permissions
        context['can_edit'] = self.request.user.has_perm('maintenance.change_maintenance')
        context['can_approve'] = self.request.user.has_perm('maintenance.approve_maintenance')
        context['can_view_costs'] = self.request.user.has_perm('maintenance.view_maintenance_costs')
        
        return context


class MaintenanceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Create new maintenance record.
    Location: maintenance/create.html
    """
    model = Maintenance
    form_class = MaintenanceCreateForm
    template_name = 'maintenance/create.html'
    permission_required = 'maintenance.add_maintenance'
    
    def get_form_kwargs(self):
        """Pass user to form for permission checks."""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """Set created_by field and handle additional processing."""
        form.instance.created_by = self.request.user
        
        # Auto-generate maintenance ID if not provided
        if not form.instance.maintenance_id:
            form.instance.maintenance_id = form.instance._generate_maintenance_id()
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f'Maintenance {self.object.maintenance_id} created successfully for device {self.object.device.device_id}.'
        )
        
        return response
    
    def get_success_url(self):
        """Redirect to detail view after creation."""
        return reverse('maintenance:detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Schedule New Maintenance'
        
        # Add device info if device_id is provided in GET parameters
        device_id = self.request.GET.get('device_id')
        if device_id:
            try:
                context['suggested_device'] = Device.objects.get(pk=device_id)
            except Device.DoesNotExist:
                pass
        
        return context


class MaintenanceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Update existing maintenance record.
    Location: maintenance/edit.html
    """
    model = Maintenance
    form_class = MaintenanceUpdateForm
    template_name = 'maintenance/edit.html'
    permission_required = 'maintenance.change_maintenance'
    
    def get_form_kwargs(self):
        """Pass user to form for permission checks."""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """Set updated_by field and handle status transitions."""
        form.instance.updated_by = self.request.user
        
        # Check if status changed
        old_status = Maintenance.objects.get(pk=self.object.pk).status
        new_status = form.instance.status
        
        if old_status != new_status:
            # Validate status transition
            if not self.object.can_transition_to(new_status):
                messages.error(
                    self.request,
                    f'Invalid status transition from {old_status} to {new_status}.'
                )
                return self.form_invalid(form)
            
            # Set appropriate timestamps for status changes
            if new_status == 'IN_PROGRESS' and not form.instance.actual_start_date:
                form.instance.actual_start_date = timezone.now()
            elif new_status == 'COMPLETED' and not form.instance.actual_end_date:
                form.instance.actual_end_date = timezone.now()
        
        response = super().form_valid(form)
        
        messages.success(
            self.request,
            f'Maintenance {self.object.maintenance_id} updated successfully.'
        )
        
        return response
    
    def get_success_url(self):
        """Redirect to detail view after update."""
        return reverse('maintenance:detail', kwargs={'pk': self.object.pk})


class MaintenanceDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Delete maintenance record.
    Location: maintenance/delete.html
    """
    model = Maintenance
    template_name = 'maintenance/delete.html'
    permission_required = 'maintenance.delete_maintenance'
    success_url = reverse_lazy('maintenance:list')
    
    def delete(self, request, *args, **kwargs):
        """Handle deletion with proper messaging."""
        maintenance = self.get_object()
        maintenance_id = maintenance.maintenance_id
        
        # Check if maintenance can be deleted
        if maintenance.status == 'IN_PROGRESS':
            messages.error(
                request,
                'Cannot delete maintenance that is currently in progress.'
            )
            return redirect('maintenance:detail', pk=maintenance.pk)
        
        response = super().delete(request, *args, **kwargs)
        
        messages.success(
            request,
            f'Maintenance {maintenance_id} deleted successfully.'
        )
        
        return response


# ============================================================================
# Status-based Maintenance Views
# ============================================================================

class ScheduledMaintenanceView(MaintenanceListView):
    """Display only scheduled maintenance records."""
    template_name = 'maintenance/scheduled.html'
    
    def get_queryset(self):
        """Filter for scheduled maintenance only."""
        return super().get_queryset().filter(status='SCHEDULED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Scheduled Maintenance'
        return context


class InProgressMaintenanceView(MaintenanceListView):
    """Display only in-progress maintenance records."""
    template_name = 'maintenance/in_progress.html'
    
    def get_queryset(self):
        """Filter for in-progress maintenance only."""
        return super().get_queryset().filter(status='IN_PROGRESS')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'In Progress Maintenance'
        return context


class CompletedMaintenanceView(MaintenanceListView):
    """Display only completed maintenance records."""
    template_name = 'maintenance/completed.html'
    
    def get_queryset(self):
        """Filter for completed maintenance only."""
        return super().get_queryset().filter(status='COMPLETED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Completed Maintenance'
        return context


class OverdueMaintenanceView(MaintenanceListView):
    """Display only overdue maintenance records."""
    template_name = 'maintenance/overdue.html'
    
    def get_queryset(self):
        """Filter for overdue maintenance only."""
        queryset = super().get_queryset().filter(
            expected_end_date__lt=timezone.now().date(),
            status__in=['SCHEDULED', 'IN_PROGRESS']
        )
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Overdue Maintenance'
        return context


class CancelledMaintenanceView(MaintenanceListView):
    """Display only cancelled maintenance records."""
    template_name = 'maintenance/cancelled.html'
    
    def get_queryset(self):
        """Filter for cancelled maintenance only."""
        return super().get_queryset().filter(status='CANCELLED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Cancelled Maintenance'
        return context


# ============================================================================
# Search and Filtering Views
# ============================================================================

class MaintenanceSearchView(LoginRequiredMixin, ListView):
    """
    Advanced search for maintenance records.
    Location: maintenance/search.html
    """
    model = Maintenance
    template_name = 'maintenance/search.html'
    context_object_name = 'maintenance_records'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter maintenance based on search form."""
        queryset = Maintenance.objects.select_related(
            'device', 'vendor', 'created_by'
        ).prefetch_related('device__subcategory__category')
        
        form = MaintenanceSearchForm(self.request.GET)
        if form.is_valid():
            # Apply search filters
            if form.cleaned_data.get('search_query'):
                query = form.cleaned_data['search_query']
                queryset = queryset.filter(
                    Q(maintenance_id__icontains=query) |
                    Q(title__icontains=query) |
                    Q(description__icontains=query) |
                    Q(device__device_id__icontains=query) |
                    Q(device__brand__icontains=query) |
                    Q(device__model__icontains=query) |
                    Q(vendor__name__icontains=query)
                )
            
            if form.cleaned_data.get('status'):
                queryset = queryset.filter(status__in=form.cleaned_data['status'])
            
            if form.cleaned_data.get('maintenance_type'):
                queryset = queryset.filter(maintenance_type__in=form.cleaned_data['maintenance_type'])
            
            if form.cleaned_data.get('priority'):
                queryset = queryset.filter(priority__in=form.cleaned_data['priority'])
            
            if form.cleaned_data.get('start_date_from'):
                queryset = queryset.filter(start_date__gte=form.cleaned_data['start_date_from'])
            
            if form.cleaned_data.get('start_date_to'):
                queryset = queryset.filter(start_date__lte=form.cleaned_data['start_date_to'])
            
            if form.cleaned_data.get('cost_min'):
                queryset = queryset.filter(actual_cost__gte=form.cleaned_data['cost_min'])
            
            if form.cleaned_data.get('cost_max'):
                queryset = queryset.filter(actual_cost__lte=form.cleaned_data['cost_max'])
            
            if form.cleaned_data.get('vendor'):
                queryset = queryset.filter(vendor=form.cleaned_data['vendor'])
            
            if form.cleaned_data.get('device_category'):
                queryset = queryset.filter(
                    device__subcategory__category_id=form.cleaned_data['device_category']
                )
            
            if form.cleaned_data.get('overdue_only'):
                today = timezone.now().date()
                queryset = queryset.filter(
                    expected_end_date__lt=today,
                    status__in=['SCHEDULED', 'IN_PROGRESS']
                )
            
            if form.cleaned_data.get('warranty_service'):
                queryset = queryset.filter(is_warranty_service=True)
        
        return queryset.order_by('-start_date')
    
    def get_context_data(self, **kwargs):
        """Add search form to context."""
        context = super().get_context_data(**kwargs)
        context['search_form'] = MaintenanceSearchForm(self.request.GET)
        context['page_title'] = 'Search Maintenance Records'
        return context


class MaintenanceFilterView(LoginRequiredMixin, View):
    """
    AJAX view for dynamic filtering.
    """
    def get(self, request):
        """Return filtered maintenance data as JSON."""
        form = MaintenanceFilterForm(request.GET)
        if form.is_valid():
            # Apply filters similar to MaintenanceListView
            queryset = Maintenance.objects.all()
            # ... apply filters ...
            
            # Return filtered data
            data = []
            for maintenance in queryset[:50]:  # Limit for performance
                data.append({
                    'id': maintenance.id,
                    'maintenance_id': maintenance.maintenance_id,
                    'device': str(maintenance.device),
                    'status': maintenance.get_status_display(),
                    'priority': maintenance.get_priority_display(),
                    'start_date': maintenance.start_date.isoformat(),
                    'url': reverse('maintenance:detail', kwargs={'pk': maintenance.pk})
                })
            
            return JsonResponse({'results': data})
        
        return JsonResponse({'error': 'Invalid form data'}, status=400)


# ============================================================================
# Bulk Operations Views
# ============================================================================

class MaintenanceBulkUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    Bulk update maintenance records.
    Location: maintenance/bulk_update.html
    """
    form_class = MaintenanceBulkUpdateForm
    template_name = 'maintenance/bulk_update.html'
    permission_required = 'maintenance.change_maintenance'
    
    def get_form_kwargs(self):
        """Pass user to form."""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """Apply bulk updates."""
        selected_ids = self.request.POST.getlist('selected_maintenance')
        if not selected_ids:
            messages.error(self.request, 'No maintenance records selected.')
            return self.form_invalid(form)
        
        maintenance_queryset = Maintenance.objects.filter(id__in=selected_ids)
        updated_count = apply_bulk_maintenance_update(
            form.cleaned_data, maintenance_queryset, self.request.user
        )
        
        messages.success(
            self.request,
            f'{updated_count} maintenance record(s) updated successfully.'
        )
        
        return redirect('maintenance:list')
    
    def get_context_data(self, **kwargs):
        """Add maintenance records to context."""
        context = super().get_context_data(**kwargs)
        
        # Get selected maintenance IDs from session or GET parameters
        selected_ids = self.request.GET.getlist('ids')
        if selected_ids:
            context['selected_maintenance'] = Maintenance.objects.filter(
                id__in=selected_ids
            ).select_related('device', 'vendor')
        
        return context


class MaintenanceBulkExportView(LoginRequiredMixin, View):
    """
    Export selected maintenance records.
    """
    def post(self, request):
        """Export maintenance records based on selection."""
        selected_ids = request.POST.getlist('selected_maintenance')
        export_format = request.POST.get('export_format', 'csv')
        
        if not selected_ids:
            messages.error(request, 'No maintenance records selected for export.')
            return redirect('maintenance:list')
        
        maintenance_queryset = Maintenance.objects.filter(
            id__in=selected_ids
        ).select_related('device', 'vendor', 'created_by')
        
        if export_format == 'csv':
            return self._export_csv(maintenance_queryset)
        elif export_format == 'excel':
            return self._export_excel(maintenance_queryset)
        elif export_format == 'pdf':
            return self._export_pdf(maintenance_queryset)
        
        messages.error(request, 'Invalid export format selected.')
        return redirect('maintenance:list')
    
    def _export_csv(self, queryset):
        """Export as CSV file."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="maintenance_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Maintenance ID', 'Device ID', 'Device Brand', 'Device Model',
            'Type', 'Priority', 'Status', 'Start Date', 'Expected End Date',
            'Actual End Date', 'Estimated Cost', 'Actual Cost', 'Vendor',
            'Technician', 'Result', 'Created Date'
        ])
        
        for maintenance in queryset:
            writer.writerow([
                maintenance.maintenance_id,
                maintenance.device.device_id if maintenance.device else '',
                maintenance.device.brand if maintenance.device else '',
                maintenance.device.model if maintenance.device else '',
                maintenance.get_maintenance_type_display(),
                maintenance.get_priority_display(),
                maintenance.get_status_display(),
                maintenance.start_date.strftime('%Y-%m-%d') if maintenance.start_date else '',
                maintenance.expected_end_date.strftime('%Y-%m-%d') if maintenance.expected_end_date else '',
                maintenance.actual_end_date.strftime('%Y-%m-%d %H:%M') if maintenance.actual_end_date else '',
                maintenance.estimated_cost,
                maintenance.actual_cost or '',
                maintenance.vendor.name if maintenance.vendor else '',
                maintenance.technician_name or '',
                maintenance.get_result_display() if maintenance.result else '',
                maintenance.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export as Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Maintenance Records"
        
        # Header row
        headers = [
            'Maintenance ID', 'Device ID', 'Device Brand', 'Device Model',
            'Type', 'Priority', 'Status', 'Start Date', 'Expected End Date',
            'Actual End Date', 'Estimated Cost (BDT)', 'Actual Cost (BDT)', 
            'Vendor', 'Technician', 'Result', 'Created Date'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row, maintenance in enumerate(queryset, 2):
            data = [
                maintenance.maintenance_id,
                maintenance.device.device_id if maintenance.device else '',
                maintenance.device.brand if maintenance.device else '',
                maintenance.device.model if maintenance.device else '',
                maintenance.get_maintenance_type_display(),
                maintenance.get_priority_display(),
                maintenance.get_status_display(),
                maintenance.start_date,
                maintenance.expected_end_date,
                maintenance.actual_end_date,
                float(maintenance.estimated_cost) if maintenance.estimated_cost else '',
                float(maintenance.actual_cost) if maintenance.actual_cost else '',
                maintenance.vendor.name if maintenance.vendor else '',
                maintenance.technician_name or '',
                maintenance.get_result_display() if maintenance.result else '',
                maintenance.created_at
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="maintenance_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def _export_pdf(self, queryset):
        """Export as PDF file."""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="maintenance_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph("Maintenance Records Export", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Export info
        export_info = Paragraph(
            f"Exported on: {timezone.now().strftime('%Y-%m-%d %H:%M')} | "
            f"Total Records: {queryset.count()}",
            styles['Normal']
        )
        story.append(export_info)
        story.append(Spacer(1, 12))
        
        # Table data
        data = [['ID', 'Device', 'Type', 'Status', 'Start Date', 'Cost (BDT)']]
        
        for maintenance in queryset:
            data.append([
                maintenance.maintenance_id,
                f"{maintenance.device.device_id}\n{maintenance.device.brand}",
                maintenance.get_maintenance_type_display(),
                maintenance.get_status_display(),
                maintenance.start_date.strftime('%Y-%m-%d'),
                f"৳{maintenance.actual_cost or maintenance.estimated_cost or 0:,.2f}"
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        return response


# ============================================================================
# Maintenance Scheduling Views
# ============================================================================

class MaintenanceScheduleView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    Schedule new maintenance (single or recurring).
    Location: maintenance/schedule.html
    """
    form_class = MaintenanceScheduleForm
    template_name = 'maintenance/schedule.html'
    permission_required = 'maintenance.add_maintenance'
    
    def form_valid(self, form):
        """Create maintenance schedule(s)."""
        try:
            created_maintenance = create_maintenance_schedule(
                form.cleaned_data, self.request.user
            )
            
            if len(created_maintenance) == 1:
                messages.success(
                    self.request,
                    f'Maintenance {created_maintenance[0].maintenance_id} scheduled successfully.'
                )
            else:
                messages.success(
                    self.request,
                    f'{len(created_maintenance)} maintenance schedules created successfully.'
                )
            
            return redirect('maintenance:list')
            
        except Exception as e:
            messages.error(
                self.request,
                f'Error creating maintenance schedule: {str(e)}'
            )
            return self.form_invalid(form)
    
    def get_context_data(self, **kwargs):
        """Add additional context."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Schedule Maintenance'
        
        # Add available devices for quick reference
        context['available_devices'] = Device.objects.filter(
            is_active=True,
            status__in=['AVAILABLE', 'ASSIGNED']
        ).select_related('subcategory__category')[:10]
        
        return context


class RecurringMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Manage recurring maintenance schedules.
    Location: maintenance/recurring.html
    """
    template_name = 'maintenance/recurring.html'
    permission_required = 'maintenance.add_maintenance'
    
    def get_context_data(self, **kwargs):
        """Add recurring maintenance data."""
        context = super().get_context_data(**kwargs)
        
        # Get maintenance with follow-up required
        context['recurring_maintenance'] = Maintenance.objects.filter(
            follow_up_required=True,
            follow_up_date__isnull=False
        ).select_related('device', 'vendor').order_by('follow_up_date')
        
        # Get preventive maintenance patterns
        context['preventive_patterns'] = Maintenance.objects.filter(
            maintenance_type='PREVENTIVE',
            status='COMPLETED'
        ).values(
            'device__subcategory__category__name'
        ).annotate(
            avg_interval=Avg('planned_duration'),
            count=Count('id')
        ).order_by('-count')
        
        return context


class MaintenanceCalendarView(LoginRequiredMixin, TemplateView):
    """
    Calendar view of maintenance schedules.
    Location: maintenance/calendar.html
    """
    template_name = 'maintenance/calendar.html'
    
    def get_context_data(self, **kwargs):
        """Add calendar data."""
        context = super().get_context_data(**kwargs)
        
        # Get current month data
        today = timezone.now().date()
        year = int(self.request.GET.get('year', today.year))
        month = int(self.request.GET.get('month', today.month))
        
        # Calculate month boundaries
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
        
        # Get maintenance for the month
        maintenance_records = Maintenance.objects.filter(
            start_date__gte=first_day,
            start_date__lte=last_day
        ).select_related('device', 'vendor')
        
        # Group by date
        calendar_data = {}
        for maintenance in maintenance_records:
            date_key = maintenance.start_date.isoformat()
            if date_key not in calendar_data:
                calendar_data[date_key] = []
            calendar_data[date_key].append(maintenance)
        
        context.update({
            'calendar_data': calendar_data,
            'current_year': year,
            'current_month': month,
            'month_name': calendar.month_name[month],
            'prev_month': month - 1 if month > 1 else 12,
            'next_month': month + 1 if month < 12 else 1,
            'prev_year': year if month > 1 else year - 1,
            'next_year': year if month < 12 else year + 1,
        })
        
        return context


class UpcomingMaintenanceView(LoginRequiredMixin, ListView):
    """
    Display upcoming maintenance in the next 30 days.
    Location: maintenance/upcoming.html
    """
    model = Maintenance
    template_name = 'maintenance/upcoming.html'
    context_object_name = 'upcoming_maintenance'
    
    def get_queryset(self):
        """Get upcoming maintenance."""
        today = timezone.now().date()
        future_date = today + timedelta(days=30)
        
        return Maintenance.objects.filter(
            start_date__gte=today,
            start_date__lte=future_date,
            status='SCHEDULED'
        ).select_related('device', 'vendor').order_by('start_date')
    
    def get_context_data(self, **kwargs):
        """Add timeline data."""
        context = super().get_context_data(**kwargs)
        
        # Group by week
        weeks = {}
        for maintenance in context['upcoming_maintenance']:
            week_start = maintenance.start_date - timedelta(
                days=maintenance.start_date.weekday()
            )
            week_key = week_start.isoformat()
            if week_key not in weeks:
                weeks[week_key] = {
                    'start_date': week_start,
                    'maintenance': []
                }
            weeks[week_key]['maintenance'].append(maintenance)
        
        context['weeks'] = weeks
        return context


class DueSoonMaintenanceView(LoginRequiredMixin, ListView):
    """
    Display maintenance due soon (next 7 days).
    Location: maintenance/due_soon.html
    """
    model = Maintenance
    template_name = 'maintenance/due_soon.html'
    context_object_name = 'due_soon_maintenance'
    
    def get_queryset(self):
        """Get maintenance due soon."""
        today = timezone.now().date()
        soon_date = today + timedelta(days=7)
        
        return Maintenance.objects.filter(
            start_date__gte=today,
            start_date__lte=soon_date,
            status='SCHEDULED'
        ).select_related('device', 'vendor').order_by('start_date')


# ============================================================================
# Quick Actions Views
# ============================================================================

class DeviceMaintenanceView(LoginRequiredMixin, DetailView):
    """
    View maintenance history for a specific device.
    Location: maintenance/device_maintenance.html
    """
    model = Device
    template_name = 'maintenance/device_maintenance.html'
    context_object_name = 'device'
    pk_url_kwarg = 'device_id'
    
    def get_context_data(self, **kwargs):
        """Add maintenance history."""
        context = super().get_context_data(**kwargs)
        
        # Get maintenance history
        context['maintenance_history'] = Maintenance.objects.filter(
            device=self.object
        ).select_related('vendor', 'created_by').order_by('-start_date')
        
        # Get active maintenance
        context['active_maintenance'] = Maintenance.objects.filter(
            device=self.object,
            status__in=['SCHEDULED', 'IN_PROGRESS'],
            is_active=True
        ).first()
        
        # Get maintenance statistics
        all_maintenance = context['maintenance_history']
        context['maintenance_stats'] = {
            'total': all_maintenance.count(),
            'completed': all_maintenance.filter(status='COMPLETED').count(),
            'preventive': all_maintenance.filter(maintenance_type='PREVENTIVE').count(),
            'corrective': all_maintenance.filter(maintenance_type='CORRECTIVE').count(),
            'total_cost': all_maintenance.aggregate(
                total=Sum('actual_cost')
            )['total'] or Decimal('0.00'),
            'avg_cost': all_maintenance.aggregate(
                avg=Avg('actual_cost')
            )['avg'] or Decimal('0.00'),
        }
        
        return context


class QuickScheduleView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """
    Quick schedule maintenance from device page.
    Location: maintenance/quick_schedule.html
    """
    form_class = MaintenanceQuickActionForm
    template_name = 'maintenance/quick_schedule.html'
    permission_required = 'maintenance.add_maintenance'
    
    def get_form_kwargs(self):
        """Pass device to form."""
        kwargs = super().get_form_kwargs()
        device_id = self.request.GET.get('device_id')
        if device_id:
            try:
                kwargs['device'] = Device.objects.get(pk=device_id)
            except Device.DoesNotExist:
                pass
        return kwargs
    
    def form_valid(self, form):
        """Create maintenance record."""
        maintenance = Maintenance(
            device=form.device,
            maintenance_type=form.cleaned_data['maintenance_type'],
            priority=form.cleaned_data['priority'],
            title=f"{form.cleaned_data['maintenance_type']} - {form.device.device_id}",
            description=form.cleaned_data['description'],
            start_date=form.cleaned_data['start_date'],
            expected_end_date=form.cleaned_data['start_date'] + timedelta(days=1),
            estimated_cost=form.cleaned_data.get('estimated_cost', Decimal('0.00')),
            created_by=self.request.user
        )
        maintenance.save()
        
        messages.success(
            self.request,
            f'Maintenance {maintenance.maintenance_id} scheduled successfully.'
        )
        
        return redirect('maintenance:detail', pk=maintenance.pk)


class ReportIssueView(LoginRequiredMixin, FormView):
    """
    Quick issue reporting.
    Location: maintenance/report_issue.html
    """
    form_class = MaintenanceQuickActionForm
    template_name = 'maintenance/report_issue.html'
    
    def get_initial(self):
        """Set initial form values for issue reporting."""
        initial = super().get_initial()
        initial['action_type'] = 'report_issue'
        initial['maintenance_type'] = 'CORRECTIVE'
        initial['priority'] = 'HIGH'
        return initial
    
    def form_valid(self, form):
        """Create corrective maintenance for reported issue."""
        maintenance = Maintenance(
            device=form.device,
            maintenance_type='CORRECTIVE',
            priority=form.cleaned_data['priority'],
            title=f"Issue Reported - {form.device.device_id}",
            description=form.cleaned_data['description'],
            problem_reported=form.cleaned_data['description'],
            start_date=form.cleaned_data['start_date'],
            expected_end_date=form.cleaned_data['start_date'] + timedelta(days=3),
            estimated_cost=form.cleaned_data.get('estimated_cost', Decimal('0.00')),
            created_by=self.request.user
        )
        maintenance.save()
        
        messages.success(
            self.request,
            f'Issue reported successfully. Maintenance {maintenance.maintenance_id} created.'
        )
        
        return redirect('maintenance:detail', pk=maintenance.pk)


# ============================================================================
# Status Change Views
# ============================================================================

class StartMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Start scheduled maintenance."""
    permission_required = 'maintenance.change_maintenance'
    
    def post(self, request, pk):
        """Start maintenance."""
        maintenance = get_object_or_404(Maintenance, pk=pk)
        
        if not maintenance.can_be_started():
            messages.error(request, 'This maintenance cannot be started.')
            return redirect('maintenance:detail', pk=pk)
        
        maintenance.status = 'IN_PROGRESS'
        maintenance.actual_start_date = timezone.now()
        maintenance.updated_by = request.user
        maintenance.save()
        
        messages.success(
            request,
            f'Maintenance {maintenance.maintenance_id} started successfully.'
        )
        
        return redirect('maintenance:detail', pk=pk)


class CompleteMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Complete in-progress maintenance."""
    permission_required = 'maintenance.change_maintenance'
    
    def post(self, request, pk):
        """Complete maintenance."""
        maintenance = get_object_or_404(Maintenance, pk=pk)
        
        if not maintenance.can_be_completed():
            messages.error(request, 'This maintenance cannot be completed.')
            return redirect('maintenance:detail', pk=pk)
        
        # Get completion data from form
        work_performed = request.POST.get('work_performed', '')
        result = request.POST.get('result', 'SUCCESS')
        actual_cost = request.POST.get('actual_cost', '')
        
        maintenance.status = 'COMPLETED'
        maintenance.actual_end_date = timezone.now()
        maintenance.work_performed = work_performed
        maintenance.result = result
        if actual_cost:
            try:
                maintenance.actual_cost = Decimal(actual_cost)
            except:
                pass
        maintenance.updated_by = request.user
        maintenance.save()
        
        messages.success(
            request,
            f'Maintenance {maintenance.maintenance_id} completed successfully.'
        )
        
        return redirect('maintenance:detail', pk=pk)


class CancelMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Cancel scheduled maintenance."""
    permission_required = 'maintenance.change_maintenance'
    
    def post(self, request, pk):
        """Cancel maintenance."""
        maintenance = get_object_or_404(Maintenance, pk=pk)
        
        if not maintenance.can_be_cancelled():
            messages.error(request, 'This maintenance cannot be cancelled.')
            return redirect('maintenance:detail', pk=pk)
        
        cancellation_reason = request.POST.get('cancellation_reason', '')
        
        maintenance.status = 'CANCELLED'
        if cancellation_reason:
            if maintenance.internal_notes:
                maintenance.internal_notes += f"\n\nCancelled: {cancellation_reason}"
            else:
                maintenance.internal_notes = f"Cancelled: {cancellation_reason}"
        maintenance.updated_by = request.user
        maintenance.save()
        
        messages.success(
            request,
            f'Maintenance {maintenance.maintenance_id} cancelled successfully.'
        )
        
        return redirect('maintenance:detail', pk=pk)


class HoldMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Put maintenance on hold."""
    permission_required = 'maintenance.change_maintenance'
    
    def post(self, request, pk):
        """Put maintenance on hold."""
        maintenance = get_object_or_404(Maintenance, pk=pk)
        
        if maintenance.status not in ['SCHEDULED', 'IN_PROGRESS']:
            messages.error(request, 'This maintenance cannot be put on hold.')
            return redirect('maintenance:detail', pk=pk)
        
        hold_reason = request.POST.get('hold_reason', '')
        
        maintenance.status = 'ON_HOLD'
        if hold_reason:
            if maintenance.internal_notes:
                maintenance.internal_notes += f"\n\nPut on hold: {hold_reason}"
            else:
                maintenance.internal_notes = f"Put on hold: {hold_reason}"
        maintenance.updated_by = request.user
        maintenance.save()
        
        messages.success(
            request,
            f'Maintenance {maintenance.maintenance_id} put on hold.'
        )
        
        return redirect('maintenance:detail', pk=pk)


class ResumeMaintenanceView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Resume maintenance from hold."""
    permission_required = 'maintenance.change_maintenance'
    
    def post(self, request, pk):
        """Resume maintenance."""
        maintenance = get_object_or_404(Maintenance, pk=pk)
        
        if maintenance.status != 'ON_HOLD':
            messages.error(request, 'This maintenance is not on hold.')
            return redirect('maintenance:detail', pk=pk)
        
        resume_reason = request.POST.get('resume_reason', '')
        
        # Determine appropriate status
        if maintenance.actual_start_date:
            maintenance.status = 'IN_PROGRESS'
        else:
            maintenance.status = 'SCHEDULED'
        
        if resume_reason:
            if maintenance.internal_notes:
                maintenance.internal_notes += f"\n\nResumed: {resume_reason}"
            else:
                maintenance.internal_notes = f"Resumed: {resume_reason}"
        maintenance.updated_by = request.user
        maintenance.save()
        
        messages.success(
            request,
            f'Maintenance {maintenance.maintenance_id} resumed.'
        )
        
        return redirect('maintenance:detail', pk=pk)


# ============================================================================
# Reports and Analytics Views
# ============================================================================

class MaintenanceReportsView(LoginRequiredMixin, TemplateView):
    """
    Main reports dashboard.
    Location: maintenance/reports.html
    """
    template_name = 'maintenance/reports.html'
    
    def get_context_data(self, **kwargs):
        """Add report options and quick stats."""
        context = super().get_context_data(**kwargs)
        
        # Add quick statistics
        all_maintenance = Maintenance.objects.all()
        context['quick_stats'] = {
            'total_maintenance': all_maintenance.count(),
            'this_month': all_maintenance.filter(
                created_at__month=timezone.now().month,
                created_at__year=timezone.now().year
            ).count(),
            'completed_this_month': all_maintenance.filter(
                status='COMPLETED',
                actual_end_date__month=timezone.now().month,
                actual_end_date__year=timezone.now().year
            ).count(),
            'total_cost_this_month': all_maintenance.filter(
                actual_end_date__month=timezone.now().month,
                actual_end_date__year=timezone.now().year
            ).aggregate(total=Sum('actual_cost'))['total'] or Decimal('0.00'),
        }
        
        # Add recent reports
        context['recent_reports'] = [
            {
                'title': 'Monthly Maintenance Summary',
                'description': 'Summary of all maintenance activities this month',
                'url': reverse('maintenance:report_summary'),
            },
            {
                'title': 'Cost Analysis Report',
                'description': 'Detailed analysis of maintenance costs',
                'url': reverse('maintenance:report_cost'),
            },
            {
                'title': 'Vendor Performance Report',
                'description': 'Performance analysis of maintenance vendors',
                'url': reverse('maintenance:report_vendor'),
            },
        ]
        
        return context


class MaintenanceSummaryReportView(LoginRequiredMixin, TemplateView):
    """
    Summary report of maintenance activities.
    Location: maintenance/report_summary.html
    """
    template_name = 'maintenance/report_summary.html'
    
    def get_context_data(self, **kwargs):
        """Generate summary report data."""
        context = super().get_context_data(**kwargs)
        
        # Date range from GET parameters
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if not date_from:
            date_from = (timezone.now().date() - timedelta(days=30)).isoformat()
        if not date_to:
            date_to = timezone.now().date().isoformat()
        
        # Filter maintenance records
        maintenance_qs = Maintenance.objects.filter(
            start_date__gte=date_from,
            start_date__lte=date_to
        )
        
        # Status summary
        context['status_summary'] = maintenance_qs.values(
            'status'
        ).annotate(
            count=Count('id'),
            total_cost=Sum('actual_cost')
        ).order_by('-count')
        
        # Type summary
        context['type_summary'] = maintenance_qs.values(
            'maintenance_type'
        ).annotate(
            count=Count('id'),
            avg_cost=Avg('actual_cost')
        ).order_by('-count')
        
        # Priority summary
        context['priority_summary'] = maintenance_qs.values(
            'priority'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Cost summary
        context['cost_summary'] = maintenance_qs.aggregate(
            total_estimated=Sum('estimated_cost'),
            total_actual=Sum('actual_cost'),
            avg_cost=Avg('actual_cost'),
            max_cost=Max('actual_cost'),
            min_cost=Min('actual_cost')
        )
        
        # Monthly trend
        context['monthly_trend'] = maintenance_qs.extra(
            select={'month': "DATE_FORMAT(start_date, '%%Y-%%m')"}
        ).values('month').annotate(
            count=Count('id'),
            total_cost=Sum('actual_cost')
        ).order_by('month')
        
        context['date_from'] = date_from
        context['date_to'] = date_to
        
        return context


class MaintenanceCostAnalysisView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Detailed cost analysis report.
    Location: maintenance/report_cost.html
    """
    template_name = 'maintenance/report_cost.html'
    permission_required = 'maintenance.view_maintenance_costs'
    
    def get_context_data(self, **kwargs):
        """Generate cost analysis data."""
        context = super().get_context_data(**kwargs)
        
        # Cost by maintenance type
        context['cost_by_type'] = Maintenance.objects.values(
            'maintenance_type'
        ).annotate(
            total_cost=Sum('actual_cost'),
            avg_cost=Avg('actual_cost'),
            count=Count('id'),
            estimated_total=Sum('estimated_cost')
        ).order_by('-total_cost')
        
        # Cost by device category
        context['cost_by_category'] = Maintenance.objects.select_related(
            'device__subcategory__category'
        ).values(
            'device__subcategory__category__name'
        ).annotate(
            total_cost=Sum('actual_cost'),
            avg_cost=Avg('actual_cost'),
            count=Count('id')
        ).order_by('-total_cost')
        
        # Cost variance analysis
        context['cost_variance'] = Maintenance.objects.filter(
            actual_cost__isnull=False,
            estimated_cost__isnull=False
        ).extra(
            select={
                'variance': 'actual_cost - estimated_cost',
                'variance_pct': '((actual_cost - estimated_cost) / estimated_cost) * 100'
            }
        ).order_by('-variance')[:10]
        
        # Monthly cost trend
        context['monthly_costs'] = Maintenance.objects.filter(
            actual_cost__isnull=False
        ).extra(
            select={'month': "DATE_FORMAT(actual_end_date, '%%Y-%%m')"}
        ).values('month').annotate(
            total_cost=Sum('actual_cost'),
            avg_cost=Avg('actual_cost'),
            count=Count('id')
        ).order_by('month')
        
        return context


class MaintenancePerformanceReportView(LoginRequiredMixin, TemplateView):
    """
    Performance analysis report.
    Location: maintenance/report_performance.html
    """
    template_name = 'maintenance/report_performance.html'
    
    def get_context_data(self, **kwargs):
        """Generate performance data."""
        context = super().get_context_data(**kwargs)
        
        # Completion rate analysis
        all_maintenance = Maintenance.objects.all()
        completed_maintenance = all_maintenance.filter(status='COMPLETED')
        
        context['completion_stats'] = {
            'total': all_maintenance.count(),
            'completed': completed_maintenance.count(),
            'completion_rate': (
                completed_maintenance.count() / all_maintenance.count() * 100
                if all_maintenance.count() > 0 else 0
            ),
            'on_time_completion': completed_maintenance.filter(
                actual_end_date__lte=F('expected_end_date')
            ).count()
        }
        
        # Duration analysis
        context['duration_stats'] = completed_maintenance.aggregate(
            avg_planned=Avg('planned_duration'),
            avg_actual=Avg('actual_duration'),
            max_duration=Max('actual_duration'),
            min_duration=Min('actual_duration')
        )
        
        # Overdue analysis
        overdue_maintenance = [m for m in all_maintenance if m.is_overdue]
        context['overdue_stats'] = {
            'count': len(overdue_maintenance),
            'percentage': (
                len(overdue_maintenance) / all_maintenance.count() * 100
                if all_maintenance.count() > 0 else 0
            )
        }
        
        # Success rate by type
        context['success_by_type'] = Maintenance.objects.filter(
            status='COMPLETED'
        ).values('maintenance_type').annotate(
            total=Count('id'),
            successful=Count(
                Case(When(result='SUCCESS', then=1), output_field=IntegerField())
            )
        ).order_by('-total')
        
        return context


class VendorPerformanceReportView(LoginRequiredMixin, TemplateView):
    """
    Vendor performance analysis.
    Location: maintenance/report_vendor.html
    """
    template_name = 'maintenance/report_vendor.html'
    
    def get_context_data(self, **kwargs):
        """Generate vendor performance data."""
        context = super().get_context_data(**kwargs)
        
        # Vendor performance summary
        context['vendor_performance'] = Maintenance.objects.filter(
            vendor__isnull=False
        ).values(
            'vendor__name',
            'vendor__id'
        ).annotate(
            total_jobs=Count('id'),
            completed_jobs=Count(
                Case(When(status='COMPLETED', then=1), output_field=IntegerField())
            ),
            avg_cost=Avg('actual_cost'),
            total_cost=Sum('actual_cost'),
            avg_rating=Avg('satisfaction_rating'),
            on_time_completion=Count(
                Case(
                    When(
                        Q(status='COMPLETED') & Q(actual_end_date__lte=F('expected_end_date')),
                        then=1
                    ),
                    output_field=IntegerField()
                )
            )
        ).order_by('-total_jobs')
        
        # Top performing vendors
        context['top_vendors'] = context['vendor_performance'][:5]
        
        # Vendor cost comparison
        context['vendor_costs'] = Maintenance.objects.filter(
            vendor__isnull=False,
            actual_cost__isnull=False
        ).values('vendor__name').annotate(
            avg_cost=Avg('actual_cost'),
            total_cost=Sum('actual_cost'),
            job_count=Count('id')
        ).order_by('-total_cost')
        
        return context


class DeviceMaintenanceHistoryView(LoginRequiredMixin, ListView):
    """
    Device maintenance history report.
    Location: maintenance/report_device_history.html
    """
    model = Maintenance
    template_name = 'maintenance/report_device_history.html'
    context_object_name = 'maintenance_records'
    paginate_by = 50
    
    def get_queryset(self):
        """Filter by device category or specific device."""
        queryset = Maintenance.objects.select_related(
            'device', 'vendor'
        ).order_by('-start_date')
        
        device_id = self.request.GET.get('device_id')
        category_id = self.request.GET.get('category_id')
        
        if device_id:
            queryset = queryset.filter(device_id=device_id)
        elif category_id:
            queryset = queryset.filter(
                device__subcategory__category_id=category_id
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add device and category data."""
        context = super().get_context_data(**kwargs)
        
        # Add filter options
        context['devices'] = Device.objects.filter(
            is_active=True
        ).select_related('subcategory__category').order_by('device_id')
        
        context['categories'] = DeviceCategory.objects.filter(
            is_active=True
        ).order_by('name')
        
        # Add selected filters
        context['selected_device'] = self.request.GET.get('device_id')
        context['selected_category'] = self.request.GET.get('category_id')
        
        return context

class MaintenanceTrendsReportView(LoginRequiredMixin, TemplateView):
    """
    Maintenance trends analysis report.
    Location: maintenance/report_trends.html
    """
    template_name = 'maintenance/report_trends.html'
    
    def get_context_data(self, **kwargs):
        """Generate trends analysis data."""
        context = super().get_context_data(**kwargs)
        
        # Yearly trend analysis
        current_year = timezone.now().year
        yearly_trends = []
        
        for year in range(current_year - 4, current_year + 1):
            year_data = Maintenance.objects.filter(
                start_date__year=year
            ).aggregate(
                total_count=Count('id'),
                total_cost=Sum('actual_cost'),
                avg_cost=Avg('actual_cost'),
                preventive_count=Count(Case(
                    When(maintenance_type='PREVENTIVE', then=1),
                    output_field=IntegerField()
                )),
                corrective_count=Count(Case(
                    When(maintenance_type='CORRECTIVE', then=1),
                    output_field=IntegerField()
                ))
            )
            year_data['year'] = year
            yearly_trends.append(year_data)
        
        context['yearly_trends'] = yearly_trends
        
        # Monthly trend for current year
        monthly_trends = []
        for month in range(1, 13):
            month_data = Maintenance.objects.filter(
                start_date__year=current_year,
                start_date__month=month
            ).aggregate(
                count=Count('id'),
                cost=Sum('actual_cost'),
                avg_cost=Avg('actual_cost')
            )
            month_data['month'] = month
            month_data['month_name'] = calendar.month_name[month]
            monthly_trends.append(month_data)
        
        context['monthly_trends'] = monthly_trends
        
        # Device category trends
        context['category_trends'] = Maintenance.objects.select_related(
            'device__subcategory__category'
        ).values(
            'device__subcategory__category__name'
        ).annotate(
            count=Count('id'),
            avg_cost=Avg('actual_cost'),
            trend_score=Avg('satisfaction_rating')
        ).order_by('-count')[:10]
        
        # Seasonal patterns
        context['seasonal_patterns'] = Maintenance.objects.extra(
            select={'quarter': 'QUARTER(start_date)'}
        ).values('quarter').annotate(
            count=Count('id'),
            avg_cost=Avg('actual_cost')
        ).order_by('quarter')
        
        return context


# ============================================================================
# Export Views
# ============================================================================

class ExportMaintenancePDFView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Export maintenance data as PDF report.
    """
    permission_required = 'maintenance.view_maintenance'
    
    def get(self, request):
        """Generate and return PDF report."""
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        status = request.GET.get('status')
        
        # Build queryset
        queryset = Maintenance.objects.select_related('device', 'vendor', 'created_by')
        
        if date_from:
            queryset = queryset.filter(start_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(start_date__lte=date_to)
        if status:
            queryset = queryset.filter(status=status)
        
        queryset = queryset.order_by('-start_date')
        
        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="maintenance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Generate PDF
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story = []
        
        # Title and header
        title = Paragraph("Maintenance Report - Bangladesh Parliament Secretariat", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Report metadata
        metadata = [
            f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            f"Period: {date_from or 'All'} to {date_to or 'All'}",
            f"Status Filter: {status or 'All'}",
            f"Total Records: {queryset.count()}"
        ]
        
        for meta in metadata:
            story.append(Paragraph(meta, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary statistics
        if queryset.exists():
            stats = queryset.aggregate(
                total_cost=Sum('actual_cost'),
                avg_cost=Avg('actual_cost'),
                completed_count=Count(Case(
                    When(status='COMPLETED', then=1),
                    output_field=IntegerField()
                ))
            )
            
            summary_title = Paragraph("Summary Statistics", styles['Heading2'])
            story.append(summary_title)
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Cost', f"৳{stats['total_cost'] or 0:,.2f}"],
                ['Average Cost', f"৳{stats['avg_cost'] or 0:,.2f}"],
                ['Completed', f"{stats['completed_count']} ({stats['completed_count']/queryset.count()*100:.1f}%)"],
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))
        
        # Detailed records table
        detail_title = Paragraph("Detailed Records", styles['Heading2'])
        story.append(detail_title)
        
        # Table headers
        data = [['ID', 'Device', 'Type', 'Status', 'Start Date', 'Cost (BDT)']]
        
        # Add data rows (limit for PDF performance)
        for maintenance in queryset[:100]:  # Limit to 100 records
            data.append([
                maintenance.maintenance_id,
                f"{maintenance.device.device_id}\n{maintenance.device.brand}",
                maintenance.get_maintenance_type_display(),
                maintenance.get_status_display(),
                maintenance.start_date.strftime('%Y-%m-%d') if maintenance.start_date else '',
                f"৳{maintenance.actual_cost or maintenance.estimated_cost or 0:,.2f}"
            ])
        
        # Create and style table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        return response


class ExportMaintenanceExcelView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Export maintenance data as Excel file.
    """
    permission_required = 'maintenance.view_maintenance'
    
    def get(self, request):
        """Generate and return Excel file."""
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        status = request.GET.get('status')
        
        # Build queryset
        queryset = Maintenance.objects.select_related('device', 'vendor', 'created_by')
        
        if date_from:
            queryset = queryset.filter(start_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(start_date__lte=date_to)
        if status:
            queryset = queryset.filter(status=status)
        
        queryset = queryset.order_by('-start_date')
        
        # Create workbook
        wb = Workbook()
        
        # Summary worksheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Summary headers
        summary_headers = [
            'Report Generated', 'Date From', 'Date To', 'Status Filter',
            'Total Records', 'Total Cost (BDT)', 'Average Cost (BDT)'
        ]
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Summary data
        stats = queryset.aggregate(
            total_cost=Sum('actual_cost'),
            avg_cost=Avg('actual_cost')
        )
        
        summary_data = [
            timezone.now().strftime('%Y-%m-%d %H:%M'),
            date_from or 'All',
            date_to or 'All',
            status or 'All',
            queryset.count(),
            float(stats['total_cost']) if stats['total_cost'] else 0,
            float(stats['avg_cost']) if stats['avg_cost'] else 0
        ]
        
        for col, value in enumerate(summary_data, 1):
            ws_summary.cell(row=2, column=col, value=value)
        
        # Auto-adjust column widths for summary
        for column in ws_summary.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Detailed data worksheet
        ws_detail = wb.create_sheet(title="Detailed Records")
        
        # Detail headers
        detail_headers = [
            'Maintenance ID', 'Device ID', 'Device Brand', 'Device Model',
            'Type', 'Priority', 'Status', 'Start Date', 'Expected End Date',
            'Actual Start Date', 'Actual End Date', 'Estimated Cost (BDT)', 
            'Actual Cost (BDT)', 'Parts Cost (BDT)', 'Labor Cost (BDT)',
            'Vendor', 'Technician', 'Result', 'Satisfaction Rating',
            'Downtime Hours', 'Created By', 'Created Date'
        ]
        
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Detail data
        for row, maintenance in enumerate(queryset, 2):
            detail_data = [
                maintenance.maintenance_id,
                maintenance.device.device_id if maintenance.device else '',
                maintenance.device.brand if maintenance.device else '',
                maintenance.device.model if maintenance.device else '',
                maintenance.get_maintenance_type_display(),
                maintenance.get_priority_display(),
                maintenance.get_status_display(),
                maintenance.start_date,
                maintenance.expected_end_date,
                maintenance.actual_start_date,
                maintenance.actual_end_date,
                float(maintenance.estimated_cost) if maintenance.estimated_cost else 0,
                float(maintenance.actual_cost) if maintenance.actual_cost else 0,
                float(maintenance.parts_cost) if maintenance.parts_cost else 0,
                float(maintenance.labor_cost) if maintenance.labor_cost else 0,
                maintenance.vendor.name if maintenance.vendor else '',
                maintenance.technician_name or '',
                maintenance.get_result_display() if maintenance.result else '',
                maintenance.satisfaction_rating or '',
                float(maintenance.downtime_hours) if maintenance.downtime_hours else 0,
                maintenance.created_by.get_full_name() if maintenance.created_by else '',
                maintenance.created_at
            ]
            
            for col, value in enumerate(detail_data, 1):
                ws_detail.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths for detail
        for column in ws_detail.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws_detail.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="maintenance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response


class ExportMaintenanceCSVView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Export maintenance data as CSV file.
    """
    permission_required = 'maintenance.view_maintenance'
    
    def get(self, request):
        """Generate and return CSV file."""
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        status = request.GET.get('status')
        
        # Build queryset
        queryset = Maintenance.objects.select_related('device', 'vendor', 'created_by')
        
        if date_from:
            queryset = queryset.filter(start_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(start_date__lte=date_to)
        if status:
            queryset = queryset.filter(status=status)
        
        queryset = queryset.order_by('-start_date')
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="maintenance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # CSV writer
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'Maintenance ID', 'Device ID', 'Device Brand', 'Device Model',
            'Type', 'Priority', 'Status', 'Start Date', 'Expected End Date',
            'Actual Start Date', 'Actual End Date', 'Estimated Cost (BDT)', 
            'Actual Cost (BDT)', 'Parts Cost (BDT)', 'Labor Cost (BDT)',
            'Vendor', 'Technician', 'Result', 'Satisfaction Rating',
            'Downtime Hours', 'Created By', 'Created Date'
        ])
        
        # Data rows
        for maintenance in queryset:
            writer.writerow([
                maintenance.maintenance_id,
                maintenance.device.device_id if maintenance.device else '',
                maintenance.device.brand if maintenance.device else '',
                maintenance.device.model if maintenance.device else '',
                maintenance.get_maintenance_type_display(),
                maintenance.get_priority_display(),
                maintenance.get_status_display(),
                maintenance.start_date.strftime('%Y-%m-%d') if maintenance.start_date else '',
                maintenance.expected_end_date.strftime('%Y-%m-%d') if maintenance.expected_end_date else '',
                maintenance.actual_start_date.strftime('%Y-%m-%d %H:%M') if maintenance.actual_start_date else '',
                maintenance.actual_end_date.strftime('%Y-%m-%d %H:%M') if maintenance.actual_end_date else '',
                maintenance.estimated_cost,
                maintenance.actual_cost or '',
                maintenance.parts_cost or '',
                maintenance.labor_cost or '',
                maintenance.vendor.name if maintenance.vendor else '',
                maintenance.technician_name or '',
                maintenance.get_result_display() if maintenance.result else '',
                maintenance.satisfaction_rating or '',
                maintenance.downtime_hours or '',
                maintenance.created_by.get_full_name() if maintenance.created_by else '',
                maintenance.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response


# ============================================================================
# Dashboard Views
# ============================================================================

class MaintenanceDashboardView(LoginRequiredMixin, TemplateView):
    """
    Main maintenance dashboard.
    Location: maintenance/dashboard.html
    """
    template_name = 'maintenance/dashboard.html'
    
    def get_context_data(self, **kwargs):
        """Add dashboard data."""
        context = super().get_context_data(**kwargs)
        
        # Overall statistics
        all_maintenance = Maintenance.objects.all()
        context['stats'] = {
            'total': all_maintenance.count(),
            'scheduled': all_maintenance.filter(status='SCHEDULED').count(),
            'in_progress': all_maintenance.filter(status='IN_PROGRESS').count(),
            'completed': all_maintenance.filter(status='COMPLETED').count(),
            'overdue': len([m for m in all_maintenance if m.is_overdue]),
        }
        
        # Recent maintenance activity
        context['recent_maintenance'] = Maintenance.objects.select_related(
            'device', 'vendor', 'created_by'
        ).order_by('-created_at')[:10]
        
        # Upcoming maintenance (next 7 days)
        upcoming_date = timezone.now().date() + timedelta(days=7)
        context['upcoming_maintenance'] = Maintenance.objects.filter(
            start_date__gte=timezone.now().date(),
            start_date__lte=upcoming_date,
            status='SCHEDULED'
        ).select_related('device', 'vendor').order_by('start_date')[:10]
        
        # Overdue maintenance
        context['overdue_maintenance'] = Maintenance.objects.filter(
            expected_end_date__lt=timezone.now().date(),
            status__in=['SCHEDULED', 'IN_PROGRESS']
        ).select_related('device', 'vendor').order_by('expected_end_date')[:10]
        
        # Monthly cost trend (last 6 months)
        monthly_costs = []
        for i in range(6):
            month_date = timezone.now().date().replace(day=1) - timedelta(days=i*30)
            month_cost = Maintenance.objects.filter(
                start_date__year=month_date.year,
                start_date__month=month_date.month
            ).aggregate(total=Sum('actual_cost'))['total'] or 0
            
            monthly_costs.append({
                'month': month_date.strftime('%B %Y'),
                'cost': float(month_cost),
                'count': Maintenance.objects.filter(
                    start_date__year=month_date.year,
                    start_date__month=month_date.month
                ).count()
            })
        
        context['monthly_costs'] = list(reversed(monthly_costs))
        
        # Maintenance by type (pie chart data)
        context['maintenance_by_type'] = Maintenance.objects.values(
            'maintenance_type'
        ).annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Device categories requiring most maintenance
        context['high_maintenance_categories'] = Maintenance.objects.select_related(
            'device__subcategory__category'
        ).values(
            'device__subcategory__category__name'
        ).annotate(
            count=Count('id'),
            avg_cost=Avg('actual_cost')
        ).order_by('-count')[:5]
        
        # Performance metrics
        completed_maintenance = all_maintenance.filter(status='COMPLETED')
        if completed_maintenance.exists():
            context['performance_metrics'] = {
                'completion_rate': completed_maintenance.count() / all_maintenance.count() * 100,
                'avg_satisfaction': completed_maintenance.aggregate(
                    avg=Avg('satisfaction_rating')
                )['avg'] or 0,
                'on_time_completion': completed_maintenance.filter(
                    actual_end_date__lte=F('expected_end_date')
                ).count() / completed_maintenance.count() * 100,
                'total_cost_this_month': all_maintenance.filter(
                    actual_end_date__month=timezone.now().month,
                    actual_end_date__year=timezone.now().year
                ).aggregate(total=Sum('actual_cost'))['total'] or 0
            }
        else:
            context['performance_metrics'] = {
                'completion_rate': 0,
                'avg_satisfaction': 0,
                'on_time_completion': 0,
                'total_cost_this_month': 0
            }
        
        return context


class MaintenanceAlertsView(LoginRequiredMixin, TemplateView):
    """
    Maintenance alerts and notifications.
    Location: maintenance/alerts.html
    """
    template_name = 'maintenance/alerts.html'
    
    def get_context_data(self, **kwargs):
        """Add alerts data."""
        context = super().get_context_data(**kwargs)
        
        today = timezone.now().date()
        
        # Critical alerts
        context['critical_alerts'] = {
            'overdue': Maintenance.objects.filter(
                expected_end_date__lt=today,
                status__in=['SCHEDULED', 'IN_PROGRESS'],
                priority__in=['CRITICAL', 'EMERGENCY']
            ).select_related('device', 'vendor'),
            
            'emergency_pending': Maintenance.objects.filter(
                maintenance_type='EMERGENCY',
                status='SCHEDULED',
                requires_approval=True,
                approved_by__isnull=True
            ).select_related('device', 'vendor'),
            
            'high_cost_variance': Maintenance.objects.filter(
                status='COMPLETED',
                actual_cost__gt=F('estimated_cost') * 1.5  # 50% over budget
            ).select_related('device', 'vendor')[:10]
        }
        
        # Warning alerts
        context['warning_alerts'] = {
            'due_soon': Maintenance.objects.filter(
                start_date__gte=today,
                start_date__lte=today + timedelta(days=3),
                status='SCHEDULED'
            ).select_related('device', 'vendor'),
            
            'pending_approval': Maintenance.objects.filter(
                requires_approval=True,
                approved_by__isnull=True,
                status='SCHEDULED'
            ).select_related('device', 'vendor'),
            
            'long_running': Maintenance.objects.filter(
                status='IN_PROGRESS',
                actual_start_date__lt=timezone.now() - timedelta(days=7)
            ).select_related('device', 'vendor')
        }
        
        # Info alerts
        context['info_alerts'] = {
            'follow_up_due': Maintenance.objects.filter(
                follow_up_required=True,
                follow_up_date__lte=today + timedelta(days=7),
                status='COMPLETED'
            ).select_related('device', 'vendor'),
            
            'warranty_ending': Maintenance.objects.filter(
                is_warranty_service=True,
                device__warranty__end_date__lte=today + timedelta(days=30),
                device__warranty__end_date__gte=today
            ).select_related('device', 'vendor', 'device__warranty')
        }
        
        return context


class MaintenanceStatisticsView(LoginRequiredMixin, TemplateView):
    """
    Detailed maintenance statistics.
    Location: maintenance/statistics.html
    """
    template_name = 'maintenance/statistics.html'
    
    def get_context_data(self, **kwargs):
        """Add detailed statistics."""
        context = super().get_context_data(**kwargs)
        
        # Time-based statistics
        today = timezone.now().date()
        
        context['time_stats'] = {
            'today': Maintenance.objects.filter(start_date=today).count(),
            'this_week': Maintenance.objects.filter(
                start_date__gte=today - timedelta(days=today.weekday())
            ).count(),
            'this_month': Maintenance.objects.filter(
                start_date__year=today.year,
                start_date__month=today.month
            ).count(),
            'this_year': Maintenance.objects.filter(
                start_date__year=today.year
            ).count()
        }
        
        # Status distribution
        context['status_distribution'] = Maintenance.objects.values(
            'status'
        ).annotate(
            count=Count('id'),
            percentage=Count('id') * 100.0 / Maintenance.objects.count()
        ).order_by('-count')
        
        # Type distribution
        context['type_distribution'] = Maintenance.objects.values(
            'maintenance_type'
        ).annotate(
            count=Count('id'),
            avg_cost=Avg('actual_cost'),
            total_cost=Sum('actual_cost')
        ).order_by('-count')
        
        # Priority distribution
        context['priority_distribution'] = Maintenance.objects.values(
            'priority'
        ).annotate(
            count=Count('id'),
            avg_duration=Avg(
                Case(
                    When(
                        actual_start_date__isnull=False,
                        actual_end_date__isnull=False,
                        then=F('actual_end_date') - F('actual_start_date')
                    ),
                    output_field=DecimalField()
                )
            )
        ).order_by('-count')
        
        # Cost statistics
        cost_stats = Maintenance.objects.aggregate(
            total_estimated=Sum('estimated_cost'),
            total_actual=Sum('actual_cost'),
            avg_estimated=Avg('estimated_cost'),
            avg_actual=Avg('actual_cost'),
            max_cost=Max('actual_cost'),
            min_cost=Min('actual_cost')
        )
        
        context['cost_stats'] = cost_stats
        context['cost_variance'] = (
            (cost_stats['total_actual'] or 0) - (cost_stats['total_estimated'] or 0)
        )
        
        # Performance metrics
        all_maintenance = Maintenance.objects.all()
        completed_maintenance = all_maintenance.filter(status='COMPLETED')
        
        if completed_maintenance.exists():
            context['performance_stats'] = {
                'completion_rate': completed_maintenance.count() / all_maintenance.count() * 100,
                'success_rate': completed_maintenance.filter(result='SUCCESS').count() / completed_maintenance.count() * 100,
                'avg_satisfaction': completed_maintenance.aggregate(avg=Avg('satisfaction_rating'))['avg'] or 0,
                'on_time_rate': completed_maintenance.filter(
                    actual_end_date__lte=F('expected_end_date')
                ).count() / completed_maintenance.count() * 100
            }
        else:
            context['performance_stats'] = {
                'completion_rate': 0,
                'success_rate': 0,
                'avg_satisfaction': 0,
                'on_time_rate': 0
            }
        
        # Top performing vendors
        context['top_vendors'] = Maintenance.objects.filter(
            vendor__isnull=False,
            status='COMPLETED'
        ).values('vendor__name').annotate(
            job_count=Count('id'),
            avg_rating=Avg('satisfaction_rating'),
            avg_cost=Avg('actual_cost'),
            on_time_count=Count(
                Case(
                    When(actual_end_date__lte=F('expected_end_date'), then=1),
                    output_field=IntegerField()
                )
            )
        ).order_by('-job_count')[:10]
        
        # Device categories with most maintenance
        context['high_maintenance_devices'] = Maintenance.objects.select_related(
            'device__subcategory__category'
        ).values(
            'device__subcategory__category__name'
        ).annotate(
            maintenance_count=Count('id'),
            avg_cost=Avg('actual_cost'),
            total_cost=Sum('actual_cost')
        ).order_by('-maintenance_count')[:10]
        
        return context


# ============================================================================
# AJAX API Endpoints
# ============================================================================

@login_required
@require_http_methods(["GET"])
def validate_device_maintenance(request):
    """
    Validate if device can have maintenance scheduled.
    """
    device_id = request.GET.get('device_id')
    
    if not device_id:
        return JsonResponse({'valid': False, 'message': 'Device ID is required'})
    
    try:
        device = Device.objects.get(pk=device_id)
        
        # Check if device is available for maintenance
        if device.status in ['RETIRED', 'LOST']:
            return JsonResponse({
                'valid': False,
                'message': f'Cannot schedule maintenance for {device.status.lower()} devices'
            })
        
        # Check for conflicting active maintenance
        active_maintenance = Maintenance.objects.filter(
            device=device,
            status__in=['SCHEDULED', 'IN_PROGRESS'],
            is_active=True
        ).first()
        
        if active_maintenance:
            return JsonResponse({
                'valid': False,
                'message': f'Device already has active maintenance: {active_maintenance.maintenance_id}',
                'conflicting_maintenance': {
                    'id': active_maintenance.maintenance_id,
                    'status': active_maintenance.get_status_display(),
                    'start_date': active_maintenance.start_date.isoformat() if active_maintenance.start_date else None
                }
            })
        
        # Device is valid for maintenance
        return JsonResponse({
            'valid': True,
            'message': 'Device is available for maintenance',
            'device_info': {
                'id': device.device_id,
                'brand': device.brand,
                'model': device.model,
                'status': device.get_status_display(),
                'category': device.subcategory.category.name if device.subcategory else 'Unknown'
            }
        })
        
    except Device.DoesNotExist:
        return JsonResponse({'valid': False, 'message': 'Device not found'})
    except Exception as e:
        return JsonResponse({'valid': False, 'message': f'Error validating device: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def suggest_maintenance_cost(request):
    """
    Suggest maintenance cost based on device type and maintenance type.
    """
    device_id = request.GET.get('device_id')
    maintenance_type = request.GET.get('maintenance_type')
    
    if not device_id or not maintenance_type:
        return JsonResponse({'error': 'Device ID and maintenance type are required'})
    
    try:
        device = Device.objects.get(pk=device_id)
        
        # Get historical costs for similar maintenance
        similar_maintenance = Maintenance.objects.filter(
            device__subcategory=device.subcategory,
            maintenance_type=maintenance_type,
            actual_cost__isnull=False
        )
        
        if similar_maintenance.exists():
            avg_cost = similar_maintenance.aggregate(avg=Avg('actual_cost'))['avg']
            recent_cost = similar_maintenance.order_by('-actual_end_date').first().actual_cost
            
            # Suggest cost based on average and recent costs
            suggested_cost = (avg_cost + recent_cost) / 2 if recent_cost else avg_cost
            
            return JsonResponse({
                'suggested_cost': float(suggested_cost),
                'avg_cost': float(avg_cost),
                'recent_cost': float(recent_cost) if recent_cost else None,
                'sample_size': similar_maintenance.count(),
                'confidence': 'high' if similar_maintenance.count() >= 5 else 'medium' if similar_maintenance.count() >= 2 else 'low'
            })
        else:
            # Fallback to general estimates
            base_costs = {
                'PREVENTIVE': 5000,
                'CORRECTIVE': 8000,
                'EMERGENCY': 15000,
                'UPGRADE': 12000,
                'INSPECTION': 2000,
                'CLEANING': 1500,
                'CALIBRATION': 3000,
                'REPLACEMENT': 10000,
                'WARRANTY': 0,
                'OTHER': 5000
            }
            
            suggested_cost = base_costs.get(maintenance_type, 5000)
            
            return JsonResponse({
                'suggested_cost': suggested_cost,
                'avg_cost': suggested_cost,
                'recent_cost': None,
                'sample_size': 0,
                'confidence': 'low',
                'note': 'Based on general estimates - no historical data available'
            })
            
    except Device.DoesNotExist:
        return JsonResponse({'error': 'Device not found'})
    except Exception as e:
        return JsonResponse({'error': f'Error suggesting cost: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def vendor_search_api(request):
    """
    Search vendors for maintenance assignment.
    """
    query = request.GET.get('q', '')
    maintenance_type = request.GET.get('type', '')
    
    vendors = Vendor.objects.filter(
        is_active=True,
        vendor_type__in=['SERVICE_PROVIDER', 'BOTH']
    )
    
    if query:
        vendors = vendors.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query) |
            Q(services_offered__icontains=query)
        )
    
    # Filter by maintenance expertise if available
    if maintenance_type:
        vendors = vendors.filter(
            Q(services_offered__icontains=maintenance_type.lower()) |
            Q(specialization__icontains=maintenance_type.lower())
        )
    
    results = []
    for vendor in vendors[:10]:  # Limit results
        # Get vendor performance metrics
        vendor_maintenance = Maintenance.objects.filter(vendor=vendor, status='COMPLETED')
        
        if vendor_maintenance.exists():
            avg_rating = vendor_maintenance.aggregate(avg=Avg('satisfaction_rating'))['avg'] or 0
            job_count = vendor_maintenance.count()
            avg_cost = vendor_maintenance.aggregate(avg=Avg('actual_cost'))['avg'] or 0
        else:
            avg_rating = 0
            job_count = 0
            avg_cost = 0
        
        results.append({
            'id': vendor.id,
            'name': vendor.name,
            'contact_person': vendor.contact_person,
            'phone': vendor.phone,
            'email': vendor.email,
            'services': vendor.services_offered,
            'avg_rating': round(avg_rating, 1),
            'job_count': job_count,
            'avg_cost': float(avg_cost),
            'location': vendor.address
        })
    
    return JsonResponse({'vendors': results})


@login_required
@require_http_methods(["GET"])
def get_device_info(request, device_id):
    """
    Get detailed device information for maintenance forms.
    """
    try:
        device = Device.objects.select_related(
            'subcategory__category', 'warranty'
        ).get(pk=device_id)
        
        # Get maintenance history
        maintenance_history = Maintenance.objects.filter(
            device=device
        ).order_by('-start_date')[:5]
        
        # Get current assignments
        current_assignment = None
        if hasattr(device, 'assignments'):
            current_assignment = device.assignments.filter(is_active=True).first()
        
        device_info = {
            'id': device.device_id,
            'brand': device.brand,
            'model': device.model,
            'serial_number': device.serial_number,
            'status': device.get_status_display(),
            'category': {
                'name': device.subcategory.category.name if device.subcategory else 'Unknown',
                'code': device.subcategory.category.code if device.subcategory else '',
            },
            'subcategory': {
                'name': device.subcategory.name if device.subcategory else 'Unknown',
                'code': device.subcategory.code if device.subcategory else '',
            },
            'specifications': device.specifications,
            'purchase_date': device.purchase_date.isoformat() if device.purchase_date else None,
            'warranty_info': {
                'start_date': device.warranty.start_date.isoformat() if hasattr(device, 'warranty') and device.warranty.start_date else None,
                'end_date': device.warranty.end_date.isoformat() if hasattr(device, 'warranty') and device.warranty.end_date else None,
                'is_active': device.warranty.is_active if hasattr(device, 'warranty') else False,
                'provider': device.warranty.provider if hasattr(device, 'warranty') else None,
            },
            'current_assignment': {
                'user': current_assignment.assigned_to.get_full_name() if current_assignment else None,
                'location': current_assignment.location.name if current_assignment and current_assignment.location else None,
                'date': current_assignment.assigned_date.isoformat() if current_assignment else None,
            } if current_assignment else None,
            'maintenance_stats': {
                'total_maintenance': maintenance_history.count(),
                'last_maintenance': maintenance_history.first().start_date.isoformat() if maintenance_history.exists() else None,
                'preventive_count': maintenance_history.filter(maintenance_type='PREVENTIVE').count(),
                'corrective_count': maintenance_history.filter(maintenance_type='CORRECTIVE').count(),
            }
        }
        
        return JsonResponse({'device': device_info})
        
    except Device.DoesNotExist:
        return JsonResponse({'error': 'Device not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving device info: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_maintenance_history(request, device_id):
    """
    Get maintenance history for a device.
    """
    try:
        device = Device.objects.get(pk=device_id)
        
        maintenance_records = Maintenance.objects.filter(
            device=device
        ).select_related('vendor', 'created_by').order_by('-start_date')
        
        history = []
        for maintenance in maintenance_records:
            history.append({
                'id': maintenance.id,
                'maintenance_id': maintenance.maintenance_id,
                'type': maintenance.get_maintenance_type_display(),
                'status': maintenance.get_status_display(),
                'priority': maintenance.get_priority_display(),
                'start_date': maintenance.start_date.isoformat() if maintenance.start_date else None,
                'end_date': maintenance.actual_end_date.isoformat() if maintenance.actual_end_date else maintenance.expected_end_date.isoformat() if maintenance.expected_end_date else None,
                'cost': float(maintenance.actual_cost) if maintenance.actual_cost else float(maintenance.estimated_cost) if maintenance.estimated_cost else 0,
                'vendor': maintenance.vendor.name if maintenance.vendor else 'Internal',
                'result': maintenance.get_result_display() if maintenance.result else None,
                'satisfaction': maintenance.satisfaction_rating,
                'created_by': maintenance.created_by.get_full_name() if maintenance.created_by else None,
                'description': maintenance.description[:100] + '...' if len(maintenance.description) > 100 else maintenance.description,
                'url': reverse('maintenance:detail', kwargs={'pk': maintenance.pk})
            })
        
        return JsonResponse({'history': history})
        
    except Device.DoesNotExist:
        return JsonResponse({'error': 'Device not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving maintenance history: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def get_cost_estimate(request):
    """
    Get cost estimate based on maintenance parameters.
    """
    try:
        data = json.loads(request.body)
        device_id = data.get('device_id')
        maintenance_type = data.get('maintenance_type')
        priority = data.get('priority')
        description = data.get('description', '')
        
        if not all([device_id, maintenance_type, priority]):
            return JsonResponse({'error': 'Missing required parameters'}, status=400)
        
        device = Device.objects.get(pk=device_id)
        
        # Base cost estimates
        base_costs = {
            'PREVENTIVE': 3000,
            'CORRECTIVE': 7000,
            'EMERGENCY': 15000,
            'UPGRADE': 12000,
            'INSPECTION': 1500,
            'CLEANING': 1000,
            'CALIBRATION': 2500,
            'REPLACEMENT': 10000,
            'WARRANTY': 0,
            'OTHER': 4000
        }
        
        # Priority multipliers
        priority_multipliers = {
            'LOW': 0.8,
            'MEDIUM': 1.0,
            'HIGH': 1.3,
            'CRITICAL': 1.6,
            'EMERGENCY': 2.0
        }
        
        # Device category multipliers (based on complexity)
        category_multipliers = {
            'COMPUTER': 1.2,
            'NETWORK': 1.5,
            'SERVER': 2.0,
            'SECURITY': 1.4,
            'COMMUNICATION': 1.1,
            'FURNITURE': 0.5,
            'OTHER': 1.0
        }
        
        base_cost = base_costs.get(maintenance_type, 4000)
        priority_mult = priority_multipliers.get(priority, 1.0)
        
        # Get category multiplier
        category_code = device.subcategory.category.code if device.subcategory else 'OTHER'
        category_mult = category_multipliers.get(category_code, 1.0)
        
        # Calculate estimate
        estimated_cost = base_cost * priority_mult * category_mult
        
        # Check for keywords in description that might affect cost
        description_lower = description.lower()
        if any(keyword in description_lower for keyword in ['replace', 'replacement', 'new']):
            estimated_cost *= 1.5
        elif any(keyword in description_lower for keyword in ['upgrade', 'improve']):
            estimated_cost *= 1.3
        elif any(keyword in description_lower for keyword in ['clean', 'dust', 'maintenance']):
            estimated_cost *= 0.7
        
        # Get historical data for better estimation
        similar_maintenance = Maintenance.objects.filter(
            device__subcategory=device.subcategory,
            maintenance_type=maintenance_type,
            priority=priority,
            actual_cost__isnull=False
        )
        
        if similar_maintenance.exists():
            historical_avg = similar_maintenance.aggregate(avg=Avg('actual_cost'))['avg']
            # Blend estimated and historical costs
            final_estimate = (estimated_cost + historical_avg) / 2
            confidence = 'high'
        else:
            final_estimate = estimated_cost
            confidence = 'medium'
        
        # Round to nearest 100
        final_estimate = round(final_estimate / 100) * 100
        
        return JsonResponse({
            'estimated_cost': final_estimate,
            'base_cost': base_cost,
            'priority_multiplier': priority_mult,
            'category_multiplier': category_mult,
            'confidence': confidence,
            'factors': {
                'maintenance_type': maintenance_type,
                'priority': priority,
                'device_category': category_code,
                'historical_data': similar_maintenance.count() > 0
            }
        })
        
    except Device.DoesNotExist:
        return JsonResponse({'error': 'Device not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error calculating estimate: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def update_maintenance_status(request, pk):
    """
    Update maintenance status via AJAX.
    """
    if not request.user.has_perm('maintenance.change_maintenance'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        maintenance = get_object_or_404(Maintenance, pk=pk)
        data = json.loads(request.body)
        new_status = data.get('status')
        notes = data.get('notes', '')
        
        if not new_status:
            return JsonResponse({'error': 'Status is required'}, status=400)
        
        # Validate status transition
        if not maintenance.can_transition_to(new_status):
            return JsonResponse({
                'error': f'Invalid status transition from {maintenance.status} to {new_status}'
            }, status=400)
        
        old_status = maintenance.status
        maintenance.status = new_status
        
        # Set appropriate timestamps
        if new_status == 'IN_PROGRESS' and not maintenance.actual_start_date:
            maintenance.actual_start_date = timezone.now()
        elif new_status == 'COMPLETED' and not maintenance.actual_end_date:
            maintenance.actual_end_date = timezone.now()
            if not maintenance.result:
                maintenance.result = 'SUCCESS'
        
        # Add notes if provided
        if notes:
            if maintenance.internal_notes:
                maintenance.internal_notes += f"\n\n[Status Update {timezone.now().strftime('%Y-%m-%d %H:%M')}]: {notes}"
            else:
                maintenance.internal_notes = f"[Status Update {timezone.now().strftime('%Y-%m-%d %H:%M')}]: {notes}"
        
        maintenance.updated_by = request.user
        maintenance.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Status updated from {old_status} to {new_status}',
            'maintenance': {
                'id': maintenance.id,
                'maintenance_id': maintenance.maintenance_id,
                'status': maintenance.get_status_display(),
                'status_badge_class': maintenance.get_status_badge_class(),
                'progress_percentage': maintenance.progress_percentage,
                'actual_start_date': maintenance.actual_start_date.isoformat() if maintenance.actual_start_date else None,
                'actual_end_date': maintenance.actual_end_date.isoformat() if maintenance.actual_end_date else None,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error updating status: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_maintenance_progress(request, pk):
    """
    Get maintenance progress information.
    """
    try:
        maintenance = get_object_or_404(Maintenance, pk=pk)
        
        progress_data = {
            'maintenance_id': maintenance.maintenance_id,
            'status': maintenance.get_status_display(),
            'status_code': maintenance.status,
            'progress_percentage': maintenance.progress_percentage,
            'is_overdue': maintenance.is_overdue,
            'days_overdue': maintenance.days_overdue if maintenance.is_overdue else 0,
            'days_until_due': maintenance.days_until_due,
            'planned_duration': maintenance.planned_duration,
            'actual_duration': maintenance.actual_duration,
            'duration_variance': maintenance.duration_variance,
            'start_date': maintenance.start_date.isoformat() if maintenance.start_date else None,
            'expected_end_date': maintenance.expected_end_date.isoformat() if maintenance.expected_end_date else None,
            'actual_start_date': maintenance.actual_start_date.isoformat() if maintenance.actual_start_date else None,
            'actual_end_date': maintenance.actual_end_date.isoformat() if maintenance.actual_end_date else None,
            'cost_info': {
                'estimated_cost': float(maintenance.estimated_cost) if maintenance.estimated_cost else 0,
                'actual_cost': float(maintenance.actual_cost) if maintenance.actual_cost else 0,
                'cost_variance': float(maintenance.cost_variance),
                'cost_variance_percentage': float(maintenance.cost_variance_percentage),
            },
            'device_info': {
                'device_id': maintenance.device.device_id,
                'brand': maintenance.device.brand,
                'model': maintenance.device.model,
                'status': maintenance.device.get_status_display(),
            },
            'vendor_info': {
                'name': maintenance.vendor.name if maintenance.vendor else None,
                'contact': maintenance.vendor.contact_person if maintenance.vendor else None,
                'phone': maintenance.vendor.phone if maintenance.vendor else None,
            } if maintenance.vendor else None,
            'last_updated': maintenance.updated_at.isoformat() if maintenance.updated_at else None,
        }
        
        return JsonResponse({'progress': progress_data})
        
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving progress: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_calendar_events(request):
    """
    Get maintenance events for calendar display.
    """
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    if not start_date or not end_date:
        return JsonResponse({'error': 'Start and end dates are required'}, status=400)
    
    try:
        maintenance_records = Maintenance.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date
        ).select_related('device', 'vendor')
        
        events = []
        for maintenance in maintenance_records:
            # Determine event color based on status and priority
            color_map = {
                'SCHEDULED': '#3b82f6',  # Blue
                'IN_PROGRESS': '#f59e0b',  # Amber
                'ON_HOLD': '#6b7280',  # Gray
                'COMPLETED': '#10b981',  # Green
                'CANCELLED': '#374151',  # Dark gray
                'FAILED': '#ef4444',  # Red
            }
            
            if maintenance.priority in ['CRITICAL', 'EMERGENCY']:
                color = '#dc2626'  # Red for critical
            else:
                color = color_map.get(maintenance.status, '#6b7280')
            
            events.append({
                'id': maintenance.id,
                'title': f"{maintenance.maintenance_id} - {maintenance.device.device_id}",
                'start': maintenance.start_date.isoformat(),
                'end': maintenance.expected_end_date.isoformat() if maintenance.expected_end_date else maintenance.start_date.isoformat(),
                'color': color,
                'borderColor': color,
                'textColor': '#ffffff',
                'extendedProps': {
                    'maintenance_id': maintenance.maintenance_id,
                    'device': f"{maintenance.device.device_id} - {maintenance.device.brand} {maintenance.device.model}",
                    'type': maintenance.get_maintenance_type_display(),
                    'status': maintenance.get_status_display(),
                    'priority': maintenance.get_priority_display(),
                    'vendor': maintenance.vendor.name if maintenance.vendor else 'Internal',
                    'estimated_cost': float(maintenance.estimated_cost) if maintenance.estimated_cost else 0,
                    'description': maintenance.description[:100] + '...' if len(maintenance.description) > 100 else maintenance.description,
                    'url': reverse('maintenance:detail', kwargs={'pk': maintenance.pk})
                }
            })
        
        return JsonResponse({'events': events})
        
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving calendar events: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_month_events(request, year, month):
    """
    Get maintenance events for a specific month.
    """
    try:
        # Calculate month boundaries
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        maintenance_records = Maintenance.objects.filter(
            start_date__gte=start_date,
            start_date__lte=end_date
        ).select_related('device', 'vendor').order_by('start_date')
        
        # Group by date
        events_by_date = {}
        for maintenance in maintenance_records:
            date_key = maintenance.start_date.isoformat()
            if date_key not in events_by_date:
                events_by_date[date_key] = []
            
            events_by_date[date_key].append({
                'id': maintenance.id,
                'maintenance_id': maintenance.maintenance_id,
                'device': maintenance.device.device_id,
                'type': maintenance.get_maintenance_type_display(),
                'status': maintenance.get_status_display(),
                'priority': maintenance.get_priority_display(),
                'status_badge_class': maintenance.get_status_badge_class(),
                'priority_badge_class': maintenance.get_priority_badge_class(),
                'url': reverse('maintenance:detail', kwargs={'pk': maintenance.pk})
            })
        
        return JsonResponse({
            'events_by_date': events_by_date,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'total_events': maintenance_records.count()
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving month events: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_dashboard_stats(request):
    """
    Get real-time dashboard statistics.
    """
    try:
        # Current statistics
        all_maintenance = Maintenance.objects.all()
        today = timezone.now().date()
        
        stats = {
            'total': all_maintenance.count(),
            'scheduled': all_maintenance.filter(status='SCHEDULED').count(),
            'in_progress': all_maintenance.filter(status='IN_PROGRESS').count(),
            'completed': all_maintenance.filter(status='COMPLETED').count(),
            'overdue': len([m for m in all_maintenance if m.is_overdue]),
            'due_today': all_maintenance.filter(start_date=today, status='SCHEDULED').count(),
            'due_this_week': all_maintenance.filter(
                start_date__gte=today,
                start_date__lte=today + timedelta(days=7),
                status='SCHEDULED'
            ).count(),
            'completed_this_month': all_maintenance.filter(
                status='COMPLETED',
                actual_end_date__month=today.month,
                actual_end_date__year=today.year
            ).count(),
            'cost_this_month': float(all_maintenance.filter(
                actual_end_date__month=today.month,
                actual_end_date__year=today.year
            ).aggregate(total=Sum('actual_cost'))['total'] or 0),
        }
        
        # Performance metrics
        completed_maintenance = all_maintenance.filter(status='COMPLETED')
        if completed_maintenance.exists():
            stats['performance'] = {
                'completion_rate': completed_maintenance.count() / all_maintenance.count() * 100,
                'avg_satisfaction': float(completed_maintenance.aggregate(
                    avg=Avg('satisfaction_rating')
                )['avg'] or 0),
                'on_time_rate': completed_maintenance.filter(
                    actual_end_date__lte=F('expected_end_date')
                ).count() / completed_maintenance.count() * 100
            }
        else:
            stats['performance'] = {
                'completion_rate': 0,
                'avg_satisfaction': 0,
                'on_time_rate': 0
            }
        
        # Recent activity
        recent_maintenance = all_maintenance.select_related('device', 'vendor').order_by('-created_at')[:5]
        stats['recent_activity'] = []
        
        for maintenance in recent_maintenance:
            stats['recent_activity'].append({
                'id': maintenance.id,
                'maintenance_id': maintenance.maintenance_id,
                'device': maintenance.device.device_id,
                'status': maintenance.get_status_display(),
                'created_at': maintenance.created_at.isoformat(),
                'url': reverse('maintenance:detail', kwargs={'pk': maintenance.pk})
            })
        
        return JsonResponse({'stats': stats})
        
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving dashboard stats: {str(e)}'}, status=500)


@login_required
@require_http_methods(["GET"])
def get_overdue_count(request):
    """
    Get count of overdue maintenance for notifications.
    """
    try:
        all_maintenance = Maintenance.objects.filter(
            status__in=['SCHEDULED', 'IN_PROGRESS'],
            is_active=True
        )
        
        overdue_count = len([m for m in all_maintenance if m.is_overdue])
        
        # Get critical overdue items
        critical_overdue = []
        for maintenance in all_maintenance:
            if maintenance.is_overdue and maintenance.priority in ['CRITICAL', 'EMERGENCY']:
                critical_overdue.append({
                    'id': maintenance.id,
                    'maintenance_id': maintenance.maintenance_id,
                    'device': maintenance.device.device_id,
                    'days_overdue': maintenance.days_overdue,
                    'priority': maintenance.get_priority_display(),
                    'url': reverse('maintenance:detail', kwargs={'pk': maintenance.pk})
                })
        
        return JsonResponse({
            'overdue_count': overdue_count,
            'critical_overdue': critical_overdue,
            'has_critical': len(critical_overdue) > 0
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving overdue count: {str(e)}'}, status=500)

# ============================================================================
# ESSENTIAL AJAX Endpoints
# ============================================================================

@login_required
@require_http_methods(["POST"])
def quick_status_update(request, pk):
    """
    Quick status update for maintenance records.
    """
    if not request.user.has_perm('maintenance.change_maintenance'):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        maintenance = get_object_or_404(Maintenance, pk=pk)
        new_status = request.POST.get('status')
        
        if not new_status:
            return JsonResponse({'error': 'Status is required'}, status=400)
        
        old_status = maintenance.status
        maintenance.status = new_status
        
        # Set timestamps
        if new_status == 'IN_PROGRESS' and not maintenance.actual_start_date:
            maintenance.actual_start_date = timezone.now()
        elif new_status == 'COMPLETED' and not maintenance.actual_end_date:
            maintenance.actual_end_date = timezone.now()
        
        maintenance.updated_by = request.user
        maintenance.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Status updated from {old_status} to {new_status}',
            'new_status': maintenance.get_status_display()
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def device_maintenance_check(request):
    """
    Check if device can have maintenance scheduled.
    """
    device_id = request.GET.get('device_id')
    
    if not device_id:
        return JsonResponse({'valid': False, 'message': 'Device ID required'})
    
    try:
        device = Device.objects.get(pk=device_id)
        
        # Check for active maintenance
        active_maintenance = Maintenance.objects.filter(
            device=device,
            status__in=['SCHEDULED', 'IN_PROGRESS'],
            is_active=True
        ).first()
        
        if active_maintenance:
            return JsonResponse({
                'valid': False,
                'message': f'Device has active maintenance: {active_maintenance.maintenance_id}'
            })
        
        return JsonResponse({
            'valid': True,
            'device_info': {
                'id': device.device_id,
                'brand': device.brand,
                'model': device.model
            }
        })
        
    except Device.DoesNotExist:
        return JsonResponse({'valid': False, 'message': 'Device not found'})


@login_required  
@require_http_methods(["GET"])
def simple_cost_estimate(request):
    """
    Simple cost estimation based on maintenance type.
    """
    maintenance_type = request.GET.get('type')
    
    # Simple cost estimates
    cost_map = {
        'PREVENTIVE': 3000,
        'CORRECTIVE': 5000,
        'EMERGENCY': 10000,
        'UPGRADE': 8000,
        'INSPECTION': 1500,
        'CLEANING': 1000,
        'CALIBRATION': 2500,
        'REPLACEMENT': 7000,
        'WARRANTY': 0,
        'OTHER': 3000
    }
    
    estimated_cost = cost_map.get(maintenance_type, 3000)
    
    return JsonResponse({
        'estimated_cost': estimated_cost,
        'maintenance_type': maintenance_type
    })