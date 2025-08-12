"""
Views for Locations app in PIMS (Parliament IT Inventory Management System)
Bangladesh Parliament Secretariat

This module defines views for location management including CRUD operations
for buildings, floors, blocks, rooms, offices, and comprehensive locations.

"""
# Standard library imports
import csv
import io
import json
import os
import zipfile
from datetime import datetime, timedelta
import calendar

# Django core imports
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Q, Count, Sum, Avg, Max, Min
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View


from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)

# Third-party imports
import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)

# Local app imports
from .models import (
    Building, Floor, Block, Room, Office, Location, LocationQRCode
)
from .forms import (
    BuildingForm, FloorForm, BlockForm, RoomForm, OfficeForm, 
    LocationForm, LocationSearchForm, CoordinateInputForm
)

from pims.utils.qr_code import (
    create_location_qr_code, 
    bulk_generate_location_qr_codes,
    get_qr_code_for_location
)

# Additional imports for specific functionalities
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ValidationError, PermissionDenied


# ============================================================================
# Building Management Views
# ============================================================================

class BuildingListView(LoginRequiredMixin, ListView):
    """Display list of all buildings with search and filtering."""
    model = Building
    template_name = 'locations/building_list.html'
    context_object_name = 'buildings'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter buildings based on search parameters."""
        queryset = Building.objects.all().order_by('code', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['status'] = self.request.GET.get('status', '')
        context['total_buildings'] = Building.objects.count()
        context['active_buildings'] = Building.objects.filter(is_active=True).count()
        context['inactive_buildings'] = Building.objects.filter(is_active=False).count()
        return context


class BuildingCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new building."""
    model = Building
    form_class = BuildingForm
    template_name = 'locations/building_create.html'
    permission_required = 'locations.add_building'
    success_url = reverse_lazy('locations:building_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Building "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Building'
        context['form_action'] = 'Create'
        return context


class BuildingDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a building."""
    model = Building
    template_name = 'locations/building_detail.html'
    context_object_name = 'building'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        building = self.object
        
        # Get related locations for this building
        context['related_locations'] = Location.objects.filter(
            building=building
        ).select_related('floor', 'block', 'room', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class BuildingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing building."""
    model = Building
    form_class = BuildingForm
    template_name = 'locations/building_edit.html'
    permission_required = 'locations.change_building'
    
    def get_success_url(self):
        return reverse('locations:building_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Building "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Building - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class BuildingDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a building."""
    model = Building
    template_name = 'locations/building_delete.html'
    permission_required = 'locations.delete_building'
    success_url = reverse_lazy('locations:building_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        building = self.object
        context['related_locations_count'] = Location.objects.filter(building=building).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle building deletion."""
        building = self.get_object()
        building_name = building.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Building "{building_name}" deleted successfully!')
        return response


class BuildingToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle building active/inactive status."""
    permission_required = 'locations.change_building'
    
    def post(self, request, pk):
        """Toggle building status."""
        building = get_object_or_404(Building, pk=pk)
        building.is_active = not building.is_active
        building.save()
        
        status = "activated" if building.is_active else "deactivated"
        messages.success(request, f'Building "{building.name}" {status} successfully!')
        
        return redirect('locations:building_detail', pk=pk)


# ============================================================================
# Floor Management Views
# ============================================================================

class FloorListView(LoginRequiredMixin, ListView):
    """Display list of all floors with search and filtering."""
    model = Floor
    template_name = 'locations/floor_list.html'
    context_object_name = 'floors'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter floors based on search parameters."""
        queryset = Floor.objects.all().order_by('floor_number', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['status'] = self.request.GET.get('status', '')
        context['total_floors'] = Floor.objects.count()
        context['active_floors'] = Floor.objects.filter(is_active=True).count()
        context['inactive_floors'] = Floor.objects.filter(is_active=False).count()
        return context


class FloorCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new floor."""
    model = Floor
    form_class = FloorForm
    template_name = 'locations/floor_create.html'
    permission_required = 'locations.add_floor'
    success_url = reverse_lazy('locations:floor_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Floor "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Floor'
        context['form_action'] = 'Create'
        return context


class FloorDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a floor."""
    model = Floor
    template_name = 'locations/floor_detail.html'
    context_object_name = 'floor'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        floor = self.object
        
        # Get related locations for this floor
        context['related_locations'] = Location.objects.filter(
            floor=floor
        ).select_related('building', 'block', 'room', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class FloorUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing floor."""
    model = Floor
    form_class = FloorForm
    template_name = 'locations/floor_edit.html'
    permission_required = 'locations.change_floor'
    
    def get_success_url(self):
        return reverse('locations:floor_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Floor "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Floor - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class FloorDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a floor."""
    model = Floor
    template_name = 'locations/floor_delete.html'
    permission_required = 'locations.delete_floor'
    success_url = reverse_lazy('locations:floor_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        floor = self.object
        context['related_locations_count'] = Location.objects.filter(floor=floor).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle floor deletion."""
        floor = self.get_object()
        floor_name = floor.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Floor "{floor_name}" deleted successfully!')
        return response


class FloorToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle floor active/inactive status."""
    permission_required = 'locations.change_floor'
    
    def post(self, request, pk):
        """Toggle floor status."""
        floor = get_object_or_404(Floor, pk=pk)
        floor.is_active = not floor.is_active
        floor.save()
        
        status = "activated" if floor.is_active else "deactivated"
        messages.success(request, f'Floor "{floor.name}" {status} successfully!')
        
        return redirect('locations:floor_detail', pk=pk)


# ============================================================================
# Block Management Views
# ============================================================================

class BlockListView(LoginRequiredMixin, ListView):
    """Display list of all blocks with search and filtering."""
    model = Block
    template_name = 'locations/block_list.html'
    context_object_name = 'blocks'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter blocks based on search parameters."""
        queryset = Block.objects.all().order_by('code', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['status'] = self.request.GET.get('status', '')
        context['total_blocks'] = Block.objects.count()
        context['active_blocks'] = Block.objects.filter(is_active=True).count()
        context['inactive_blocks'] = Block.objects.filter(is_active=False).count()
        return context


class BlockCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new block."""
    model = Block
    form_class = BlockForm
    template_name = 'locations/block_create.html'
    permission_required = 'locations.add_block'
    success_url = reverse_lazy('locations:block_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Block "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Block'
        context['form_action'] = 'Create'
        return context


class BlockDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a block."""
    model = Block
    template_name = 'locations/block_detail.html'
    context_object_name = 'block'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        block = self.object
        
        # Get related locations for this block
        context['related_locations'] = Location.objects.filter(
            block=block
        ).select_related('building', 'floor', 'room', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class BlockUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing block."""
    model = Block
    form_class = BlockForm
    template_name = 'locations/block_edit.html'
    permission_required = 'locations.change_block'
    
    def get_success_url(self):
        return reverse('locations:block_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Block "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Block - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class BlockDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a block."""
    model = Block
    template_name = 'locations/block_delete.html'
    permission_required = 'locations.delete_block'
    success_url = reverse_lazy('locations:block_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        block = self.object
        context['related_locations_count'] = Location.objects.filter(block=block).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle block deletion."""
        block = self.get_object()
        block_name = block.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Block "{block_name}" deleted successfully!')
        return response


class BlockToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle block active/inactive status."""
    permission_required = 'locations.change_block'
    
    def post(self, request, pk):
        """Toggle block status."""
        block = get_object_or_404(Block, pk=pk)
        block.is_active = not block.is_active
        block.save()
        
        status = "activated" if block.is_active else "deactivated"
        messages.success(request, f'Block "{block.name}" {status} successfully!')
        
        return redirect('locations:block_detail', pk=pk)


# ============================================================================
# Room Management Views
# ============================================================================

class RoomListView(LoginRequiredMixin, ListView):
    """Display list of all rooms with search and filtering."""
    model = Room
    template_name = 'locations/room_list.html'
    context_object_name = 'rooms'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter rooms based on search parameters."""
        queryset = Room.objects.all().order_by('room_number', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(room_number__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Room type filter
        room_type = self.request.GET.get('room_type')
        if room_type:
            queryset = queryset.filter(room_type=room_type)
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['room_type'] = self.request.GET.get('room_type', '')
        context['status'] = self.request.GET.get('status', '')
        context['room_types'] = Room.ROOM_TYPES
        context['total_rooms'] = Room.objects.count()
        context['active_rooms'] = Room.objects.filter(is_active=True).count()
        context['inactive_rooms'] = Room.objects.filter(is_active=False).count()
        return context


class RoomCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new room."""
    model = Room
    form_class = RoomForm
    template_name = 'locations/room_create.html'
    permission_required = 'locations.add_room'
    success_url = reverse_lazy('locations:room_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Room "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Room'
        context['form_action'] = 'Create'
        return context


class RoomDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a room."""
    model = Room
    template_name = 'locations/room_detail.html'
    context_object_name = 'room'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        room = self.object
        
        # Get related locations for this room
        context['related_locations'] = Location.objects.filter(
            room=room
        ).select_related('building', 'floor', 'block', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class RoomUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing room."""
    model = Room
    form_class = RoomForm
    template_name = 'locations/room_edit.html'
    permission_required = 'locations.change_room'
    
    def get_success_url(self):
        return reverse('locations:room_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Room "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Room - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class RoomDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a room."""
    model = Room
    template_name = 'locations/room_delete.html'
    permission_required = 'locations.delete_room'
    success_url = reverse_lazy('locations:room_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        room = self.object
        context['related_locations_count'] = Location.objects.filter(room=room).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle room deletion."""
        room = self.get_object()
        room_name = room.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Room "{room_name}" deleted successfully!')
        return response


class RoomToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle room active/inactive status."""
    permission_required = 'locations.change_room'
    
    def post(self, request, pk):
        """Toggle room status."""
        room = get_object_or_404(Room, pk=pk)
        room.is_active = not room.is_active
        room.save()
        
        status = "activated" if room.is_active else "deactivated"
        messages.success(request, f'Room "{room.name}" {status} successfully!')
        
        return redirect('locations:room_detail', pk=pk)


class RoomByTypeView(LoginRequiredMixin, ListView):
    """Display rooms filtered by type."""
    model = Room
    template_name = 'locations/room_by_type.html'
    context_object_name = 'rooms'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter rooms by type."""
        room_type = self.kwargs.get('room_type')
        return Room.objects.filter(
            room_type=room_type,
            is_active=True
        ).order_by('room_number', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        room_type = self.kwargs.get('room_type')
        context['room_type'] = room_type
        context['room_type_display'] = dict(Room.ROOM_TYPES).get(room_type, room_type)
        context['total_rooms'] = self.get_queryset().count()
        return context


# ============================================================================
# Office Management Views
# ============================================================================

class OfficeListView(LoginRequiredMixin, ListView):
    """Display list of all offices with search and filtering."""
    model = Office
    template_name = 'locations/office_list.html'
    context_object_name = 'offices'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter offices based on search parameters."""
        queryset = Office.objects.all().order_by('office_code', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(office_code__icontains=search) |
                Q(head_of_office__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Office type filter
        office_type = self.request.GET.get('office_type')
        if office_type:
            queryset = queryset.filter(office_type=office_type)
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['office_type'] = self.request.GET.get('office_type', '')
        context['status'] = self.request.GET.get('status', '')
        context['office_types'] = Office.OFFICE_TYPES
        context['total_offices'] = Office.objects.count()
        context['active_offices'] = Office.objects.filter(is_active=True).count()
        context['inactive_offices'] = Office.objects.filter(is_active=False).count()
        return context


class OfficeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new office."""
    model = Office
    form_class = OfficeForm
    template_name = 'locations/office_create.html'
    permission_required = 'locations.add_office'
    success_url = reverse_lazy('locations:office_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Office "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Office'
        context['form_action'] = 'Create'
        return context


class OfficeDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about an office."""
    model = Office
    template_name = 'locations/office_detail.html'
    context_object_name = 'office'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        office = self.object
        
        # Get related locations for this office
        context['related_locations'] = Location.objects.filter(
            office=office
        ).select_related('building', 'floor', 'block', 'room')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class OfficeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing office."""
    model = Office
    form_class = OfficeForm
    template_name = 'locations/office_edit.html'
    permission_required = 'locations.change_office'
    
    def get_success_url(self):
        return reverse('locations:office_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Office "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Office - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class OfficeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete an office."""
    model = Office
    template_name = 'locations/office_delete.html'
    permission_required = 'locations.delete_office'
    success_url = reverse_lazy('locations:office_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        office = self.object
        context['related_locations_count'] = Location.objects.filter(office=office).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle office deletion."""
        office = self.get_object()
        office_name = office.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Office "{office_name}" deleted successfully!')
        return response


class OfficeToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle office active/inactive status."""
    permission_required = 'locations.change_office'
    
    def post(self, request, pk):
        """Toggle office status."""
        office = get_object_or_404(Office, pk=pk)
        office.is_active = not office.is_active
        office.save()
        
        status = "activated" if office.is_active else "deactivated"
        messages.success(request, f'Office "{office.name}" {status} successfully!')
        
        return redirect('locations:office_detail', pk=pk)


class OfficeByTypeView(LoginRequiredMixin, ListView):
    """Display offices filtered by type."""
    model = Office
    template_name = 'locations/office_by_type.html'
    context_object_name = 'offices'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter offices by type."""
        office_type = self.kwargs.get('office_type')
        return Office.objects.filter(
            office_type=office_type,
            is_active=True
        ).order_by('office_code', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        office_type = self.kwargs.get('office_type')
        context['office_type'] = office_type
        context['office_type_display'] = dict(Office.OFFICE_TYPES).get(office_type, office_type)
        context['total_offices'] = self.get_queryset().count()
        
        # Add office type statistics
        context['office_stats'] = {
            'speaker': Office.objects.filter(office_type='speaker', is_active=True).count(),
            'secretary': Office.objects.filter(office_type='secretary', is_active=True).count(),
            'mp': Office.objects.filter(office_type='mp', is_active=True).count(),
            'wing': Office.objects.filter(office_type='wing', is_active=True).count(),
            'branch': Office.objects.filter(office_type='branch', is_active=True).count(),
        }
        
        return context

# ============================================================================
# Location Management Views - Main Entity
# ============================================================================

class LocationListView(LoginRequiredMixin, ListView):
    """Display list of all locations with comprehensive search and filtering."""
    model = Location
    template_name = 'locations/location_list.html'
    context_object_name = 'locations'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter locations based on search parameters."""
        queryset = Location.objects.select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code', 'name')
        
        # Apply search form filters
        form = LocationSearchForm(self.request.GET)
        if form.is_valid():
            queryset = form.filter_queryset(queryset)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search_form'] = LocationSearchForm(self.request.GET)
        context['total_locations'] = Location.objects.count()
        context['active_locations'] = Location.objects.filter(is_active=True).count()
        context['inactive_locations'] = Location.objects.filter(is_active=False).count()
        context['locations_with_coordinates'] = Location.objects.filter(
            latitude__isnull=False, longitude__isnull=False
        ).count()
        
        # Component counts
        context['component_counts'] = {
            'buildings': Building.objects.filter(is_active=True).count(),
            'floors': Floor.objects.filter(is_active=True).count(),
            'blocks': Block.objects.filter(is_active=True).count(),
            'rooms': Room.objects.filter(is_active=True).count(),
            'offices': Office.objects.filter(is_active=True).count(),
        }
        
        return context


class LocationCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new location."""
    model = Location
    form_class = LocationForm
    template_name = 'locations/location_create.html'
    permission_required = 'locations.add_location'
    success_url = reverse_lazy('locations:list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Location "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Location'
        context['form_action'] = 'Create'
        
        # Add component data for dynamic forms
        context['components'] = {
            'buildings': Building.objects.filter(is_active=True),
            'floors': Floor.objects.filter(is_active=True),
            'blocks': Block.objects.filter(is_active=True),
            'rooms': Room.objects.filter(is_active=True),
            'offices': Office.objects.filter(is_active=True),
        }
        
        return context


class LocationDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a location."""
    model = Location
    template_name = 'locations/location_detail.html'
    context_object_name = 'location'
    
    def get_queryset(self):
        """Optimize queryset with related data."""
        return Location.objects.select_related(
            'building', 'floor', 'block', 'room', 'office'
        )
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        location = self.object
        
        # Get location components
        context['components'] = location.get_related_components()
        context['has_coordinates'] = location.has_coordinates()
        context['coordinate_string'] = location.coordinate_string
        context['full_description'] = location.get_full_location_description()
        
        # Find related assignments/devices (if implemented)
        # context['related_assignments'] = Assignment.objects.filter(location=location)
        
        return context


class LocationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing location."""
    model = Location
    form_class = LocationForm
    template_name = 'locations/location_edit.html'
    permission_required = 'locations.change_location'
    
    def get_success_url(self):
        return reverse('locations:detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Location "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Location - {self.object.name}'
        context['form_action'] = 'Update'
        
        # Add component data for dynamic forms
        context['components'] = {
            'buildings': Building.objects.filter(is_active=True),
            'floors': Floor.objects.filter(is_active=True),
            'blocks': Block.objects.filter(is_active=True),
            'rooms': Room.objects.filter(is_active=True),
            'offices': Office.objects.filter(is_active=True),
        }
        
        return context


class LocationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a location."""
    model = Location
    template_name = 'locations/location_delete.html'
    permission_required = 'locations.delete_location'
    success_url = reverse_lazy('locations:list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        location = self.object
        
        # Check for related data that would be affected
        # context['related_assignments_count'] = Assignment.objects.filter(location=location).count()
        # context['related_devices_count'] = Device.objects.filter(location=location).count()
        
        context['components'] = location.get_related_components()
        context['full_description'] = location.get_full_location_description()
        
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle location deletion."""
        location = self.get_object()
        location_name = location.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Location "{location_name}" deleted successfully!')
        return response


class LocationSearchView(LoginRequiredMixin, ListView):
    """Advanced location search with filters."""
    model = Location
    template_name = 'locations/location_search.html'
    context_object_name = 'locations'
    paginate_by = 20
    
    def get_queryset(self):
        """Apply search filters."""
        queryset = Location.objects.select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code', 'name')
        
        # Apply search form filters
        form = LocationSearchForm(self.request.GET)
        if form.is_valid():
            queryset = form.filter_queryset(queryset)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add search context."""
        context = super().get_context_data(**kwargs)
        context['search_form'] = LocationSearchForm(self.request.GET)
        context['search_performed'] = bool(self.request.GET)
        context['total_results'] = self.get_queryset().count()
        return context


class LocationFilterView(LoginRequiredMixin, View):
    """AJAX endpoint for dynamic location filtering."""
    
    def get(self, request):
        """Return filtered locations as JSON."""
        form = LocationSearchForm(request.GET)
        
        if form.is_valid():
            queryset = Location.objects.select_related(
                'building', 'floor', 'block', 'room', 'office'
            )
            queryset = form.filter_queryset(queryset)
            
            locations = []
            for location in queryset[:50]:  # Limit results for performance
                locations.append({
                    'id': location.id,
                    'code': location.location_code,
                    'name': location.name,
                    'building': location.building.name if location.building else None,
                    'floor': location.floor.name if location.floor else None,
                    'room': location.room.name if location.room else None,
                    'office': location.office.name if location.office else None,
                    'has_coordinates': location.has_coordinates(),
                    'is_active': location.is_active,
                    'url': reverse('locations:detail', kwargs={'pk': location.pk})
                })
            
            return JsonResponse({
                'success': True,
                'locations': locations,
                'total': queryset.count()
            })
        
        return JsonResponse({
            'success': False,
            'errors': form.errors
        })


class LocationCoordinatesView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update location coordinates."""
    model = Location
    form_class = CoordinateInputForm
    template_name = 'locations/location_coordinates.html'
    permission_required = 'locations.change_location'
    
    def get_success_url(self):
        return reverse('locations:detail', kwargs={'pk': self.object.pk})
    
    def get_initial(self):
        """Set initial form data."""
        location = self.get_object()
        return {
            'latitude': location.latitude,
            'longitude': location.longitude
        }
    
    def form_valid(self, form):
        """Update location coordinates."""
        location = self.get_object()
        location.latitude = form.cleaned_data['latitude']
        location.longitude = form.cleaned_data['longitude']
        location.save()
        
        messages.success(
            self.request, 
            f'Coordinates updated for location "{location.name}"!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        """Add context data."""
        context = super().get_context_data(**kwargs)
        context['location'] = self.get_object()
        context['page_title'] = f'Update Coordinates - {self.get_object().name}'
        return context


# ============================================================================
# API Endpoints - Phase 3
# ============================================================================

class BuildingAPIView(LoginRequiredMixin, View):
    """API endpoint for building data."""
    
    def get(self, request):
        """Return buildings as JSON."""
        buildings = Building.objects.filter(is_active=True).order_by('code', 'name')
        
        # Optional filtering
        search = request.GET.get('search')
        if search:
            buildings = buildings.filter(
                Q(name__icontains=search) | Q(code__icontains=search)
            )
        
        data = []
        for building in buildings:
            data.append({
                'id': building.id,
                'code': building.code,
                'name': building.name,
                'description': building.description,
                'locations_count': Location.objects.filter(building=building).count()
            })
        
        return JsonResponse({
            'success': True,
            'buildings': data,
            'total': len(data)
        })


class FloorAPIView(LoginRequiredMixin, View):
    """API endpoint for floor data."""
    
    def get(self, request):
        """Return floors as JSON."""
        floors = Floor.objects.filter(is_active=True).order_by('floor_number', 'name')
        
        # Optional filtering
        search = request.GET.get('search')
        if search:
            floors = floors.filter(name__icontains=search)
        
        data = []
        for floor in floors:
            data.append({
                'id': floor.id,
                'name': floor.name,
                'floor_number': floor.floor_number,
                'description': floor.description,
                'locations_count': Location.objects.filter(floor=floor).count()
            })
        
        return JsonResponse({
            'success': True,
            'floors': data,
            'total': len(data)
        })


class BlockAPIView(LoginRequiredMixin, View):
    """API endpoint for block data."""
    
    def get(self, request):
        """Return blocks as JSON."""
        blocks = Block.objects.filter(is_active=True).order_by('code', 'name')
        
        # Optional filtering
        search = request.GET.get('search')
        if search:
            blocks = blocks.filter(
                Q(name__icontains=search) | Q(code__icontains=search)
            )
        
        data = []
        for block in blocks:
            data.append({
                'id': block.id,
                'code': block.code,
                'name': block.name,
                'description': block.description,
                'locations_count': Location.objects.filter(block=block).count()
            })
        
        return JsonResponse({
            'success': True,
            'blocks': data,
            'total': len(data)
        })


class RoomAPIView(LoginRequiredMixin, View):
    """API endpoint for room data."""
    
    def get(self, request):
        """Return rooms as JSON."""
        rooms = Room.objects.filter(is_active=True).order_by('room_number', 'name')
        
        # Optional filtering
        search = request.GET.get('search')
        if search:
            rooms = rooms.filter(
                Q(name__icontains=search) | Q(room_number__icontains=search)
            )
        
        room_type = request.GET.get('room_type')
        if room_type:
            rooms = rooms.filter(room_type=room_type)
        
        data = []
        for room in rooms:
            data.append({
                'id': room.id,
                'room_number': room.room_number,
                'name': room.name,
                'room_type': room.room_type,
                'room_type_display': room.get_room_type_display(),
                'capacity': room.capacity,
                'area_sqft': str(room.area_sqft) if room.area_sqft else None,
                'locations_count': Location.objects.filter(room=room).count()
            })
        
        return JsonResponse({
            'success': True,
            'rooms': data,
            'total': len(data)
        })


class OfficeAPIView(LoginRequiredMixin, View):
    """API endpoint for office data."""
    
    def get(self, request):
        """Return offices as JSON."""
        offices = Office.objects.filter(is_active=True).order_by('office_code', 'name')
        
        # Optional filtering
        search = request.GET.get('search')
        if search:
            offices = offices.filter(
                Q(name__icontains=search) | 
                Q(office_code__icontains=search) |
                Q(head_of_office__icontains=search)
            )
        
        office_type = request.GET.get('office_type')
        if office_type:
            offices = offices.filter(office_type=office_type)
        
        data = []
        for office in offices:
            data.append({
                'id': office.id,
                'office_code': office.office_code,
                'name': office.name,
                'office_type': office.office_type,
                'office_type_display': office.get_office_type_display(),
                'head_of_office': office.head_of_office,
                'contact_number': office.contact_number,
                'email': office.email,
                'locations_count': Location.objects.filter(office=office).count()
            })
        
        return JsonResponse({
            'success': True,
            'offices': data,
            'total': len(data)
        })


class LocationCoordinatesAPIView(LoginRequiredMixin, View):
    """API endpoint for location coordinates (for mapping)."""
    
    def get(self, request):
        """Return locations with coordinates as JSON."""
        locations = Location.objects.filter(
            latitude__isnull=False,
            longitude__isnull=False,
            is_active=True
        ).select_related('building', 'floor', 'block', 'room', 'office')
        
        data = []
        for location in locations:
            data.append({
                'id': location.id,
                'code': location.location_code,
                'name': location.name,
                'latitude': float(location.latitude),
                'longitude': float(location.longitude),
                'building': location.building.name if location.building else None,
                'floor': location.floor.name if location.floor else None,
                'room': location.room.name if location.room else None,
                'office': location.office.name if location.office else None,
                'description': location.get_full_location_description(),
                'url': reverse('locations:detail', kwargs={'pk': location.pk})
            })
        
        return JsonResponse({
            'success': True,
            'locations': data,
            'total': len(data)
        })


class BuildingCodeValidationView(LoginRequiredMixin, View):
    """AJAX endpoint for building code validation."""
    
    def get(self, request):
        """Check if building code exists."""
        code = request.GET.get('code', '').upper().strip()
        exclude_id = request.GET.get('exclude_id')
        
        if not code:
            return JsonResponse({
                'valid': False,
                'message': 'Code is required.'
            })
        
        # Check for existing code
        existing = Building.objects.filter(code=code)
        if exclude_id:
            existing = existing.exclude(id=exclude_id)
        
        if existing.exists():
            return JsonResponse({
                'valid': False,
                'message': f'Building code "{code}" already exists.'
            })
        
        return JsonResponse({
            'valid': True,
            'message': f'Building code "{code}" is available.'
        })


class BuildingLookupView(LoginRequiredMixin, View):
    """AJAX endpoint for building lookup by code."""
    
    def get(self, request, code):
        """Lookup building by code."""
        try:
            building = Building.objects.get(code=code.upper())
            return JsonResponse({
                'found': True,
                'building': {
                    'id': building.id,
                    'code': building.code,
                    'name': building.name,
                    'description': building.description,
                    'is_active': building.is_active,
                    'locations_count': Location.objects.filter(building=building).count(),
                    'url': reverse('locations:building_detail', kwargs={'pk': building.pk})
                }
            })
        except Building.DoesNotExist:
            return JsonResponse({
                'found': False,
                'message': f'Building with code "{code}" not found.'
            })


class LocationCodeValidationView(LoginRequiredMixin, View):
    """AJAX endpoint for location code validation."""
    
    def get(self, request):
        """Check if location code exists."""
        code = request.GET.get('code', '').upper().strip()
        exclude_id = request.GET.get('exclude_id')
        
        if not code:
            return JsonResponse({
                'valid': False,
                'message': 'Code is required.'
            })
        
        # Check for existing code
        existing = Location.objects.filter(location_code=code)
        if exclude_id:
            existing = existing.exclude(id=exclude_id)
        
        if existing.exists():
            return JsonResponse({
                'valid': False,
                'message': f'Location code "{code}" already exists.'
            })
        
        return JsonResponse({
            'valid': True,
            'message': f'Location code "{code}" is available.'
        })


class LocationLookupView(LoginRequiredMixin, View):
    """AJAX endpoint for location lookup by code."""
    
    def get(self, request, code):
        """Lookup location by code."""
        try:
            location = Location.objects.select_related(
                'building', 'floor', 'block', 'room', 'office'
            ).get(location_code=code.upper())
            
            return JsonResponse({
                'found': True,
                'location': {
                    'id': location.id,
                    'code': location.location_code,
                    'name': location.name,
                    'address': location.address,
                    'building': {
                        'id': location.building.id,
                        'name': location.building.name,
                        'code': location.building.code
                    } if location.building else None,
                    'floor': {
                        'id': location.floor.id,
                        'name': location.floor.name,
                        'number': location.floor.floor_number
                    } if location.floor else None,
                    'block': {
                        'id': location.block.id,
                        'name': location.block.name,
                        'code': location.block.code
                    } if location.block else None,
                    'room': {
                        'id': location.room.id,
                        'name': location.room.name,
                        'number': location.room.room_number,
                        'type': location.room.room_type
                    } if location.room else None,
                    'office': {
                        'id': location.office.id,
                        'name': location.office.name,
                        'code': location.office.office_code,
                        'type': location.office.office_type
                    } if location.office else None,
                    'coordinates': {
                        'latitude': float(location.latitude) if location.latitude else None,
                        'longitude': float(location.longitude) if location.longitude else None
                    },
                    'has_coordinates': location.has_coordinates(),
                    'is_active': location.is_active,
                    'description': location.get_full_location_description(),
                    'url': reverse('locations:detail', kwargs={'pk': location.pk})
                }
            })
        except Location.DoesNotExist:
            return JsonResponse({
                'found': False,
                'message': f'Location with code "{code}" not found.'
            })


class OfficeCodeValidationView(LoginRequiredMixin, View):
    """AJAX endpoint for office code validation."""
    
    def get(self, request):
        """Check if office code exists."""
        code = request.GET.get('code', '').upper().strip()
        exclude_id = request.GET.get('exclude_id')
        
        if not code:
            return JsonResponse({
                'valid': False,
                'message': 'Code is required.'
            })
        
        # Check for existing code
        existing = Office.objects.filter(office_code=code)
        if exclude_id:
            existing = existing.exclude(id=exclude_id)
        
        if existing.exists():
            return JsonResponse({
                'valid': False,
                'message': f'Office code "{code}" already exists.'
            })
        
        return JsonResponse({
            'valid': True,
            'message': f'Office code "{code}" is available.'
        })


class OfficeCodeLookupView(LoginRequiredMixin, View):
    """AJAX endpoint for office lookup by code."""
    
    def get(self, request, code):
        """Lookup office by code."""
        try:
            office = Office.objects.get(office_code=code.upper())
            return JsonResponse({
                'found': True,
                'office': {
                    'id': office.id,
                    'code': office.office_code,
                    'name': office.name,
                    'office_type': office.office_type,
                    'office_type_display': office.get_office_type_display(),
                    'head_of_office': office.head_of_office,
                    'contact_number': office.contact_number,
                    'email': office.email,
                    'description': office.description,
                    'is_active': office.is_active,
                    'locations_count': Location.objects.filter(office=office).count(),
                    'url': reverse('locations:office_detail', kwargs={'pk': office.pk})
                }
            })
        except Office.DoesNotExist:
            return JsonResponse({
                'found': False,
                'message': f'Office with code "{code}" not found.'
            })


# ============================================================================
# Helper Functions for Views
# ============================================================================

def get_location_statistics():
    """Get location statistics for dashboard and analytics."""
    return {
        'total_locations': Location.objects.count(),
        'active_locations': Location.objects.filter(is_active=True).count(),
        'locations_with_coordinates': Location.objects.filter(
            latitude__isnull=False, longitude__isnull=False
        ).count(),
        'buildings_count': Building.objects.filter(is_active=True).count(),
        'floors_count': Floor.objects.filter(is_active=True).count(),
        'blocks_count': Block.objects.filter(is_active=True).count(),
        'rooms_count': Room.objects.filter(is_active=True).count(),
        'offices_count': Office.objects.filter(is_active=True).count(),
    }


class BuildingCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new building."""
    model = Building
    form_class = BuildingForm
    template_name = 'locations/building_create.html'
    permission_required = 'locations.add_building'
    success_url = reverse_lazy('locations:building_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Building "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Building'
        context['form_action'] = 'Create'
        return context


class BuildingDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a building."""
    model = Building
    template_name = 'locations/building_detail.html'
    context_object_name = 'building'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        building = self.object
        
        # Get related locations for this building
        context['related_locations'] = Location.objects.filter(
            building=building
        ).select_related('floor', 'block', 'room', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class BuildingUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing building."""
    model = Building
    form_class = BuildingForm
    template_name = 'locations/building_edit.html'
    permission_required = 'locations.change_building'
    
    def get_success_url(self):
        return reverse('locations:building_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Building "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Building - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class BuildingDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a building."""
    model = Building
    template_name = 'locations/building_delete.html'
    permission_required = 'locations.delete_building'
    success_url = reverse_lazy('locations:building_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        building = self.object
        context['related_locations_count'] = Location.objects.filter(building=building).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle building deletion."""
        building = self.get_object()
        building_name = building.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Building "{building_name}" deleted successfully!')
        return response


class BuildingToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle building active/inactive status."""
    permission_required = 'locations.change_building'
    
    def post(self, request, pk):
        """Toggle building status."""
        building = get_object_or_404(Building, pk=pk)
        building.is_active = not building.is_active
        building.save()
        
        status = "activated" if building.is_active else "deactivated"
        messages.success(request, f'Building "{building.name}" {status} successfully!')
        
        return redirect('locations:building_detail', pk=pk)


# ============================================================================
# Floor Management Views
# ============================================================================

class FloorListView(LoginRequiredMixin, ListView):
    """Display list of all floors with search and filtering."""
    model = Floor
    template_name = 'locations/floor_list.html'
    context_object_name = 'floors'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter floors based on search parameters."""
        queryset = Floor.objects.all().order_by('floor_number', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['status'] = self.request.GET.get('status', '')
        context['total_floors'] = Floor.objects.count()
        context['active_floors'] = Floor.objects.filter(is_active=True).count()
        context['inactive_floors'] = Floor.objects.filter(is_active=False).count()
        return context


class FloorCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new floor."""
    model = Floor
    form_class = FloorForm
    template_name = 'locations/floor_create.html'
    permission_required = 'locations.add_floor'
    success_url = reverse_lazy('locations:floor_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Floor "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Floor'
        context['form_action'] = 'Create'
        return context


class FloorDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a floor."""
    model = Floor
    template_name = 'locations/floor_detail.html'
    context_object_name = 'floor'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        floor = self.object
        
        # Get related locations for this floor
        context['related_locations'] = Location.objects.filter(
            floor=floor
        ).select_related('building', 'block', 'room', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class FloorUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing floor."""
    model = Floor
    form_class = FloorForm
    template_name = 'locations/floor_edit.html'
    permission_required = 'locations.change_floor'
    
    def get_success_url(self):
        return reverse('locations:floor_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Floor "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Floor - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class FloorDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a floor."""
    model = Floor
    template_name = 'locations/floor_delete.html'
    permission_required = 'locations.delete_floor'
    success_url = reverse_lazy('locations:floor_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        floor = self.object
        context['related_locations_count'] = Location.objects.filter(floor=floor).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle floor deletion."""
        floor = self.get_object()
        floor_name = floor.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Floor "{floor_name}" deleted successfully!')
        return response


class FloorToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle floor active/inactive status."""
    permission_required = 'locations.change_floor'
    
    def post(self, request, pk):
        """Toggle floor status."""
        floor = get_object_or_404(Floor, pk=pk)
        floor.is_active = not floor.is_active
        floor.save()
        
        status = "activated" if floor.is_active else "deactivated"
        messages.success(request, f'Floor "{floor.name}" {status} successfully!')
        
        return redirect('locations:floor_detail', pk=pk)


# ============================================================================
# Block Management Views
# ============================================================================

class BlockListView(LoginRequiredMixin, ListView):
    """Display list of all blocks with search and filtering."""
    model = Block
    template_name = 'locations/block_list.html'
    context_object_name = 'blocks'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter blocks based on search parameters."""
        queryset = Block.objects.all().order_by('code', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['status'] = self.request.GET.get('status', '')
        context['total_blocks'] = Block.objects.count()
        context['active_blocks'] = Block.objects.filter(is_active=True).count()
        context['inactive_blocks'] = Block.objects.filter(is_active=False).count()
        return context


class BlockCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new block."""
    model = Block
    form_class = BlockForm
    template_name = 'locations/block_create.html'
    permission_required = 'locations.add_block'
    success_url = reverse_lazy('locations:block_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Block "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Block'
        context['form_action'] = 'Create'
        return context


class BlockDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a block."""
    model = Block
    template_name = 'locations/block_detail.html'
    context_object_name = 'block'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        block = self.object
        
        # Get related locations for this block
        context['related_locations'] = Location.objects.filter(
            block=block
        ).select_related('building', 'floor', 'room', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class BlockUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing block."""
    model = Block
    form_class = BlockForm
    template_name = 'locations/block_edit.html'
    permission_required = 'locations.change_block'
    
    def get_success_url(self):
        return reverse('locations:block_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Block "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Block - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class BlockDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a block."""
    model = Block
    template_name = 'locations/block_delete.html'
    permission_required = 'locations.delete_block'
    success_url = reverse_lazy('locations:block_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        block = self.object
        context['related_locations_count'] = Location.objects.filter(block=block).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle block deletion."""
        block = self.get_object()
        block_name = block.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Block "{block_name}" deleted successfully!')
        return response


class BlockToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle block active/inactive status."""
    permission_required = 'locations.change_block'
    
    def post(self, request, pk):
        """Toggle block status."""
        block = get_object_or_404(Block, pk=pk)
        block.is_active = not block.is_active
        block.save()
        
        status = "activated" if block.is_active else "deactivated"
        messages.success(request, f'Block "{block.name}" {status} successfully!')
        
        return redirect('locations:block_detail', pk=pk)


# ============================================================================
# Room Management Views
# ============================================================================

class RoomListView(LoginRequiredMixin, ListView):
    """Display list of all rooms with search and filtering."""
    model = Room
    template_name = 'locations/room_list.html'
    context_object_name = 'rooms'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter rooms based on search parameters."""
        queryset = Room.objects.all().order_by('room_number', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(room_number__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Room type filter
        room_type = self.request.GET.get('room_type')
        if room_type:
            queryset = queryset.filter(room_type=room_type)
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['room_type'] = self.request.GET.get('room_type', '')
        context['status'] = self.request.GET.get('status', '')
        context['room_types'] = Room.ROOM_TYPES
        context['total_rooms'] = Room.objects.count()
        context['active_rooms'] = Room.objects.filter(is_active=True).count()
        context['inactive_rooms'] = Room.objects.filter(is_active=False).count()
        return context


class RoomCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new room."""
    model = Room
    form_class = RoomForm
    template_name = 'locations/room_create.html'
    permission_required = 'locations.add_room'
    success_url = reverse_lazy('locations:room_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Room "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Room'
        context['form_action'] = 'Create'
        return context


class RoomDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about a room."""
    model = Room
    template_name = 'locations/room_detail.html'
    context_object_name = 'room'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        room = self.object
        
        # Get related locations for this room
        context['related_locations'] = Location.objects.filter(
            room=room
        ).select_related('building', 'floor', 'block', 'office')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class RoomUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing room."""
    model = Room
    form_class = RoomForm
    template_name = 'locations/room_edit.html'
    permission_required = 'locations.change_room'
    
    def get_success_url(self):
        return reverse('locations:room_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Room "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Room - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class RoomDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a room."""
    model = Room
    template_name = 'locations/room_delete.html'
    permission_required = 'locations.delete_room'
    success_url = reverse_lazy('locations:room_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        room = self.object
        context['related_locations_count'] = Location.objects.filter(room=room).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle room deletion."""
        room = self.get_object()
        room_name = room.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Room "{room_name}" deleted successfully!')
        return response


class RoomToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle room active/inactive status."""
    permission_required = 'locations.change_room'
    
    def post(self, request, pk):
        """Toggle room status."""
        room = get_object_or_404(Room, pk=pk)
        room.is_active = not room.is_active
        room.save()
        
        status = "activated" if room.is_active else "deactivated"
        messages.success(request, f'Room "{room.name}" {status} successfully!')
        
        return redirect('locations:room_detail', pk=pk)


class RoomByTypeView(LoginRequiredMixin, ListView):
    """Display rooms filtered by type."""
    model = Room
    template_name = 'locations/room_by_type.html'
    context_object_name = 'rooms'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter rooms by type."""
        room_type = self.kwargs.get('room_type')
        return Room.objects.filter(
            room_type=room_type,
            is_active=True
        ).order_by('room_number', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        room_type = self.kwargs.get('room_type')
        context['room_type'] = room_type
        context['room_type_display'] = dict(Room.ROOM_TYPES).get(room_type, room_type)
        context['total_rooms'] = self.get_queryset().count()
        return context


# ============================================================================
# Office Management Views
# ============================================================================

class OfficeListView(LoginRequiredMixin, ListView):
    """Display list of all offices with search and filtering."""
    model = Office
    template_name = 'locations/office_list.html'
    context_object_name = 'offices'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter offices based on search parameters."""
        queryset = Office.objects.all().order_by('office_code', 'name')
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(office_code__icontains=search) |
                Q(head_of_office__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Office type filter
        office_type = self.request.GET.get('office_type')
        if office_type:
            queryset = queryset.filter(office_type=office_type)
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['search'] = self.request.GET.get('search', '')
        context['office_type'] = self.request.GET.get('office_type', '')
        context['status'] = self.request.GET.get('status', '')
        context['office_types'] = Office.OFFICE_TYPES
        context['total_offices'] = Office.objects.count()
        context['active_offices'] = Office.objects.filter(is_active=True).count()
        context['inactive_offices'] = Office.objects.filter(is_active=False).count()
        return context


class OfficeCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new office."""
    model = Office
    form_class = OfficeForm
    template_name = 'locations/office_create.html'
    permission_required = 'locations.add_office'
    success_url = reverse_lazy('locations:office_list')
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Office "{self.object.name}" created successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Create New Office'
        context['form_action'] = 'Create'
        return context


class OfficeDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about an office."""
    model = Office
    template_name = 'locations/office_detail.html'
    context_object_name = 'office'
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        office = self.object
        
        # Get related locations for this office
        context['related_locations'] = Location.objects.filter(
            office=office
        ).select_related('building', 'floor', 'block', 'room')
        
        context['locations_count'] = context['related_locations'].count()
        context['active_locations_count'] = context['related_locations'].filter(is_active=True).count()
        
        return context


class OfficeUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing office."""
    model = Office
    form_class = OfficeForm
    template_name = 'locations/office_edit.html'
    permission_required = 'locations.change_office'
    
    def get_success_url(self):
        return reverse('locations:office_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        """Handle successful form submission."""
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Office "{self.object.name}" updated successfully!'
        )
        return response
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Edit Office - {self.object.name}'
        context['form_action'] = 'Update'
        return context


class OfficeDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete an office."""
    model = Office
    template_name = 'locations/office_delete.html'
    permission_required = 'locations.delete_office'
    success_url = reverse_lazy('locations:office_list')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        office = self.object
        context['related_locations_count'] = Location.objects.filter(office=office).count()
        return context
    
    def delete(self, request, *args, **kwargs):
        """Handle office deletion."""
        office = self.get_object()
        office_name = office.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Office "{office_name}" deleted successfully!')
        return response


class OfficeToggleStatusView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Toggle office active/inactive status."""
    permission_required = 'locations.change_office'
    
    def post(self, request, pk):
        """Toggle office status."""
        office = get_object_or_404(Office, pk=pk)
        office.is_active = not office.is_active
        office.save()
        
        status = "activated" if office.is_active else "deactivated"
        messages.success(request, f'Office "{office.name}" {status} successfully!')
        
        return redirect('locations:office_detail', pk=pk)


class OfficeByTypeView(LoginRequiredMixin, ListView):
    """Display offices filtered by type."""
    model = Office
    template_name = 'locations/office_by_type.html'
    context_object_name = 'offices'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter offices by type."""
        office_type = self.kwargs.get('office_type')
        return Office.objects.filter(
            office_type=office_type,
            is_active=True
        ).order_by('office_code', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        office_type = self.kwargs.get('office_type')
        context['office_type'] = office_type
        context['office_type_display'] = dict(Office.OFFICE_TYPES).get(office_type, office_type)
        context['total_offices'] = self.get_queryset().count()
        
        # Add office type statistics
        context['office_stats'] = {
            'speaker': Office.objects.filter(office_type='speaker', is_active=True).count(),
            'secretary': Office.objects.filter(office_type='secretary', is_active=True).count(),
            'mp': Office.objects.filter(office_type='mp', is_active=True).count(),
            'wing': Office.objects.filter(office_type='wing', is_active=True).count(),
            'branch': Office.objects.filter(office_type='branch', is_active=True).count(),
        }
        
        return context
    # ============================================================================
# Basic Utilities Views
# ============================================================================

class AvailableLocationsView(LoginRequiredMixin, ListView):
    """Display list of all available (active) locations."""
    model = Location
    template_name = 'locations/available_locations.html'
    context_object_name = 'locations'
    paginate_by = 25
    
    def get_queryset(self):
        """Return only active locations."""
        return Location.objects.filter(
            is_active=True
        ).select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Available Locations'
        context['total_available'] = self.get_queryset().count()
        context['locations_with_coordinates'] = self.get_queryset().filter(
            latitude__isnull=False, longitude__isnull=False
        ).count()
        
        # Component breakdown
        context['component_stats'] = {
            'with_building': self.get_queryset().filter(building__isnull=False).count(),
            'with_floor': self.get_queryset().filter(floor__isnull=False).count(),
            'with_block': self.get_queryset().filter(block__isnull=False).count(),
            'with_room': self.get_queryset().filter(room__isnull=False).count(),
            'with_office': self.get_queryset().filter(office__isnull=False).count(),
        }
        
        return context


class InactiveLocationsView(LoginRequiredMixin, ListView):
    """Display list of all inactive locations."""
    model = Location
    template_name = 'locations/inactive_locations.html'
    context_object_name = 'locations'
    paginate_by = 25
    
    def get_queryset(self):
        """Return only inactive locations."""
        return Location.objects.filter(
            is_active=False
        ).select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Inactive Locations'
        context['total_inactive'] = self.get_queryset().count()
        
        # Reason analysis (if we had a reason field)
        context['component_stats'] = {
            'with_building': self.get_queryset().filter(building__isnull=False).count(),
            'with_floor': self.get_queryset().filter(floor__isnull=False).count(),
            'with_block': self.get_queryset().filter(block__isnull=False).count(),
            'with_room': self.get_queryset().filter(room__isnull=False).count(),
            'with_office': self.get_queryset().filter(office__isnull=False).count(),
        }
        
        return context


class LocationsWithCoordinatesView(LoginRequiredMixin, ListView):
    """Display locations that have GPS coordinates."""
    model = Location
    template_name = 'locations/locations_with_coordinates.html'
    context_object_name = 'locations'
    paginate_by = 25
    
    def get_queryset(self):
        """Return locations with coordinates."""
        return Location.objects.filter(
            latitude__isnull=False,
            longitude__isnull=False
        ).select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Locations with GPS Coordinates'
        context['total_with_coordinates'] = self.get_queryset().count()
        context['total_locations'] = Location.objects.count()
        context['percentage_mapped'] = round(
            (context['total_with_coordinates'] / context['total_locations']) * 100, 1
        ) if context['total_locations'] > 0 else 0
        
        return context


class LocationsWithoutCoordinatesView(LoginRequiredMixin, ListView):
    """Display locations that don't have GPS coordinates."""
    model = Location
    template_name = 'locations/locations_without_coordinates.html'
    context_object_name = 'locations'
    paginate_by = 25
    
    def get_queryset(self):
        """Return locations without coordinates."""
        return Location.objects.filter(
            Q(latitude__isnull=True) | Q(longitude__isnull=True)
        ).select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code', 'name')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Locations without GPS Coordinates'
        context['total_without_coordinates'] = self.get_queryset().count()
        context['total_locations'] = Location.objects.count()
        context['percentage_unmapped'] = round(
            (context['total_without_coordinates'] / context['total_locations']) * 100, 1
        ) if context['total_locations'] > 0 else 0
        
        return context


class LocationHierarchyView(LoginRequiredMixin, TemplateView):
    """Display location hierarchy and relationships."""
    template_name = 'locations/location_hierarchy.html'
    
    def get_context_data(self, **kwargs):
        """Build hierarchical location data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Location Hierarchy'
        
        # Build hierarchy structure
        buildings = Building.objects.filter(is_active=True).prefetch_related(
            'location_set__floor', 'location_set__block', 
            'location_set__room', 'location_set__office'
        )
        
        hierarchy = []
        for building in buildings:
            building_data = {
                'building': building,
                'floors': {},
                'blocks': {},
                'rooms': {},
                'offices': {},
                'locations': building.location_set.filter(is_active=True)
            }
            
            # Group by components
            for location in building_data['locations']:
                if location.floor:
                    if location.floor.id not in building_data['floors']:
                        building_data['floors'][location.floor.id] = {
                            'floor': location.floor,
                            'locations': []
                        }
                    building_data['floors'][location.floor.id]['locations'].append(location)
                
                if location.block:
                    if location.block.id not in building_data['blocks']:
                        building_data['blocks'][location.block.id] = {
                            'block': location.block,
                            'locations': []
                        }
                    building_data['blocks'][location.block.id]['locations'].append(location)
                
                if location.room:
                    if location.room.id not in building_data['rooms']:
                        building_data['rooms'][location.room.id] = {
                            'room': location.room,
                            'locations': []
                        }
                    building_data['rooms'][location.room.id]['locations'].append(location)
                
                if location.office:
                    if location.office.id not in building_data['offices']:
                        building_data['offices'][location.office.id] = {
                            'office': location.office,
                            'locations': []
                        }
                    building_data['offices'][location.office.id]['locations'].append(location)
            
            hierarchy.append(building_data)
        
        context['hierarchy'] = hierarchy
        context['total_buildings'] = len(hierarchy)
        context['total_locations'] = Location.objects.filter(is_active=True).count()
        
        return context


class LocationRelationshipsView(LoginRequiredMixin, TemplateView):
    """Display location component relationships and dependencies."""
    template_name = 'locations/location_relationships.html'
    
    def get_context_data(self, **kwargs):
        """Analyze location relationships."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Location Relationships'
        
        # Component usage statistics
        context['stats'] = {
            'buildings': {
                'total': Building.objects.filter(is_active=True).count(),
                'used': Building.objects.filter(
                    location__isnull=False, is_active=True
                ).distinct().count(),
                'unused': Building.objects.filter(
                    location__isnull=True, is_active=True
                ).count()
            },
            'floors': {
                'total': Floor.objects.filter(is_active=True).count(),
                'used': Floor.objects.filter(
                    location__isnull=False, is_active=True
                ).distinct().count(),
                'unused': Floor.objects.filter(
                    location__isnull=True, is_active=True
                ).count()
            },
            'blocks': {
                'total': Block.objects.filter(is_active=True).count(),
                'used': Block.objects.filter(
                    location__isnull=False, is_active=True
                ).distinct().count(),
                'unused': Block.objects.filter(
                    location__isnull=True, is_active=True
                ).count()
            },
            'rooms': {
                'total': Room.objects.filter(is_active=True).count(),
                'used': Room.objects.filter(
                    location__isnull=False, is_active=True
                ).distinct().count(),
                'unused': Room.objects.filter(
                    location__isnull=True, is_active=True
                ).count()
            },
            'offices': {
                'total': Office.objects.filter(is_active=True).count(),
                'used': Office.objects.filter(
                    location__isnull=False, is_active=True
                ).distinct().count(),
                'unused': Office.objects.filter(
                    location__isnull=True, is_active=True
                ).count()
            }
        }
        
        # Component combinations
        context['combinations'] = Location.objects.filter(is_active=True).aggregate(
            only_building=Count('id', filter=Q(
                building__isnull=False,
                floor__isnull=True, block__isnull=True,
                room__isnull=True, office__isnull=True
            )),
            building_floor=Count('id', filter=Q(
                building__isnull=False, floor__isnull=False,
                block__isnull=True, room__isnull=True, office__isnull=True
            )),
            building_room=Count('id', filter=Q(
                building__isnull=False, room__isnull=False,
                floor__isnull=True, block__isnull=True, office__isnull=True
            )),
            building_office=Count('id', filter=Q(
                building__isnull=False, office__isnull=False,
                floor__isnull=True, block__isnull=True, room__isnull=True
            )),
            full_hierarchy=Count('id', filter=Q(
                building__isnull=False, floor__isnull=False,
                block__isnull=False, room__isnull=False, office__isnull=False
            ))
        )
        
        return context


# ============================================================================
# Export Functionality Views
# ============================================================================

class LocationExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export locations in various formats."""
    permission_required = 'locations.view_location'
    
    def get(self, request):
        """Export locations based on format parameter."""
        export_format = request.GET.get('format', 'csv').lower()
        
        # Get filtered queryset
        form = LocationSearchForm(request.GET)
        queryset = Location.objects.select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code')
        
        if form.is_valid():
            queryset = form.filter_queryset(queryset)
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        elif export_format == 'pdf':
            return self._export_pdf(queryset)
        else:
            messages.error(request, 'Invalid export format.')
            return redirect('locations:list')
    
    def _export_csv(self, queryset):
        """Export as CSV file."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="locations_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Location Code', 'Name', 'Address', 'Building', 'Floor', 
            'Block', 'Room', 'Office', 'Latitude', 'Longitude', 
            'Status', 'Created', 'Updated'
        ])
        
        for location in queryset:
            writer.writerow([
                location.location_code,
                location.name,
                location.address,
                location.building.name if location.building else '',
                location.floor.name if location.floor else '',
                location.block.name if location.block else '',
                location.room.name if location.room else '',
                location.office.name if location.office else '',
                location.latitude or '',
                location.longitude or '',
                'Active' if location.is_active else 'Inactive',
                location.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                location.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export as Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Locations"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Location Code', 'Name', 'Address', 'Building', 'Floor',
            'Block', 'Room', 'Office', 'Latitude', 'Longitude',
            'Status', 'Created', 'Updated'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, location in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=location.location_code)
            ws.cell(row=row, column=2, value=location.name)
            ws.cell(row=row, column=3, value=location.address)
            ws.cell(row=row, column=4, value=location.building.name if location.building else '')
            ws.cell(row=row, column=5, value=location.floor.name if location.floor else '')
            ws.cell(row=row, column=6, value=location.block.name if location.block else '')
            ws.cell(row=row, column=7, value=location.room.name if location.room else '')
            ws.cell(row=row, column=8, value=location.office.name if location.office else '')
            ws.cell(row=row, column=9, value=str(location.latitude) if location.latitude else '')
            ws.cell(row=row, column=10, value=str(location.longitude) if location.longitude else '')
            ws.cell(row=row, column=11, value='Active' if location.is_active else 'Inactive')
            ws.cell(row=row, column=12, value=location.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=13, value=location.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="locations_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_pdf(self, queryset):
        """Export as PDF file."""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="locations_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        story.append(Paragraph('Bangladesh Parliament Secretariat', title_style))
        story.append(Paragraph('Location Export Report', title_style))
        story.append(Spacer(1, 20))
        
        # Summary
        summary_data = [
            ['Export Date:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Locations:', str(queryset.count())],
            ['Active Locations:', str(queryset.filter(is_active=True).count())],
            ['With Coordinates:', str(queryset.filter(latitude__isnull=False, longitude__isnull=False).count())]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*72, 3*72])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Location data table
        data = [['Code', 'Name', 'Building', 'Floor', 'Room', 'Office', 'Status']]
        
        for location in queryset[:100]:  # Limit for PDF
            data.append([
                location.location_code,
                location.name[:20] + '...' if len(location.name) > 20 else location.name,
                location.building.code if location.building else '-',
                str(location.floor.floor_number) if location.floor else '-',
                location.room.room_number if location.room else '-',
                location.office.office_code if location.office else '-',
                'Active' if location.is_active else 'Inactive'
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        if queryset.count() > 100:
            story.append(Spacer(1, 12))
            story.append(Paragraph(f'Note: Only first 100 locations shown. Total: {queryset.count()}', styles['Normal']))
        
        doc.build(story)
        return response


class BuildingExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export buildings in various formats."""
    permission_required = 'locations.view_building'
    
    def get(self, request):
        """Export buildings based on format parameter."""
        export_format = request.GET.get('format', 'csv').lower()
        
        queryset = Building.objects.annotate(
            locations_count=Count('location')
        ).order_by('code')
        
        # Apply filters
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(code__icontains=search)
            )
        
        status = request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        else:
            messages.error(request, 'Invalid export format.')
            return redirect('locations:building_list')
    
    def _export_csv(self, queryset):
        """Export buildings as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="buildings_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Code', 'Name', 'Description', 'Locations Count', 'Status', 'Created', 'Updated'])
        
        for building in queryset:
            writer.writerow([
                building.code,
                building.name,
                building.description,
                building.locations_count,
                'Active' if building.is_active else 'Inactive',
                building.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                building.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export buildings as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Buildings"
        
        # Headers
        headers = ['Code', 'Name', 'Description', 'Locations Count', 'Status', 'Created', 'Updated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, building in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=building.code)
            ws.cell(row=row, column=2, value=building.name)
            ws.cell(row=row, column=3, value=building.description)
            ws.cell(row=row, column=4, value=building.locations_count)
            ws.cell(row=row, column=5, value='Active' if building.is_active else 'Inactive')
            ws.cell(row=row, column=6, value=building.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=7, value=building.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="buildings_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response
class FloorExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export floors in various formats."""
    permission_required = 'locations.view_floor'
    
    def get(self, request):
        """Export floors based on format parameter."""
        export_format = request.GET.get('format', 'csv').lower()
        
        queryset = Floor.objects.annotate(
            locations_count=Count('location')
        ).order_by('floor_number')
        
        # Apply filters
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        status = request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        else:
            messages.error(request, 'Invalid export format.')
            return redirect('locations:floor_list')
    
    def _export_csv(self, queryset):
        """Export floors as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="floors_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Floor Number', 'Description', 'Locations Count', 'Status', 'Created', 'Updated'])
        
        for floor in queryset:
            writer.writerow([
                floor.name,
                floor.floor_number,
                floor.description,
                floor.locations_count,
                'Active' if floor.is_active else 'Inactive',
                floor.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                floor.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export floors as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Floors"
        
        headers = ['Name', 'Floor Number', 'Description', 'Locations Count', 'Status', 'Created', 'Updated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, floor in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=floor.name)
            ws.cell(row=row, column=2, value=floor.floor_number)
            ws.cell(row=row, column=3, value=floor.description)
            ws.cell(row=row, column=4, value=floor.locations_count)
            ws.cell(row=row, column=5, value='Active' if floor.is_active else 'Inactive')
            ws.cell(row=row, column=6, value=floor.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=7, value=floor.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="floors_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


class BlockExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export blocks in various formats."""
    permission_required = 'locations.view_block'
    
    def get(self, request):
        """Export blocks based on format parameter."""
        export_format = request.GET.get('format', 'csv').lower()
        
        queryset = Block.objects.annotate(
            locations_count=Count('location')
        ).order_by('code')
        
        # Apply filters
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(code__icontains=search)
            )
        
        status = request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        else:
            messages.error(request, 'Invalid export format.')
            return redirect('locations:block_list')
    
    def _export_csv(self, queryset):
        """Export blocks as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="blocks_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Code', 'Name', 'Description', 'Locations Count', 'Status', 'Created', 'Updated'])
        
        for block in queryset:
            writer.writerow([
                block.code,
                block.name,
                block.description,
                block.locations_count,
                'Active' if block.is_active else 'Inactive',
                block.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                block.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export blocks as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Blocks"
        
        headers = ['Code', 'Name', 'Description', 'Locations Count', 'Status', 'Created', 'Updated']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, block in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=block.code)
            ws.cell(row=row, column=2, value=block.name)
            ws.cell(row=row, column=3, value=block.description)
            ws.cell(row=row, column=4, value=block.locations_count)
            ws.cell(row=row, column=5, value='Active' if block.is_active else 'Inactive')
            ws.cell(row=row, column=6, value=block.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=7, value=block.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="blocks_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


class RoomExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export rooms in various formats."""
    permission_required = 'locations.view_room'
    
    def get(self, request):
        """Export rooms based on format parameter."""
        export_format = request.GET.get('format', 'csv').lower()
        
        queryset = Room.objects.annotate(
            locations_count=Count('location')
        ).order_by('room_number')
        
        # Apply filters
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(room_number__icontains=search)
            )
        
        room_type = request.GET.get('room_type')
        if room_type:
            queryset = queryset.filter(room_type=room_type)
        
        status = request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        else:
            messages.error(request, 'Invalid export format.')
            return redirect('locations:room_list')
    
    def _export_csv(self, queryset):
        """Export rooms as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="rooms_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Room Number', 'Name', 'Room Type', 'Capacity', 'Area (sq ft)',
            'Description', 'Locations Count', 'Status', 'Created', 'Updated'
        ])
        
        for room in queryset:
            writer.writerow([
                room.room_number,
                room.name,
                room.get_room_type_display(),
                room.capacity or '',
                room.area_sqft or '',
                room.description,
                room.locations_count,
                'Active' if room.is_active else 'Inactive',
                room.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                room.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export rooms as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rooms"
        
        headers = [
            'Room Number', 'Name', 'Room Type', 'Capacity', 'Area (sq ft)',
            'Description', 'Locations Count', 'Status', 'Created', 'Updated'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, room in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=room.room_number)
            ws.cell(row=row, column=2, value=room.name)
            ws.cell(row=row, column=3, value=room.get_room_type_display())
            ws.cell(row=row, column=4, value=room.capacity or '')
            ws.cell(row=row, column=5, value=str(room.area_sqft) if room.area_sqft else '')
            ws.cell(row=row, column=6, value=room.description)
            ws.cell(row=row, column=7, value=room.locations_count)
            ws.cell(row=row, column=8, value='Active' if room.is_active else 'Inactive')
            ws.cell(row=row, column=9, value=room.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=10, value=room.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="rooms_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


class OfficeExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export offices in various formats."""
    permission_required = 'locations.view_office'
    
    def get(self, request):
        """Export offices based on format parameter."""
        export_format = request.GET.get('format', 'csv').lower()
        
        queryset = Office.objects.annotate(
            locations_count=Count('location')
        ).order_by('office_code')
        
        # Apply filters
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(office_code__icontains=search) |
                Q(head_of_office__icontains=search)
            )
        
        office_type = request.GET.get('office_type')
        if office_type:
            queryset = queryset.filter(office_type=office_type)
        
        status = request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        elif export_format == 'excel':
            return self._export_excel(queryset)
        else:
            messages.error(request, 'Invalid export format.')
            return redirect('locations:office_list')
    
    def _export_csv(self, queryset):
        """Export offices as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="offices_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Office Code', 'Name', 'Office Type', 'Head of Office', 
            'Contact Number', 'Email', 'Description', 'Locations Count', 
            'Status', 'Created', 'Updated'
        ])
        
        for office in queryset:
            writer.writerow([
                office.office_code,
                office.name,
                office.get_office_type_display(),
                office.head_of_office,
                office.contact_number,
                office.email,
                office.description,
                office.locations_count,
                'Active' if office.is_active else 'Inactive',
                office.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                office.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export offices as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Offices"
        
        headers = [
            'Office Code', 'Name', 'Office Type', 'Head of Office',
            'Contact Number', 'Email', 'Description', 'Locations Count',
            'Status', 'Created', 'Updated'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, office in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=office.office_code)
            ws.cell(row=row, column=2, value=office.name)
            ws.cell(row=row, column=3, value=office.get_office_type_display())
            ws.cell(row=row, column=4, value=office.head_of_office)
            ws.cell(row=row, column=5, value=office.contact_number)
            ws.cell(row=row, column=6, value=office.email)
            ws.cell(row=row, column=7, value=office.description)
            ws.cell(row=row, column=8, value=office.locations_count)
            ws.cell(row=row, column=9, value='Active' if office.is_active else 'Inactive')
            ws.cell(row=row, column=10, value=office.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=11, value=office.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="offices_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


class AllLocationsExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export all location-related data in a comprehensive format."""
    permission_required = 'locations.view_location'
    
    def get(self, request):
        """Export all location data as Excel workbook with multiple sheets."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each entity
        self._create_locations_sheet(wb)
        self._create_buildings_sheet(wb)
        self._create_floors_sheet(wb)
        self._create_blocks_sheet(wb)
        self._create_rooms_sheet(wb)
        self._create_offices_sheet(wb)
        self._create_summary_sheet(wb)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="all_locations_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response
    
    def _create_locations_sheet(self, wb):
        """Create locations sheet."""
        ws = wb.create_sheet(title="Locations")
        
        headers = [
            'Code', 'Name', 'Address', 'Building', 'Floor', 'Block',
            'Room', 'Office', 'Latitude', 'Longitude', 'Status', 'Created'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        locations = Location.objects.select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('location_code')
        
        for row, location in enumerate(locations, 2):
            ws.cell(row=row, column=1, value=location.location_code)
            ws.cell(row=row, column=2, value=location.name)
            ws.cell(row=row, column=3, value=location.address)
            ws.cell(row=row, column=4, value=location.building.name if location.building else '')
            ws.cell(row=row, column=5, value=location.floor.name if location.floor else '')
            ws.cell(row=row, column=6, value=location.block.name if location.block else '')
            ws.cell(row=row, column=7, value=location.room.name if location.room else '')
            ws.cell(row=row, column=8, value=location.office.name if location.office else '')
            ws.cell(row=row, column=9, value=str(location.latitude) if location.latitude else '')
            ws.cell(row=row, column=10, value=str(location.longitude) if location.longitude else '')
            ws.cell(row=row, column=11, value='Active' if location.is_active else 'Inactive')
            ws.cell(row=row, column=12, value=location.created_at.strftime('%Y-%m-%d'))
    
    def _create_buildings_sheet(self, wb):
        """Create buildings sheet."""
        ws = wb.create_sheet(title="Buildings")
        
        headers = ['Code', 'Name', 'Description', 'Locations Count', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        buildings = Building.objects.annotate(
            locations_count=Count('location')
        ).order_by('code')
        
        for row, building in enumerate(buildings, 2):
            ws.cell(row=row, column=1, value=building.code)
            ws.cell(row=row, column=2, value=building.name)
            ws.cell(row=row, column=3, value=building.description)
            ws.cell(row=row, column=4, value=building.locations_count)
            ws.cell(row=row, column=5, value='Active' if building.is_active else 'Inactive')
    
    def _create_floors_sheet(self, wb):
        """Create floors sheet."""
        ws = wb.create_sheet(title="Floors")
        
        headers = ['Name', 'Floor Number', 'Description', 'Locations Count', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        floors = Floor.objects.annotate(
            locations_count=Count('location')
        ).order_by('floor_number')
        
        for row, floor in enumerate(floors, 2):
            ws.cell(row=row, column=1, value=floor.name)
            ws.cell(row=row, column=2, value=floor.floor_number)
            ws.cell(row=row, column=3, value=floor.description)
            ws.cell(row=row, column=4, value=floor.locations_count)
            ws.cell(row=row, column=5, value='Active' if floor.is_active else 'Inactive')
    
    def _create_blocks_sheet(self, wb):
        """Create blocks sheet."""
        ws = wb.create_sheet(title="Blocks")
        
        headers = ['Code', 'Name', 'Description', 'Locations Count', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        blocks = Block.objects.annotate(
            locations_count=Count('location')
        ).order_by('code')
        
        for row, block in enumerate(blocks, 2):
            ws.cell(row=row, column=1, value=block.code)
            ws.cell(row=row, column=2, value=block.name)
            ws.cell(row=row, column=3, value=block.description)
            ws.cell(row=row, column=4, value=block.locations_count)
            ws.cell(row=row, column=5, value='Active' if block.is_active else 'Inactive')
    
    def _create_rooms_sheet(self, wb):
        """Create rooms sheet."""
        ws = wb.create_sheet(title="Rooms")
        
        headers = ['Room Number', 'Name', 'Type', 'Capacity', 'Area', 'Locations Count', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        rooms = Room.objects.annotate(
            locations_count=Count('location')
        ).order_by('room_number')
        
        for row, room in enumerate(rooms, 2):
            ws.cell(row=row, column=1, value=room.room_number)
            ws.cell(row=row, column=2, value=room.name)
            ws.cell(row=row, column=3, value=room.get_room_type_display())
            ws.cell(row=row, column=4, value=room.capacity or '')
            ws.cell(row=row, column=5, value=str(room.area_sqft) if room.area_sqft else '')
            ws.cell(row=row, column=6, value=room.locations_count)
            ws.cell(row=row, column=7, value='Active' if room.is_active else 'Inactive')
    
    def _create_offices_sheet(self, wb):
        """Create offices sheet."""
        ws = wb.create_sheet(title="Offices")
        
        headers = ['Code', 'Name', 'Type', 'Head of Office', 'Contact', 'Locations Count', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        offices = Office.objects.annotate(
            locations_count=Count('location')
        ).order_by('office_code')
        
        for row, office in enumerate(offices, 2):
            ws.cell(row=row, column=1, value=office.office_code)
            ws.cell(row=row, column=2, value=office.name)
            ws.cell(row=row, column=3, value=office.get_office_type_display())
            ws.cell(row=row, column=4, value=office.head_of_office)
            ws.cell(row=row, column=5, value=office.contact_number)
            ws.cell(row=row, column=6, value=office.locations_count)
            ws.cell(row=row, column=7, value='Active' if office.is_active else 'Inactive')
    
    def _create_summary_sheet(self, wb):
        """Create summary sheet."""
        ws = wb.create_sheet(title="Summary", index=0)
        
        # Title
        ws.cell(row=1, column=1, value='Bangladesh Parliament Secretariat')
        ws.cell(row=2, column=1, value='Location Data Export Summary')
        ws.cell(row=3, column=1, value=f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}')
        
        # Summary data
        summary_data = [
            ['Entity', 'Total', 'Active', 'Inactive'],
            ['Locations', Location.objects.count(), Location.objects.filter(is_active=True).count(), Location.objects.filter(is_active=False).count()],
            ['Buildings', Building.objects.count(), Building.objects.filter(is_active=True).count(), Building.objects.filter(is_active=False).count()],
            ['Floors', Floor.objects.count(), Floor.objects.filter(is_active=True).count(), Floor.objects.filter(is_active=False).count()],
            ['Blocks', Block.objects.count(), Block.objects.filter(is_active=True).count(), Block.objects.filter(is_active=False).count()],
            ['Rooms', Room.objects.count(), Room.objects.filter(is_active=True).count(), Room.objects.filter(is_active=False).count()],
            ['Offices', Office.objects.count(), Office.objects.filter(is_active=True).count(), Office.objects.filter(is_active=False).count()],
        ]
        
        for row, data in enumerate(summary_data, 5):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 5:  # Header row
                    cell.font = Font(bold=True)


class LocationTemplateExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Export location import templates."""
    permission_required = 'locations.add_location'
    
    def get(self, request):
        """Export template file for location imports."""
        template_type = request.GET.get('type', 'location').lower()
        
        wb = Workbook()
        ws = wb.active
        
        if template_type == 'location':
            ws.title = "Location Import Template"
            headers = [
                'location_code', 'name', 'address', 'building_code', 
                'floor_number', 'block_code', 'room_number', 'office_code',
                'latitude', 'longitude', 'is_active', 'description'
            ]
            sample_data = [
                'LOC001', 'Sample Location', '123 Parliament St', 'MPB',
                '1', 'BLK01', 'R101', 'OFF01', '23.7465', '90.3915', 'TRUE', 'Sample description'
            ]
        
        elif template_type == 'building':
            ws.title = "Building Import Template"
            headers = ['code', 'name', 'description', 'is_active']
            sample_data = ['MPB', 'Main Parliament Building', 'Primary building', 'TRUE']
        
        elif template_type == 'room':
            ws.title = "Room Import Template"
            headers = ['room_number', 'name', 'room_type', 'capacity', 'area_sqft', 'description', 'is_active']
            sample_data = ['R101', 'Conference Room 1', 'conference', '20', '500', 'Main conference room', 'TRUE']
        
        elif template_type == 'office':
            ws.title = "Office Import Template"
            headers = ['office_code', 'name', 'office_type', 'head_of_office', 'contact_number', 'email', 'description', 'is_active']
            sample_data = ['OFF01', 'IT Department', 'wing', 'John Doe', '+8801234567890', 'it@parliament.gov.bd', 'IT Wing Office', 'TRUE']
        
        else:
            messages.error(request, 'Invalid template type.')
            return redirect('locations:list')
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add sample data
        for col, data in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{template_type}_import_template.xlsx"'
        wb.save(response)
        
        return response

class LocationDashboardView(LoginRequiredMixin, TemplateView):
    """Main dashboard for location management with key metrics and charts."""
    template_name = 'locations/dashboard.html'
    
    def get_context_data(self, **kwargs):
        """Gather dashboard data and statistics."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Location Management Dashboard'
        
        # Core statistics
        context['stats'] = {
            'total_locations': Location.objects.count(),
            'active_locations': Location.objects.filter(is_active=True).count(),
            'inactive_locations': Location.objects.filter(is_active=False).count(),
            'locations_with_coordinates': Location.objects.filter(
                latitude__isnull=False, longitude__isnull=False
            ).count(),
            'total_buildings': Building.objects.filter(is_active=True).count(),
            'total_floors': Floor.objects.filter(is_active=True).count(),
            'total_blocks': Block.objects.filter(is_active=True).count(),
            'total_rooms': Room.objects.filter(is_active=True).count(),
            'total_offices': Office.objects.filter(is_active=True).count(),
        }
        
        # Calculate percentages
        total_locs = context['stats']['total_locations']
        if total_locs > 0:
            context['percentages'] = {
                'active_percent': round((context['stats']['active_locations'] / total_locs) * 100, 1),
                'mapped_percent': round((context['stats']['locations_with_coordinates'] / total_locs) * 100, 1),
            }
        else:
            context['percentages'] = {'active_percent': 0, 'mapped_percent': 0}
        
        # Component usage analysis
        context['component_usage'] = {
            'buildings_used': Building.objects.filter(location__isnull=False, is_active=True).distinct().count(),
            'floors_used': Floor.objects.filter(location__isnull=False, is_active=True).distinct().count(),
            'blocks_used': Block.objects.filter(location__isnull=False, is_active=True).distinct().count(),
            'rooms_used': Room.objects.filter(location__isnull=False, is_active=True).distinct().count(),
            'offices_used': Office.objects.filter(location__isnull=False, is_active=True).distinct().count(),
        }
        
        # Room type distribution
        context['room_types'] = Room.objects.filter(is_active=True).values(
            'room_type'
        ).annotate(
            count=Count('id'),
            locations_count=Count('location')
        ).order_by('-count')
        
        # Office type distribution
        context['office_types'] = Office.objects.filter(is_active=True).values(
            'office_type'
        ).annotate(
            count=Count('id'),
            locations_count=Count('location')
        ).order_by('-count')
        
        # Recent activity
        context['recent_locations'] = Location.objects.select_related(
            'building', 'floor', 'room', 'office'
        ).order_by('-created_at')[:10]
        
        context['recent_buildings'] = Building.objects.order_by('-created_at')[:5]
        
        # Floor distribution
        context['floor_distribution'] = Floor.objects.filter(is_active=True).values(
            'floor_number'
        ).annotate(
            count=Count('location')
        ).order_by('floor_number')
        
        # Top buildings by location count
        context['top_buildings'] = Building.objects.filter(is_active=True).annotate(
            location_count=Count('location')
        ).order_by('-location_count')[:10]
        
        # Geographic coverage
        context['geographic_stats'] = {
            'mapped_locations': Location.objects.filter(
                latitude__isnull=False, longitude__isnull=False, is_active=True
            ).count(),
            'unmapped_locations': Location.objects.filter(
                Q(latitude__isnull=True) | Q(longitude__isnull=True), is_active=True
            ).count()
        }
        
        # Weekly creation trends (last 8 weeks)
        from datetime import timedelta
        end_date = timezone.now()
        start_date = end_date - timedelta(weeks=8)
        
        weekly_data = []
        for week in range(8):
            week_start = start_date + timedelta(weeks=week)
            week_end = week_start + timedelta(weeks=1)
            
            locations_created = Location.objects.filter(
                created_at__gte=week_start,
                created_at__lt=week_end
            ).count()
            
            weekly_data.append({
                'week': week_start.strftime('%Y-W%W'),
                'locations': locations_created
            })
        
        context['weekly_trends'] = weekly_data
        
        return context


class LocationAnalyticsView(LoginRequiredMixin, TemplateView):
    """Advanced analytics and insights for location data."""
    template_name = 'locations/analytics.html'
    
    def get_context_data(self, **kwargs):
        """Generate comprehensive analytics data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Location Analytics'
        
        # Hierarchy completeness analysis
        context['hierarchy_analysis'] = {
            'complete_hierarchy': Location.objects.filter(
                building__isnull=False,
                floor__isnull=False,
                block__isnull=False,
                room__isnull=False,
                office__isnull=False,
                is_active=True
            ).count(),
            'partial_hierarchy': Location.objects.filter(
                is_active=True
            ).exclude(
                building__isnull=False,
                floor__isnull=False,
                block__isnull=False,
                room__isnull=False,
                office__isnull=False
            ).count(),
            'minimal_info': Location.objects.filter(
                building__isnull=True,
                floor__isnull=True,
                block__isnull=True,
                room__isnull=True,
                office__isnull=True,
                is_active=True
            ).count()
        }
        
        # Component relationship matrix
        context['relationship_matrix'] = {}
        relationships = [
            ('building', 'floor'), ('building', 'block'), ('building', 'room'), ('building', 'office'),
            ('floor', 'block'), ('floor', 'room'), ('floor', 'office'),
            ('block', 'room'), ('block', 'office'),
            ('room', 'office')
        ]
        
        for rel1, rel2 in relationships:
            filter_kwargs = {
                f'{rel1}__isnull': False,
                f'{rel2}__isnull': False,
                'is_active': True
            }
            count = Location.objects.filter(**filter_kwargs).count()
            context['relationship_matrix'][f'{rel1}_{rel2}'] = count
        
        # Capacity analysis (for rooms)
        context['capacity_analysis'] = {
            'rooms_with_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).count(),
            'total_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).aggregate(total=models.Sum('capacity'))['total'] or 0,
            'avg_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).aggregate(avg=models.Avg('capacity'))['avg'] or 0,
            'capacity_ranges': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).aggregate(
                small=Count('id', filter=Q(capacity__lt=10)),
                medium=Count('id', filter=Q(capacity__gte=10, capacity__lt=50)),
                large=Count('id', filter=Q(capacity__gte=50, capacity__lt=100)),
                xlarge=Count('id', filter=Q(capacity__gte=100))
            )
        }
        
        # Area analysis (for rooms)
        context['area_analysis'] = {
            'rooms_with_area': Room.objects.filter(
                area_sqft__isnull=False, is_active=True
            ).count(),
            'total_area': Room.objects.filter(
                area_sqft__isnull=False, is_active=True
            ).aggregate(total=models.Sum('area_sqft'))['total'] or 0,
            'avg_area': Room.objects.filter(
                area_sqft__isnull=False, is_active=True
            ).aggregate(avg=models.Avg('area_sqft'))['avg'] or 0
        }
        
        # Office hierarchy analysis
        context['office_hierarchy'] = {}
        for office_type, display_name in Office.OFFICE_TYPES:
            context['office_hierarchy'][office_type] = {
                'count': Office.objects.filter(office_type=office_type, is_active=True).count(),
                'with_locations': Office.objects.filter(
                    office_type=office_type, is_active=True, location__isnull=False
                ).distinct().count(),
                'avg_locations': Office.objects.filter(
                    office_type=office_type, is_active=True
                ).annotate(
                    loc_count=Count('location')
                ).aggregate(
                    avg=models.Avg('loc_count')
                )['avg'] or 0
            }
        
        # Data quality metrics
        context['data_quality'] = {
            'locations_with_description': Location.objects.filter(
                description__isnull=False, description__gt='', is_active=True
            ).count(),
            'buildings_with_description': Building.objects.filter(
                description__isnull=False, description__gt='', is_active=True
            ).count(),
            'rooms_with_complete_info': Room.objects.filter(
                capacity__isnull=False, area_sqft__isnull=False, is_active=True
            ).count(),
            'offices_with_contact': Office.objects.filter(
                Q(contact_number__isnull=False, contact_number__gt='') |
                Q(email__isnull=False, email__gt=''),
                is_active=True
            ).count()
        }
        
        # Growth trends (monthly for last 12 months)
        from datetime import datetime, timedelta
        import calendar
        
        monthly_data = []
        current_date = timezone.now()
        
        for i in range(12):
            month_start = current_date.replace(day=1) - timedelta(days=32*i)
            month_start = month_start.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            locations_created = Location.objects.filter(
                created_at__gte=month_start,
                created_at__lte=month_end
            ).count()
            
            buildings_created = Building.objects.filter(
                created_at__gte=month_start,
                created_at__lte=month_end
            ).count()
            
            monthly_data.insert(0, {
                'month': month_start.strftime('%Y-%m'),
                'month_name': calendar.month_name[month_start.month],
                'locations': locations_created,
                'buildings': buildings_created
            })
        
        context['monthly_trends'] = monthly_data
        
        return context


class LocationStatisticsView(LoginRequiredMixin, View):
    """AJAX endpoint for location statistics."""
    
    def get(self, request):
        """Return location statistics as JSON."""
        stat_type = request.GET.get('type', 'overview')
        
        if stat_type == 'overview':
            data = {
                'total_locations': Location.objects.count(),
                'active_locations': Location.objects.filter(is_active=True).count(),
                'with_coordinates': Location.objects.filter(
                    latitude__isnull=False, longitude__isnull=False
                ).count(),
                'component_counts': {
                    'buildings': Building.objects.filter(is_active=True).count(),
                    'floors': Floor.objects.filter(is_active=True).count(),
                    'blocks': Block.objects.filter(is_active=True).count(),
                    'rooms': Room.objects.filter(is_active=True).count(),
                    'offices': Office.objects.filter(is_active=True).count(),
                }
            }
        
        elif stat_type == 'room_types':
            data = list(Room.objects.filter(is_active=True).values(
                'room_type'
            ).annotate(
                count=Count('id')
            ).order_by('-count'))
            
            for item in data:
                item['room_type_display'] = dict(Room.ROOM_TYPES).get(
                    item['room_type'], item['room_type']
                )
        
        elif stat_type == 'office_types':
            data = list(Office.objects.filter(is_active=True).values(
                'office_type'
            ).annotate(
                count=Count('id')
            ).order_by('-count'))
            
            for item in data:
                item['office_type_display'] = dict(Office.OFFICE_TYPES).get(
                    item['office_type'], item['office_type']
                )
        
        elif stat_type == 'floor_distribution':
            data = list(Floor.objects.filter(is_active=True).values(
                'floor_number'
            ).annotate(
                location_count=Count('location')
            ).order_by('floor_number'))
        
        elif stat_type == 'building_usage':
            data = list(Building.objects.filter(is_active=True).annotate(
                location_count=Count('location')
            ).values(
                'code', 'name', 'location_count'
            ).order_by('-location_count')[:10])
        
        else:
            return JsonResponse({'error': 'Invalid statistics type'}, status=400)
        
        return JsonResponse({
            'success': True,
            'data': data,
            'generated_at': timezone.now().isoformat()
        })


class BuildingStatsView(LoginRequiredMixin, TemplateView):
    """Detailed statistics for buildings."""
    template_name = 'locations/building_stats.html'
    
    def get_context_data(self, **kwargs):
        """Generate building-specific statistics."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Building Statistics'
        
        # Building overview
        context['building_stats'] = {
            'total': Building.objects.count(),
            'active': Building.objects.filter(is_active=True).count(),
            'inactive': Building.objects.filter(is_active=False).count(),
            'with_locations': Building.objects.filter(
                location__isnull=False, is_active=True
            ).distinct().count(),
            'without_locations': Building.objects.filter(
                location__isnull=True, is_active=True
            ).count()
        }
        
        # Buildings with location counts
        context['buildings_data'] = Building.objects.filter(is_active=True).annotate(
            location_count=Count('location'),
            active_location_count=Count('location', filter=Q(location__is_active=True))
        ).order_by('-location_count')
        
        # Building utilization analysis
        context['utilization'] = {
            'high_usage': Building.objects.filter(is_active=True).annotate(
                loc_count=Count('location')
            ).filter(loc_count__gte=10).count(),
            'medium_usage': Building.objects.filter(is_active=True).annotate(
                loc_count=Count('location')
            ).filter(loc_count__gte=5, loc_count__lt=10).count(),
            'low_usage': Building.objects.filter(is_active=True).annotate(
                loc_count=Count('location')
            ).filter(loc_count__gte=1, loc_count__lt=5).count(),
            'no_usage': Building.objects.filter(is_active=True).annotate(
                loc_count=Count('location')
            ).filter(loc_count=0).count()
        }
        
        return context


class FloorStatsView(LoginRequiredMixin, TemplateView):
    """Detailed statistics for floors."""
    template_name = 'locations/floor_stats.html'
    
    def get_context_data(self, **kwargs):
        """Generate floor-specific statistics."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Floor Statistics'
        
        # Floor overview
        context['floor_stats'] = {
            'total': Floor.objects.count(),
            'active': Floor.objects.filter(is_active=True).count(),
            'inactive': Floor.objects.filter(is_active=False).count(),
            'with_locations': Floor.objects.filter(
                location__isnull=False, is_active=True
            ).distinct().count()
        }
        
        # Floor distribution
        context['floor_distribution'] = Floor.objects.filter(is_active=True).annotate(
            location_count=Count('location')
        ).values(
            'floor_number', 'name', 'location_count'
        ).order_by('floor_number')
        
        # Floor level analysis
        context['level_analysis'] = {
            'basement_floors': Floor.objects.filter(
                floor_number__lt=0, is_active=True
            ).count(),
            'ground_floor': Floor.objects.filter(
                floor_number=0, is_active=True
            ).count(),
            'upper_floors': Floor.objects.filter(
                floor_number__gt=0, is_active=True
            ).count(),
            'highest_floor': Floor.objects.filter(is_active=True).aggregate(
                max_floor=models.Max('floor_number')
            )['max_floor'] or 0,
            'lowest_floor': Floor.objects.filter(is_active=True).aggregate(
                min_floor=models.Min('floor_number')
            )['min_floor'] or 0
        }
        
        return context


class RoomStatsView(LoginRequiredMixin, TemplateView):
    """Detailed statistics for rooms."""
    template_name = 'locations/room_stats.html'
    
    def get_context_data(self, **kwargs):
        """Generate room-specific statistics."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Room Statistics'
        
        # Room overview
        context['room_stats'] = {
            'total': Room.objects.count(),
            'active': Room.objects.filter(is_active=True).count(),
            'inactive': Room.objects.filter(is_active=False).count(),
            'with_locations': Room.objects.filter(
                location__isnull=False, is_active=True
            ).distinct().count()
        }
        
        # Room type breakdown
        context['room_type_stats'] = []
        for room_type, display_name in Room.ROOM_TYPES:
            stats = {
                'type': room_type,
                'display_name': display_name,
                'count': Room.objects.filter(room_type=room_type, is_active=True).count(),
                'with_capacity': Room.objects.filter(
                    room_type=room_type, capacity__isnull=False, is_active=True
                ).count(),
                'avg_capacity': Room.objects.filter(
                    room_type=room_type, capacity__isnull=False, is_active=True
                ).aggregate(avg=models.Avg('capacity'))['avg'] or 0,
                'total_capacity': Room.objects.filter(
                    room_type=room_type, capacity__isnull=False, is_active=True
                ).aggregate(total=models.Sum('capacity'))['total'] or 0
            }
            context['room_type_stats'].append(stats)
        
        # Capacity analysis
        context['capacity_stats'] = {
            'rooms_with_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).count(),
            'total_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).aggregate(total=models.Sum('capacity'))['total'] or 0,
            'avg_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).aggregate(avg=models.Avg('capacity'))['avg'] or 0,
            'max_capacity': Room.objects.filter(
                capacity__isnull=False, is_active=True
            ).aggregate(max=models.Max('capacity'))['max'] or 0
        }
        
        # Area analysis
        context['area_stats'] = {
            'rooms_with_area': Room.objects.filter(
                area_sqft__isnull=False, is_active=True
            ).count(),
            'total_area': Room.objects.filter(
                area_sqft__isnull=False, is_active=True
            ).aggregate(total=models.Sum('area_sqft'))['total'] or 0,
            'avg_area': Room.objects.filter(
                area_sqft__isnull=False, is_active=True
            ).aggregate(avg=models.Avg('area_sqft'))['avg'] or 0
        }
        
        return context


class OfficeStatsView(LoginRequiredMixin, TemplateView):
    """Detailed statistics for offices."""
    template_name = 'locations/office_stats.html'
    
    def get_context_data(self, **kwargs):
        """Generate office-specific statistics."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Office Statistics'
        
        # Office overview
        context['office_stats'] = {
            'total': Office.objects.count(),
            'active': Office.objects.filter(is_active=True).count(),
            'inactive': Office.objects.filter(is_active=False).count(),
            'with_locations': Office.objects.filter(
                location__isnull=False, is_active=True
            ).distinct().count()
        }
        
        # Office type breakdown
        context['office_type_stats'] = []
        for office_type, display_name in Office.OFFICE_TYPES:
            stats = {
                'type': office_type,
                'display_name': display_name,
                'count': Office.objects.filter(office_type=office_type, is_active=True).count(),
                'with_head': Office.objects.filter(
                    office_type=office_type, head_of_office__isnull=False,
                    head_of_office__gt='', is_active=True
                ).count(),
                'with_contact': Office.objects.filter(
                    office_type=office_type, is_active=True
                ).filter(
                    Q(contact_number__isnull=False, contact_number__gt='') |
                    Q(email__isnull=False, email__gt='')
                ).count(),
                'location_count': Office.objects.filter(
                    office_type=office_type, is_active=True
                ).aggregate(
                    total_locations=Count('location')
                )['total_locations'] or 0
            }
            context['office_type_stats'].append(stats)
        
        # Contact information completeness
        context['contact_stats'] = {
            'with_phone': Office.objects.filter(
                contact_number__isnull=False, contact_number__gt='', is_active=True
            ).count(),
            'with_email': Office.objects.filter(
                email__isnull=False, email__gt='', is_active=True
            ).count(),
            'with_both': Office.objects.filter(
                contact_number__isnull=False, contact_number__gt='',
                email__isnull=False, email__gt='', is_active=True
            ).count(),
            'with_head': Office.objects.filter(
                head_of_office__isnull=False, head_of_office__gt='', is_active=True
            ).count()
        }
        
        return context
# ============================================================================
# Import Functionality Views
# ============================================================================

class BuildingImportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Import buildings from Excel/CSV files."""
    template_name = 'locations/building_import.html'
    permission_required = 'locations.add_building'
    
    def get_context_data(self, **kwargs):
        """Add import context."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Import Buildings'
        context['entity_name'] = 'Buildings'
        context['template_fields'] = [
            'code (required)', 'name (required)', 'description', 'is_active (TRUE/FALSE)'
        ]
        return context
    
    def post(self, request):
        """Handle building import file upload and processing."""
        if 'import_file' not in request.FILES:
            messages.error(request, 'No file selected for import.')
            return redirect('locations:building_import')
        
        import_file = request.FILES['import_file']
        
        # Validate file type
        if not import_file.name.endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, 'Invalid file format. Please upload Excel (.xlsx, .xls) or CSV file.')
            return redirect('locations:building_import')
        
        try:
            if import_file.name.endswith('.csv'):
                result = self._import_csv(import_file)
            else:
                result = self._import_excel(import_file)
            
            if result['success']:
                messages.success(
                    request,
                    f"Buildings imported successfully! {result['created']} created, "
                    f"{result['updated']} updated, {result['errors']} errors."
                )
                if result['error_details']:
                    for error in result['error_details'][:5]:  # Show first 5 errors
                        messages.warning(request, error)
            else:
                messages.error(request, f"Import failed: {result['message']}")
                
        except Exception as e:
            messages.error(request, f'Import error: {str(e)}')
        
        return redirect('locations:building_import')
    
    def _import_csv(self, file):
        """Import buildings from CSV file."""
        import csv
        import io
        
        try:
            file_content = file.read().decode('utf-8')
            csv_data = csv.DictReader(io.StringIO(file_content))
            return self._process_building_data(csv_data)
        except UnicodeDecodeError:
            # Try with different encoding
            file.seek(0)
            file_content = file.read().decode('utf-8-sig')
            csv_data = csv.DictReader(io.StringIO(file_content))
            return self._process_building_data(csv_data)
    
    def _import_excel(self, file):
        """Import buildings from Excel file."""
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1] if cell.value]
        
        # Convert to dict format
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                row_dict = dict(zip(headers, row))
                data.append(row_dict)
        
        return self._process_building_data(data)
    
    def _process_building_data(self, data):
        """Process building import data and create/update buildings."""
        created_count = 0
        updated_count = 0
        error_count = 0
        error_details = []
        
        for row_num, row in enumerate(data, 2):
            try:
                # Validate required fields
                code = row.get('code', '').strip()
                name = row.get('name', '').strip()
                
                if not code:
                    error_details.append(f"Row {row_num}: Building code is required")
                    error_count += 1
                    continue
                
                if not name:
                    error_details.append(f"Row {row_num}: Building name is required")
                    error_count += 1
                    continue
                
                # Prepare building data
                building_data = {
                    'name': name,
                    'description': row.get('description', '').strip(),
                    'is_active': str(row.get('is_active', 'TRUE')).upper() in ['TRUE', '1', 'YES', 'Y']
                }
                
                # Create or update building
                building, created = Building.objects.update_or_create(
                    code=code.upper(),
                    defaults=building_data
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
                    
            except Exception as e:
                error_details.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': error_count,
            'error_details': error_details
        }

class LocationImportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Import locations from Excel/CSV files."""
    template_name = 'locations/location_import.html'
    permission_required = 'locations.add_location'
    
    def get_context_data(self, **kwargs):
        """Add import context."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Import Locations'
        return context
    
    def post(self, request):
        """Handle file upload and import."""
        if 'import_file' not in request.FILES:
            messages.error(request, 'No file selected for import.')
            return redirect('locations:import')
        
        import_file = request.FILES['import_file']
        
        if not import_file.name.endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, 'Invalid file format. Please upload Excel or CSV file.')
            return redirect('locations:import')
        
        try:
            if import_file.name.endswith('.csv'):
                result = self._import_csv(import_file)
            else:
                result = self._import_excel(import_file)
            
            if result['success']:
                messages.success(
                    request, 
                    f"Import completed successfully! {result['created']} locations created, "
                    f"{result['updated']} updated, {result['errors']} errors."
                )
            else:
                messages.error(request, f"Import failed: {result['message']}")
                
        except Exception as e:
            messages.error(request, f'Import error: {str(e)}')
        
        return redirect('locations:import')
    
    def _import_csv(self, file):
        """Import from CSV file."""
        import csv
        import io
        
        file_content = file.read().decode('utf-8')
        csv_data = csv.DictReader(io.StringIO(file_content))
        
        return self._process_building_data(csv_data)
    
    def _import_excel(self, file):
        """Import buildings from Excel."""
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        
        return self._process_building_data(data)
    
    def _process_building_data(self, data):
        """Process building import data."""
        created_count = 0
        updated_count = 0
        
        for row in data:
            building_data = {
                'name': row.get('name', ''),
                'description': row.get('description', ''),
                'is_active': str(row.get('is_active', 'TRUE')).upper() == 'TRUE'
            }
            
            building, created = Building.objects.update_or_create(
                code=row['code'].upper(),
                defaults=building_data
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return {'created': created_count, 'updated': updated_count}


class RoomImportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Import rooms from Excel/CSV files."""
    template_name = 'locations/room_import.html'
    permission_required = 'locations.add_room'
    
    def post(self, request):
        """Handle room import."""
        if 'import_file' not in request.FILES:
            messages.error(request, 'No file selected for import.')
            return redirect('locations:room_import')
        
        import_file = request.FILES['import_file']
        
        try:
            if import_file.name.endswith('.csv'):
                result = self._import_csv(import_file)
            else:
                result = self._import_excel(import_file)
            
            messages.success(
                request,
                f"Rooms imported: {result['created']} created, {result['updated']} updated."
            )
        except Exception as e:
            messages.error(request, f'Import error: {str(e)}')
        
        return redirect('locations:room_import')
    
    def _import_csv(self, file):
        """Import rooms from CSV."""
        import csv
        import io
        
        file_content = file.read().decode('utf-8')
        csv_data = csv.DictReader(io.StringIO(file_content))
        
        return self._process_room_data(csv_data)
    
    def _import_excel(self, file):
        """Import rooms from Excel."""
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        
        return self._process_room_data(data)
    
    def _process_room_data(self, data):
        """Process room import data."""
        created_count = 0
        updated_count = 0
        
        for row in data:
            room_data = {
                'name': row.get('name', ''),
                'room_type': row.get('room_type', 'office'),
                'description': row.get('description', ''),
                'is_active': str(row.get('is_active', 'TRUE')).upper() == 'TRUE'
            }
            
            # Handle optional numeric fields
            if row.get('capacity'):
                try:
                    room_data['capacity'] = int(row['capacity'])
                except (ValueError, TypeError):
                    pass
            
            if row.get('area_sqft'):
                try:
                    room_data['area_sqft'] = float(row['area_sqft'])
                except (ValueError, TypeError):
                    pass
            
            room, created = Room.objects.update_or_create(
                room_number=row['room_number'],
                defaults=room_data
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return {'created': created_count, 'updated': updated_count}


class OfficeImportView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Import offices from Excel/CSV files."""
    template_name = 'locations/office_import.html'
    permission_required = 'locations.add_office'
    
    def post(self, request):
        """Handle office import."""
        if 'import_file' not in request.FILES:
            messages.error(request, 'No file selected for import.')
            return redirect('locations:office_import')
        
        import_file = request.FILES['import_file']
        
        try:
            if import_file.name.endswith('.csv'):
                result = self._import_csv(import_file)
            else:
                result = self._import_excel(import_file)
            
            messages.success(
                request,
                f"Offices imported: {result['created']} created, {result['updated']} updated."
            )
        except Exception as e:
            messages.error(request, f'Import error: {str(e)}')
        
        return redirect('locations:office_import')
    
    def _import_csv(self, file):
        """Import offices from CSV."""
        import csv
        import io
        
        file_content = file.read().decode('utf-8')
        csv_data = csv.DictReader(io.StringIO(file_content))
        
        return self._process_office_data(csv_data)
    
    def _import_excel(self, file):
        """Import offices from Excel."""
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        
        return self._process_office_data(data)
    
    def _process_office_data(self, data):
        """Process office import data."""
        created_count = 0
        updated_count = 0
        
        for row in data:
            office_data = {
                'name': row.get('name', ''),
                'office_type': row.get('office_type', 'wing'),
                'head_of_office': row.get('head_of_office', ''),
                'contact_number': row.get('contact_number', ''),
                'email': row.get('email', ''),
                'description': row.get('description', ''),
                'is_active': str(row.get('is_active', 'TRUE')).upper() == 'TRUE'
            }
            
            office, created = Office.objects.update_or_create(
                office_code=row['office_code'].upper(),
                defaults=office_data
            )
            
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        return {'created': created_count, 'updated': updated_count}


# ============================================================================
# Reports Views
# ============================================================================

class LocationReportsView(LoginRequiredMixin, TemplateView):
    """Main reports dashboard for locations."""
    template_name = 'locations/reports.html'
    
    def get_context_data(self, **kwargs):
        """Add reports context."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Location Reports'
        
        # Available report types
        context['report_types'] = [
            {
                'name': 'Location Hierarchy Report',
                'description': 'Complete hierarchy structure with all relationships',
                'url': reverse('locations:hierarchy_report'),
                'icon': 'fas fa-sitemap'
            },
            {
                'name': 'Location Summary Report',
                'description': 'Summary statistics and key metrics',
                'url': reverse('locations:summary_report'),
                'icon': 'fas fa-chart-bar'
            },
            {
                'name': 'Coordinates Export',
                'description': 'GPS coordinates for mapping applications',
                'url': reverse('locations:coordinates_export'),
                'icon': 'fas fa-map-marked'
            },
            {
                'name': 'Building Utilization Report',
                'description': 'Building usage and capacity analysis',
                'url': '#',
                'icon': 'fas fa-building'
            },
            {
                'name': 'Room Occupancy Report',
                'description': 'Room types and capacity breakdown',
                'url': '#',
                'icon': 'fas fa-door-open'
            }
        ]
        
        return context


class LocationHierarchyReportView(LoginRequiredMixin, View):
    """Generate detailed hierarchy report."""
    
    def get(self, request):
        """Generate hierarchy report in requested format."""
        export_format = request.GET.get('format', 'pdf').lower()
        
        if export_format == 'pdf':
            return self._generate_pdf_report()
        elif export_format == 'excel':
            return self._generate_excel_report()
        else:
            messages.error(request, 'Invalid report format.')
            return redirect('locations:reports')
    
    def _generate_pdf_report(self):
        """Generate PDF hierarchy report."""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="location_hierarchy_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        
        story.append(Paragraph('Bangladesh Parliament Secretariat', title_style))
        story.append(Paragraph('Location Hierarchy Report', title_style))
        story.append(Spacer(1, 20))
        
        # Generate hierarchy data
        buildings = Building.objects.filter(is_active=True).prefetch_related(
            'location_set__floor', 'location_set__block', 
            'location_set__room', 'location_set__office'
        )
        
        for building in buildings:
            # Building header
            building_style = ParagraphStyle(
                'BuildingHeader',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12
            )
            
            story.append(Paragraph(f'Building: {building.name} ({building.code})', building_style))
            
            # Location table for this building
            locations = building.location_set.filter(is_active=True)
            if locations.exists():
                data = [['Code', 'Name', 'Floor', 'Room', 'Office', 'Coordinates']]
                
                for location in locations:
                    coords = f"{location.latitude}, {location.longitude}" if location.has_coordinates() else "N/A"
                    data.append([
                        location.location_code,
                        location.name[:30] + '...' if len(location.name) > 30 else location.name,
                        str(location.floor.floor_number) if location.floor else '-',
                        location.room.room_number if location.room else '-',
                        location.office.office_code if location.office else '-',
                        coords
                    ])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 7)
                ]))
                
                story.append(table)
            else:
                story.append(Paragraph('No active locations in this building.', styles['Normal']))
            
            story.append(Spacer(1, 20))
        
        doc.build(story)
        return response
    
    def _generate_excel_report(self):
        """Generate Excel hierarchy report."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        summary_ws.cell(row=1, column=1, value='Location Hierarchy Report')
        summary_ws.cell(row=2, column=1, value=f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}')
        
        # Create sheet for each building
        buildings = Building.objects.filter(is_active=True)
        
        for building in buildings:
            # Create worksheet for building
            safe_name = building.code.replace('/', '_')[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=safe_name)
            
            # Headers
            headers = ['Code', 'Name', 'Address', 'Floor', 'Block', 'Room', 'Office', 'Latitude', 'Longitude', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Building locations
            locations = building.location_set.select_related(
                'floor', 'block', 'room', 'office'
            ).order_by('location_code')
            
            for row, location in enumerate(locations, 2):
                ws.cell(row=row, column=1, value=location.location_code)
                ws.cell(row=row, column=2, value=location.name)
                ws.cell(row=row, column=3, value=location.address)
                ws.cell(row=row, column=4, value=location.floor.name if location.floor else '')
                ws.cell(row=row, column=5, value=location.block.name if location.block else '')
                ws.cell(row=row, column=6, value=location.room.name if location.room else '')
                ws.cell(row=row, column=7, value=location.office.name if location.office else '')
                ws.cell(row=row, column=8, value=str(location.latitude) if location.latitude else '')
                ws.cell(row=row, column=9, value=str(location.longitude) if location.longitude else '')
                ws.cell(row=row, column=10, value='Active' if location.is_active else 'Inactive')
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="location_hierarchy_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        
        return response


class LocationSummaryReportView(LoginRequiredMixin, View):
    """Generate location summary report with statistics."""
    
    def get(self, request):
        """Generate summary report."""
        export_format = request.GET.get('format', 'pdf').lower()
        
        if export_format == 'pdf':
            return self._generate_pdf_summary()
        elif export_format == 'excel':
            return self._generate_excel_summary()
        else:
            messages.error(request, 'Invalid report format.')
            return redirect('locations:reports')
    
    def _generate_pdf_summary(self):
        """Generate PDF summary report."""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="location_summary_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        
        story.append(Paragraph('Bangladesh Parliament Secretariat', title_style))
        story.append(Paragraph('Location Summary Report', title_style))
        story.append(Paragraph(f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        story.append(Spacer(1, 30))
        
        # Statistics table
        stats_data = [
            ['Metric', 'Count'],
            ['Total Locations', str(Location.objects.count())],
            ['Active Locations', str(Location.objects.filter(is_active=True).count())],
            ['Locations with Coordinates', str(Location.objects.filter(latitude__isnull=False, longitude__isnull=False).count())],
            ['Buildings', str(Building.objects.filter(is_active=True).count())],
            ['Floors', str(Floor.objects.filter(is_active=True).count())],
            ['Blocks', str(Block.objects.filter(is_active=True).count())],
            ['Rooms', str(Room.objects.filter(is_active=True).count())],
            ['Offices', str(Office.objects.filter(is_active=True).count())],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*72, 2*72])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 30))
        
        # Room type breakdown
        story.append(Paragraph('Room Type Distribution', styles['Heading2']))
        room_data = [['Room Type', 'Count']]
        
        for room_type, display_name in Room.ROOM_TYPES:
            count = Room.objects.filter(room_type=room_type, is_active=True).count()
            room_data.append([display_name, str(count)])
        
        room_table = Table(room_data, colWidths=[3*72, 2*72])
        room_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(room_table)
        
        doc.build(story)
        return response
    
    def _generate_excel_summary(self):
        """Generate Excel summary report."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Location Summary"
        
        # Title
        ws.cell(row=1, column=1, value='Bangladesh Parliament Secretariat')
        ws.cell(row=2, column=1, value='Location Summary Report')
        ws.cell(row=3, column=1, value=f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}')
        
        # Statistics
        ws.cell(row=5, column=1, value='Metric')
        ws.cell(row=5, column=2, value='Count')
        
        stats = [
            ('Total Locations', Location.objects.count()),
            ('Active Locations', Location.objects.filter(is_active=True).count()),
            ('Locations with Coordinates', Location.objects.filter(latitude__isnull=False, longitude__isnull=False).count()),
            ('Buildings', Building.objects.filter(is_active=True).count()),
            ('Floors', Floor.objects.filter(is_active=True).count()),
            ('Blocks', Block.objects.filter(is_active=True).count()),
            ('Rooms', Room.objects.filter(is_active=True).count()),
            ('Offices', Office.objects.filter(is_active=True).count()),
        ]
        
        for row, (metric, count) in enumerate(stats, 6):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=count)
        
        # Format headers
        for cell in ws[5]:
            cell.font = Font(bold=True)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="location_summary_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        
        return response


class CoordinatesExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Export location coordinates for mapping applications.
    Supports CSV, GeoJSON, and KML formats for various mapping tools.
    """
    permission_required = 'locations.view_location'
    
    def get(self, request):
        """Export coordinates in various formats."""
        export_format = request.GET.get('format', 'csv').lower()
        
        try:
            # Get locations with coordinates
            locations = Location.objects.filter(
                latitude__isnull=False,
                longitude__isnull=False,
                is_active=True
            ).select_related('building', 'floor', 'room', 'office')
            
            if not locations.exists():
                messages.warning(request, 'No locations with coordinates found.')
                return redirect('locations:reports')
            
            if export_format == 'csv':
                return self._export_csv(locations)
            elif export_format == 'geojson':
                return self._export_geojson(locations)
            elif export_format == 'kml':
                return self._export_kml(locations)
            else:
                messages.error(request, 'Invalid export format. Supported: CSV, GeoJSON, KML.')
                return redirect('locations:reports')
                
        except Exception as e:
            messages.error(request, f'Error during coordinate export: {str(e)}')
            return redirect('locations:reports')
    
    def _export_csv(self, locations):
        """Export as CSV for general use."""
        try:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="coordinates_{timezone.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'Location Code', 'Name', 'Latitude', 'Longitude', 'Building', 
                'Floor', 'Room', 'Office', 'Description'
            ])
            
            for location in locations:
                writer.writerow([
                    location.location_code,
                    location.name,
                    location.latitude,
                    location.longitude,
                    location.building.name if location.building else '',
                    location.floor.name if location.floor else '',
                    location.room.name if location.room else '',
                    location.office.name if location.office else '',
                    location.get_full_location_description()
                ])
            
            return response
            
        except Exception as e:
            messages.error(self.request, f'Error generating CSV: {str(e)}')
            return redirect('locations:reports')
    
    def _export_geojson(self, locations):
        """Export as GeoJSON for web mapping."""
        try:
            features = []
            
            for location in locations:
                try:
                    longitude = float(location.longitude)
                    latitude = float(location.latitude)
                except (ValueError, TypeError):
                    continue  # Skip invalid coordinates
                
                feature = {
                    "type": "Feature",
                    "geometry": {
                        "type": "Point",
                        "coordinates": [longitude, latitude]
                    },
                    "properties": {
                        "code": location.location_code,
                        "name": location.name,
                        "building": location.building.name if location.building else None,
                        "floor": location.floor.name if location.floor else None,
                        "room": location.room.name if location.room else None,
                        "office": location.office.name if location.office else None,
                        "description": location.get_full_location_description()
                    }
                }
                features.append(feature)
            
            geojson = {
                "type": "FeatureCollection",
                "features": features
            }
            
            response = HttpResponse(
                json.dumps(geojson, indent=2),
                content_type='application/geo+json'
            )
            response['Content-Disposition'] = f'attachment; filename="coordinates_{timezone.now().strftime("%Y%m%d")}.geojson"'
            
            return response
            
        except Exception as e:
            messages.error(self.request, f'Error generating GeoJSON: {str(e)}')
            return redirect('locations:reports')
    
    def _export_kml(self, locations):
        """Export as KML for Google Earth."""
        try:
            kml_content = '''<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>Bangladesh Parliament Secretariat Locations</name>
    <description>Location coordinates export from PIMS</description>
    
'''
            
            for location in locations:
                # Escape special characters for XML
                name = str(location.location_code) + ' - ' + str(location.name)
                name = name.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                
                building_name = location.building.name if location.building else 'N/A'
                floor_name = location.floor.name if location.floor else 'N/A'
                room_name = location.room.name if location.room else 'N/A'
                office_name = location.office.name if location.office else 'N/A'
                description = location.get_full_location_description()
                
                # Escape description content
                description = str(description).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                
                kml_content += f'''    <Placemark>
      <name>{name}</name>
      <description><![CDATA[
        <b>Building:</b> {building_name}<br/>
        <b>Floor:</b> {floor_name}<br/>
        <b>Room:</b> {room_name}<br/>
        <b>Office:</b> {office_name}<br/>
        <b>Description:</b> {description}
      ]]></description>
      <Point>
        <coordinates>{location.longitude},{location.latitude},0</coordinates>
      </Point>
    </Placemark>
    
'''
            
            kml_content += '''  </Document>
</kml>'''
            
            response = HttpResponse(kml_content, content_type='application/vnd.google-earth.kml+xml')
            response['Content-Disposition'] = f'attachment; filename="coordinates_{timezone.now().strftime("%Y%m%d")}.kml"'
            
            return response
            
        except Exception as e:
            messages.error(self.request, f'Error generating KML: {str(e)}')
            return redirect('locations:reports')

# ============================================================================
# Phase 9: QR Code Management Views
# ============================================================================
class LocationQRCodeView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """
    View and manage QR code for a specific location.
    """
    model = Location
    template_name = 'locations/qrcode.html'
    context_object_name = 'location'
    permission_required = 'locations.view_location'
    
    def get_context_data(self, **kwargs):
        """Add QR code context."""
        context = super().get_context_data(**kwargs)
        location = self.get_object()
        
        # Get active QR code using centralized function
        qr_code = get_qr_code_for_location(location)
        
        context.update({
            'qr_code': qr_code,
            'has_qr_code': qr_code is not None,
            'page_title': f'QR Code - {location.name}',
        })
        
        return context


class LocationQRCodeGenerateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Generate QR code for a specific location.
    """
    permission_required = 'locations.change_location'
    
    def post(self, request, pk):
        """Generate QR code for location using centralized function."""
        location = get_object_or_404(Location, pk=pk)
        
        try:
            # Use centralized QR generation
            qr_code = create_location_qr_code(location, request)
            
            if qr_code:
                messages.success(
                    request, 
                    f'QR code generated successfully for location "{location.name}"!'
                )
            else:
                messages.error(request, 'Error generating QR code. Please try again.')
                
        except Exception as e:
            messages.error(request, f'Error generating QR code: {str(e)}')
        
        return redirect('locations:qrcode', pk=pk)


class LocationQRCodeDownloadView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Download QR code image for a location.
    """
    permission_required = 'locations.view_location'
    
    def get(self, request, pk):
        """Download QR code image using the new model approach."""
        location = get_object_or_404(Location, pk=pk)
        qr_code = get_qr_code_for_location(location)
        
        if not qr_code or not qr_code.qr_code:
            messages.error(request, 'QR code not found. Please generate it first.')
            return redirect('locations:qrcode', pk=pk)
        
        try:
            response = HttpResponse(qr_code.qr_code.read(), content_type='image/png')
            filename = f"{location.name.replace(' ', '_')}_qrcode.png"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            messages.error(request, f'Error downloading QR code: {str(e)}')
            return redirect('locations:qrcode', pk=pk)


class BulkQRCodeGenerateView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    Generate QR codes for multiple locations in bulk.
    """
    template_name = 'locations/bulk_qrcode_generate.html'
    permission_required = 'locations.change_location'
    
    def get_context_data(self, **kwargs):
        """Add bulk generation context."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Bulk QR Code Generation'
        context['is_bulk_view'] = True  # Flag to modify template behavior
        
        # Get filter options
        context['buildings'] = Building.objects.filter(is_active=True).order_by('name')
        
        # Get all active locations
        all_locations = Location.objects.filter(is_active=True)
        
        # Get locations without QR codes
        locations_with_qr = LocationQRCode.objects.filter(
            is_active=True
        ).values_list('location_id', flat=True)
        
        locations_without_qr = all_locations.exclude(id__in=locations_with_qr)
        
        # Statistics for bulk operations
        context['bulk_stats'] = {
            'total_locations': all_locations.count(),
            'locations_without_qr': locations_without_qr.count(),
            'locations_with_qr': len(locations_with_qr),
            'coverage_percentage': round(
                (len(locations_with_qr) / all_locations.count()) * 100, 1
            ) if all_locations.count() > 0 else 0
        }
        
        # Locations for selection (limit for performance)
        context['locations_for_bulk'] = all_locations.select_related(
            'building', 'floor', 'block', 'room', 'office'
        ).order_by('building__name', 'name')[:500]  # Limit for UI performance
        
        # Office and room types for filtering
        context['office_types'] = Office.objects.filter(
            is_active=True
        ).values_list('office_type', flat=True).distinct()
        
        context['room_types'] = Room.objects.filter(
            is_active=True
        ).values_list('room_type', flat=True).distinct()
        
        return context
    
    def post(self, request):
        """Process bulk QR code generation using centralized functions."""
        # Get filter parameters
        building_id = request.POST.get('building')
        office_type = request.POST.get('office_type')
        room_type = request.POST.get('room_type')
        regenerate_existing = request.POST.get('regenerate_existing') == 'on'
        only_with_coordinates = request.POST.get('only_with_coordinates') == 'on'
        
        # Build queryset
        queryset = Location.objects.filter(is_active=True).select_related(
            'building', 'floor', 'block', 'room', 'office'
        )
        
        if building_id:
            queryset = queryset.filter(building_id=building_id)
        
        if office_type:
            queryset = queryset.filter(office__office_type=office_type)
        
        if room_type:
            queryset = queryset.filter(room__room_type=room_type)
        
        if only_with_coordinates:
            queryset = queryset.filter(
                latitude__isnull=False, 
                longitude__isnull=False
            )
        
        # Use centralized bulk generation
        results = bulk_generate_location_qr_codes(
            queryset, 
            request, 
            regenerate_existing=regenerate_existing
        )
        
        # Prepare success message
        message_parts = []
        if results['generated'] > 0:
            message_parts.append(f'Generated {results["generated"]} new QR codes')
        if results['updated'] > 0:
            message_parts.append(f'Updated {results["updated"]} existing QR codes')
        if results['skipped'] > 0:
            message_parts.append(f'Skipped {results["skipped"]} existing QR codes')
        if results['errors'] > 0:
            message_parts.append(f'{results["errors"]} errors occurred')
        
        if message_parts:
            if results['errors'] == 0:
                messages.success(request, '. '.join(message_parts) + '!')
            else:
                messages.warning(request, '. '.join(message_parts) + '.')
                # Show specific error locations
                if results['error_locations']:
                    for error in results['error_locations'][:5]:  # Show first 5 errors
                        messages.error(request, f'Error: {error}')
        
        return redirect('locations:bulk_qrcode_generate')


class BulkQRCodeDownloadView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Download multiple QR codes as a ZIP file.
    """
    permission_required = 'locations.view_location'
    
    def post(self, request):
        """Create and download ZIP file with QR codes."""
        # Get filter parameters
        building_id = request.POST.get('building')
        office_type = request.POST.get('office_type')
        room_type = request.POST.get('room_type')
        only_with_coordinates = request.POST.get('only_with_coordinates') == 'on'
        
        # Build queryset
        queryset = Location.objects.filter(is_active=True).select_related(
            'building', 'floor', 'block', 'room', 'office'
        )
        
        if building_id:
            queryset = queryset.filter(building_id=building_id)
        
        if office_type:
            queryset = queryset.filter(office__office_type=office_type)
        
        if room_type:
            queryset = queryset.filter(room__room_type=room_type)
        
        if only_with_coordinates:
            queryset = queryset.filter(
                latitude__isnull=False, 
                longitude__isnull=False
            )
        
        # Create ZIP file in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            added_count = 0
            
            for location in queryset:
                qr_code_path = f'qrcodes/locations/{location.location_code}.png'
                
                if default_storage.exists(qr_code_path):
                    try:
                        # Read QR code file
                        qr_file = default_storage.open(qr_code_path)
                        qr_content = qr_file.read()
                        qr_file.close()
                        
                        # Add to ZIP with descriptive filename
                        filename = f"{location.location_code}_{location.name.replace(' ', '_')}.png"
                        zip_file.writestr(filename, qr_content)
                        added_count += 1
                        
                    except Exception:
                        continue
        
        if added_count == 0:
            messages.error(request, 'No QR codes found for the selected criteria.')
            return redirect('locations:bulk_qrcode_generate')
        
        zip_buffer.seek(0)
        
        # Create response
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        filename = f'location_qrcodes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        messages.success(request, f'Downloaded {added_count} QR codes in ZIP file.')
        
        return response


# ============================================================================
# Phase 10: Advanced Features - Map View
# ============================================================================

class LocationMapView(LoginRequiredMixin, TemplateView):
    """
    Display interactive map with all location coordinates.
    Location: locations/views.py - Phase 10 Advanced Features
    """
    template_name = 'locations/location_map.html'
    
    def get_context_data(self, **kwargs):
        """Add map context data."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Location Map - Bangladesh Parliament Secretariat'
        
        # Get locations with coordinates
        locations_with_coords = Location.objects.filter(
            latitude__isnull=False,
            longitude__isnull=False,
            is_active=True
        ).select_related('building', 'floor', 'block', 'room', 'office')
        
        # Prepare map data
        map_locations = []
        for location in locations_with_coords:
            map_location = {
                'id': location.id,
                'code': location.location_code,
                'name': location.name,
                'latitude': float(location.latitude),
                'longitude': float(location.longitude),
                'building': {
                    'name': location.building.name,
                    'code': location.building.code
                } if location.building else None,
                'floor': {
                    'name': location.floor.name,
                    'number': location.floor.floor_number
                } if location.floor else None,
                'room': {
                    'name': location.room.name,
                    'number': location.room.room_number,
                    'type': location.room.get_room_type_display()
                } if location.room else None,
                'office': {
                    'name': location.office.name,
                    'code': location.office.office_code,
                    'type': location.office.get_office_type_display()
                } if location.office else None,
                'description': location.get_full_location_description(),
                'url': reverse('locations:detail', kwargs={'pk': location.pk}),
                'qr_url': reverse('locations:qrcode', kwargs={'pk': location.pk})
            }
            map_locations.append(map_location)
        
        context['map_locations'] = json.dumps(map_locations)
        context['total_mapped_locations'] = len(map_locations)
        context['total_locations'] = Location.objects.filter(is_active=True).count()
        
        # Calculate map center (average coordinates)
        if map_locations:
            avg_lat = sum(loc['latitude'] for loc in map_locations) / len(map_locations)
            avg_lng = sum(loc['longitude'] for loc in map_locations) / len(map_locations)
            context['map_center'] = [avg_lat, avg_lng]
        else:
            # Default to Bangladesh Parliament coordinates
            context['map_center'] = [23.7465, 90.3915]
        
        # Statistics for map
        context['map_stats'] = {
            'buildings_with_coords': len(set(
                loc['building']['code'] for loc in map_locations 
                if loc['building']
            )),
            'offices_with_coords': len([
                loc for loc in map_locations if loc['office']
            ]),
            'rooms_with_coords': len([
                loc for loc in map_locations if loc['room']
            ]),
            'coverage_percentage': round(
                (len(map_locations) / context['total_locations']) * 100, 1
            ) if context['total_locations'] > 0 else 0
        }
        
        # Filter options for map
        context['filter_options'] = {
            'buildings': list(set(
                (loc['building']['code'], loc['building']['name']) 
                for loc in map_locations if loc['building']
            )),
            'office_types': list(set(
                (loc['office']['type'], loc['office']['type']) 
                for loc in map_locations if loc['office']
            )),
            'room_types': list(set(
                (loc['room']['type'], loc['room']['type']) 
                for loc in map_locations if loc['room']
            ))
        }
        
        return context


# ============================================================================
# Additional Utility Views for QR Code Management
# ============================================================================

class QRCodeStatsView(LoginRequiredMixin, TemplateView):
    """
    Display QR code statistics and management dashboard.
    """
    template_name = 'locations/qrcode_stats.html'
    
    def get_context_data(self, **kwargs):
        """Add QR code statistics."""
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'QR Code Statistics'
        
        # Get all active locations
        all_locations = Location.objects.filter(is_active=True)
        
        # Count locations with and without QR codes
        locations_with_qr = 0
        locations_without_qr = 0
        qr_file_sizes = []
        
        for location in all_locations:
            qr_code_path = f'qrcodes/locations/{location.location_code}.png'
            if default_storage.exists(qr_code_path):
                locations_with_qr += 1
                try:
                    # Get file size
                    file_size = default_storage.size(qr_code_path)
                    qr_file_sizes.append(file_size)
                except:
                    pass
            else:
                locations_without_qr += 1
        
        context['qr_stats'] = {
            'total_locations': all_locations.count(),
            'with_qr_codes': locations_with_qr,
            'without_qr_codes': locations_without_qr,
            'coverage_percentage': round(
                (locations_with_qr / all_locations.count()) * 100, 1
            ) if all_locations.count() > 0 else 0,
            'total_file_size': sum(qr_file_sizes) if qr_file_sizes else 0,
            'average_file_size': round(
                sum(qr_file_sizes) / len(qr_file_sizes), 2
            ) if qr_file_sizes else 0
        }
        
        # Building breakdown
        building_stats = []
        for building in Building.objects.filter(is_active=True):
            building_locations = all_locations.filter(building=building)
            with_qr = 0
            
            for location in building_locations:
                qr_code_path = f'qrcodes/locations/{location.location_code}.png'
                if default_storage.exists(qr_code_path):
                    with_qr += 1
            
            building_stats.append({
                'building': building,
                'total_locations': building_locations.count(),
                'with_qr_codes': with_qr,
                'without_qr_codes': building_locations.count() - with_qr,
                'coverage_percentage': round(
                    (with_qr / building_locations.count()) * 100, 1
                ) if building_locations.count() > 0 else 0
            })
        
        context['building_stats'] = building_stats
        
        return context


class QRCodeCleanupView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Clean up orphaned QR code files.
    """
    permission_required = 'locations.delete_location'
    
    def post(self, request):
        """Clean up orphaned QR code files."""
        try:
            qr_directory = 'qrcodes/locations/'
            
            if not default_storage.exists(qr_directory):
                messages.info(request, 'No QR code directory found.')
                return redirect('locations:qrcode_stats')
            
            # Get all QR code files
            dirs, files = default_storage.listdir(qr_directory)
            
            # Get all valid location codes
            valid_codes = set(
                Location.objects.values_list('location_code', flat=True)
            )
            
            deleted_count = 0
            
            for file in files:
                if file.endswith('.png'):
                    # Extract location code from filename
                    location_code = file.replace('.png', '')
                    
                    if location_code not in valid_codes:
                        # Delete orphaned file
                        file_path = os.path.join(qr_directory, file)
                        default_storage.delete(file_path)
                        deleted_count += 1
            
            if deleted_count > 0:
                messages.success(
                    request, 
                    f'Cleanup completed. Deleted {deleted_count} orphaned QR code files.'
                )
            else:
                messages.info(request, 'No orphaned QR code files found.')
            
        except Exception as e:
            messages.error(request, f'Error during cleanup: {str(e)}')
        
        return redirect('locations:qrcode_stats')


class LocationQRCodeRegenerateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Regenerate QR code for a specific location.
    """
    permission_required = 'locations.change_location'
    
    def post(self, request, pk):
        """Regenerate QR code for location using centralized function."""
        location = get_object_or_404(Location, pk=pk)
        
        try:
            # Use centralized QR generation (will automatically deactivate old ones)
            qr_code = create_location_qr_code(location, request)
            
            if qr_code:
                messages.success(
                    request, 
                    f'QR code regenerated successfully for location "{location.name}"!'
                )
            else:
                messages.error(request, 'Error regenerating QR code. Please try again.')
            
        except Exception as e:
            messages.error(request, f'Error regenerating QR code: {str(e)}')
        
        return redirect('locations:qrcode', pk=pk)