"""
Views for Assignments app in PIMS (Parliament IT Inventory Management System)
Bangladesh Parliament Secretariat

This module defines comprehensive views for assignment management including:
- Assignment CRUD operations with device availability validation
- Status-based assignment filtering and management
- QR code generation, viewing, and download functionality
- Assignment returns, transfers, and extensions
- Search, filtering, and export capabilities
- Quick assignment workflow for efficient device assignment
- Dashboard and reporting views

Features:
- Modern class-based views with Bootstrap 5.3 styling
- Comprehensive assignment management with business logic validation
- QR code integration for assignment tracking
- Export functionality (Excel, PDF, CSV)
- Real-time status updates and device management
"""

# Standard library imports
import csv
import json
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
import zipfile

# Django core imports
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import Q, Count, Sum, Avg, Max, Min, F, Case, When
from django.http import (
    HttpResponse, HttpResponseRedirect, JsonResponse, 
    Http404, HttpResponseBadRequest, FileResponse
)
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, 
    TemplateView, FormView
)

# Third-party imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# Local app imports
from .models import Assignment, AssignmentQRCode
from .forms import (
    AssignmentForm, QuickAssignmentForm, AssignmentFilterForm,
    AssignmentReturnForm, AssignmentTransferForm
)
from devices.models import Device
from locations.models import Location
from pims.utils.qr_code import create_assignment_qr_code

User = get_user_model()


# ============================================================================
# Core Assignment CRUD Views
# ============================================================================

class AssignmentListView(LoginRequiredMixin, ListView):
    """
    Display list of all assignments with advanced filtering and search.
    """
    model = Assignment
    template_name = 'assignments/list.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def get_queryset(self):
        """Filter assignments based on search parameters."""
        queryset = Assignment.objects.select_related(
            'device', 'assigned_to', 'assigned_location', 'assigned_by'
        ).prefetch_related('qr_codes')
        
        # Apply search form filters
        form = AssignmentFilterForm(self.request.GET)
        if form.is_valid():
            # Search filter
            search = form.cleaned_data.get('search')
            if search:
                queryset = queryset.filter(
                    Q(assignment_id__icontains=search) |
                    Q(device__device_id__icontains=search) |
                    Q(device__name__icontains=search) |
                    Q(assigned_to__first_name__icontains=search) |
                    Q(assigned_to__last_name__icontains=search) |
                    Q(assigned_to__username__icontains=search) |
                    Q(purpose__icontains=search)
                )
            
            # Status filter
            status = form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(status=status)
            
            # Assignment type filter
            assignment_type = form.cleaned_data.get('assignment_type')
            if assignment_type:
                queryset = queryset.filter(assignment_type=assignment_type)
            
            # Employee filter
            assigned_to = form.cleaned_data.get('assigned_to')
            if assigned_to:
                queryset = queryset.filter(assigned_to=assigned_to)
            
            # Location filter
            assigned_location = form.cleaned_data.get('assigned_location')
            if assigned_location:
                queryset = queryset.filter(assigned_location=assigned_location)
            
            # Date range filters
            assigned_date_from = form.cleaned_data.get('assigned_date_from')
            assigned_date_to = form.cleaned_data.get('assigned_date_to')
            if assigned_date_from:
                queryset = queryset.filter(assigned_date__gte=assigned_date_from)
            if assigned_date_to:
                queryset = queryset.filter(assigned_date__lte=assigned_date_to)
            
            # Return date filters
            return_date_from = form.cleaned_data.get('expected_return_date_from')
            return_date_to = form.cleaned_data.get('expected_return_date_to')
            if return_date_from:
                queryset = queryset.filter(expected_return_date__gte=return_date_from)
            if return_date_to:
                queryset = queryset.filter(expected_return_date__lte=return_date_to)
            
            # Overdue filter
            is_overdue = form.cleaned_data.get('is_overdue')
            if is_overdue == 'yes':
                queryset = queryset.filter(
                    status='ASSIGNED',
                    expected_return_date__lt=date.today()
                )
            elif is_overdue == 'no':
                queryset = queryset.filter(
                    Q(expected_return_date__gte=date.today()) |
                    Q(expected_return_date__isnull=True) |
                    ~Q(status='ASSIGNED')
                )
        
        return queryset.order_by('-assigned_date', '-created_at')
    
    def get_context_data(self, **kwargs):
        """Add additional context data."""
        context = super().get_context_data(**kwargs)
        context['filter_form'] = AssignmentFilterForm(self.request.GET)
        
        # Statistics
        total_assignments = Assignment.objects.count()
        active_assignments = Assignment.objects.filter(status='ASSIGNED', is_active=True).count()
        overdue_assignments = Assignment.objects.filter(
            status='ASSIGNED',
            expected_return_date__lt=date.today()
        ).count()
        returned_assignments = Assignment.objects.filter(status='RETURNED').count()
        
        context.update({
            'total_assignments': total_assignments,
            'active_assignments': active_assignments,
            'overdue_assignments': overdue_assignments,
            'returned_assignments': returned_assignments,
        })
        
        return context


class AssignmentDetailView(LoginRequiredMixin, DetailView):
    """Display detailed information about an assignment."""
    model = Assignment
    template_name = 'assignments/detail.html'
    context_object_name = 'assignment'
    
    def get_queryset(self):
        return Assignment.objects.select_related(
            'device', 'assigned_to', 'assigned_location', 'assigned_by'
        ).prefetch_related('qr_codes')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assignment = self.object
        
        # Add QR code information
        qr_code = assignment.qr_codes.filter(is_active=True).first()
        context['qr_code'] = qr_code
        
        # Add assignment history and related information
        context['days_assigned'] = assignment.days_assigned()
        if assignment.status == 'OVERDUE':
            context['days_overdue'] = assignment.days_overdue()
        
        # Check if user can perform actions
        context['can_edit'] = self.request.user.has_perm('assignments.change_assignment')
        context['can_return'] = (
            assignment.status == 'ASSIGNED' and 
            self.request.user.has_perm('assignments.change_assignment')
        )
        context['can_transfer'] = (
            assignment.status == 'ASSIGNED' and 
            self.request.user.has_perm('assignments.change_assignment')
        )
        
        return context


class AssignmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new assignment."""
    model = Assignment
    form_class = AssignmentForm
    template_name = 'assignments/create.html'
    permission_required = 'assignments.add_assignment'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """Handle successful form submission."""
        with transaction.atomic():
            # Save assignment
            response = super().form_valid(form)
            assignment = self.object
            
            # Update device status
            if assignment.device and assignment.status == 'ASSIGNED':
                assignment.device.status = 'ASSIGNED'
                assignment.device.save()
            
            # Generate QR code
            try:
                create_assignment_qr_code(assignment, self.request)
                messages.success(
                    self.request, 
                    f'Assignment {assignment.assignment_id} created successfully! QR code generated.'
                )
            except Exception as e:
                messages.warning(
                    self.request,
                    f'Assignment created but QR code generation failed: {str(e)}'
                )
            
            return response
    
    def form_invalid(self, form):
        """Handle form validation errors."""
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


class AssignmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing assignment."""
    model = Assignment
    form_class = AssignmentForm
    template_name = 'assignments/edit.html'
    permission_required = 'assignments.change_assignment'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """Handle successful form submission."""
        original_device = Assignment.objects.get(pk=self.object.pk).device
        original_status = Assignment.objects.get(pk=self.object.pk).status
        
        with transaction.atomic():
            response = super().form_valid(form)
            assignment = self.object
            
            # Handle device changes
            if original_device != assignment.device:
                # Release old device
                if original_device and original_status == 'ASSIGNED':
                    original_device.status = 'AVAILABLE'
                    original_device.save()
                
                # Assign new device
                if assignment.device and assignment.status == 'ASSIGNED':
                    assignment.device.status = 'ASSIGNED'
                    assignment.device.save()
                
                # Generate new QR code
                try:
                    create_assignment_qr_code(assignment, self.request)
                    messages.success(self.request, 'Device changed and QR code updated.')
                except:
                    messages.warning(self.request, 'Device changed but QR generation failed.')
            
            # Handle status changes
            elif original_status != assignment.status:
                if assignment.status == 'RETURNED' and assignment.device:
                    assignment.device.status = 'AVAILABLE'
                    assignment.device.save()
                elif assignment.status == 'ASSIGNED' and assignment.device:
                    assignment.device.status = 'ASSIGNED'
                    assignment.device.save()
            
            messages.success(self.request, f'Assignment {assignment.assignment_id} updated successfully.')
            return response


class AssignmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete an assignment."""
    model = Assignment
    template_name = 'assignments/delete.html'
    permission_required = 'assignments.delete_assignment'
    success_url = reverse_lazy('assignments:list')
    
    def delete(self, request, *args, **kwargs):
        """Handle assignment deletion."""
        assignment = self.get_object()
        
        with transaction.atomic():
            # Free up device if assignment is active
            if assignment.status == 'ASSIGNED' and assignment.device:
                assignment.device.status = 'AVAILABLE'
                assignment.device.save()
            
            assignment_id = assignment.assignment_id
            response = super().delete(request, *args, **kwargs)
            messages.success(request, f'Assignment {assignment_id} deleted successfully.')
            return response


# ============================================================================
# Status-based Assignment Views
# ============================================================================

class ActiveAssignmentsView(AssignmentListView):
    """Display only active assignments."""
    template_name = 'assignments/active.html'
    
    def get_queryset(self):
        return super().get_queryset().filter(status='ASSIGNED', is_active=True)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Active Assignments'
        return context


class ReturnedAssignmentsView(AssignmentListView):
    """Display only returned assignments."""
    template_name = 'assignments/returned.html'
    
    def get_queryset(self):
        return super().get_queryset().filter(status='RETURNED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Returned Assignments'
        return context


class OverdueAssignmentsView(AssignmentListView):
    """Display only overdue assignments."""
    template_name = 'assignments/overdue.html'
    
    def get_queryset(self):
        return super().get_queryset().filter(
            status='ASSIGNED',
            expected_return_date__lt=date.today()
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Overdue Assignments'
        return context


class CancelledAssignmentsView(AssignmentListView):
    """Display only cancelled assignments."""
    template_name = 'assignments/cancelled.html'
    
    def get_queryset(self):
        return super().get_queryset().filter(status='CANCELLED')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Cancelled Assignments'
        return context


# ============================================================================
# Quick Assignment Workflow
# ============================================================================

class QuickAssignmentView(LoginRequiredMixin, TemplateView):
    """Quick assignment interface for fast device assignment."""
    template_name = 'assignments/quick.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['available_devices'] = Device.objects.filter(
            status='AVAILABLE', is_active=True
        ).count()
        context['active_employees'] = User.objects.filter(
            is_active=True, is_active_employee=True
        ).count()
        return context


class QuickAssignmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create assignment using quick form."""
    model = Assignment
    form_class = QuickAssignmentForm
    template_name = 'assignments/quick_create.html'
    permission_required = 'assignments.add_assignment'
    success_url = reverse_lazy('assignments:list')
    
    def form_valid(self, form):
        """Handle successful quick assignment creation."""
        form.instance.assigned_by = self.request.user
        form.instance.status = 'ASSIGNED'
        form.instance.is_active = True
        form.instance.condition_at_assignment = 'GOOD'
        
        with transaction.atomic():
            response = super().form_valid(form)
            assignment = self.object
            
            # Update device status
            if assignment.device:
                assignment.device.status = 'ASSIGNED'
                assignment.device.save()
            
            # Generate QR code
            try:
                create_assignment_qr_code(assignment, self.request)
            except:
                pass  # Silent fail for quick assignment
            
            messages.success(
                self.request,
                f'Quick assignment {assignment.assignment_id} created successfully!'
            )
            return response


# ============================================================================
# Assignment Operations (Return, Transfer, Extend)
# ============================================================================

class AssignmentReturnView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Handle assignment returns."""
    model = Assignment
    form_class = AssignmentReturnForm
    template_name = 'assignments/return.html'
    permission_required = 'assignments.change_assignment'
    
    def get_queryset(self):
        return Assignment.objects.filter(status='ASSIGNED')
    
    def form_valid(self, form):
        """Handle successful assignment return."""
        with transaction.atomic():
            response = super().form_valid(form)
            assignment = self.object
            
            # Update device status
            if assignment.device:
                assignment.device.status = 'AVAILABLE'
                assignment.device.save()
            
            messages.success(
                self.request,
                f'Assignment {assignment.assignment_id} returned successfully. Device is now available.'
            )
            return redirect('assignments:detail', pk=assignment.pk)


class AssignmentTransferView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Handle assignment transfers."""
    form_class = AssignmentTransferForm
    template_name = 'assignments/transfer.html'
    permission_required = 'assignments.change_assignment'
    
    def dispatch(self, request, *args, **kwargs):
        self.assignment = get_object_or_404(Assignment, pk=kwargs['pk'], status='ASSIGNED')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['assignment'] = self.assignment
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assignment'] = self.assignment
        return context
    
    def form_valid(self, form):
        """Handle successful assignment transfer."""
        with transaction.atomic():
            new_assigned_to = form.cleaned_data.get('new_assigned_to')
            new_assigned_location = form.cleaned_data.get('new_assigned_location')
            transfer_reason = form.cleaned_data.get('transfer_reason')
            
            # Update assignment
            if new_assigned_to:
                self.assignment.assigned_to = new_assigned_to
            if new_assigned_location:
                self.assignment.assigned_location = new_assigned_location
            
            # Add transfer note
            current_notes = self.assignment.assignment_notes or ''
            transfer_note = f"\nTransfer on {date.today()}: {transfer_reason}"
            self.assignment.assignment_notes = (current_notes + transfer_note).strip()
            
            self.assignment.save()
            
            # Regenerate QR code for new assignment details
            try:
                create_assignment_qr_code(self.assignment, self.request)
            except:
                pass
            
            messages.success(
                self.request,
                f'Assignment {self.assignment.assignment_id} transferred successfully.'
            )
            return redirect('assignments:detail', pk=self.assignment.pk)


class AssignmentExtendView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Handle assignment extensions via AJAX."""
    permission_required = 'assignments.change_assignment'
    
    def post(self, request, pk):
        assignment = get_object_or_404(Assignment, pk=pk, status='ASSIGNED')
        
        try:
            new_return_date = request.POST.get('new_return_date')
            reason = request.POST.get('reason', '')
            
            if new_return_date:
                new_date = datetime.strptime(new_return_date, '%Y-%m-%d').date()
                assignment.extend_assignment(new_date, reason)
                
                return JsonResponse({
                    'success': True,
                    'message': 'Assignment extended successfully.',
                    'new_date': new_date.isoformat()
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'New return date is required.'
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })


# ============================================================================
# QR Code Operations
# ============================================================================

class AssignmentQRCodeView(LoginRequiredMixin, DetailView):
    """Display assignment QR code."""
    model = Assignment
    template_name = 'assignments/qr_code.html'
    context_object_name = 'assignment'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assignment = self.object
        qr_code = assignment.qr_codes.filter(is_active=True).first()
        context['qr_code'] = qr_code
        return context


class AssignmentQRCodeDownloadView(LoginRequiredMixin, View):
    """Download assignment QR code."""
    
    def get(self, request, pk):
        qr_code = get_object_or_404(AssignmentQRCode, pk=pk, is_active=True)
        
        if qr_code.qr_code and qr_code.file_exists:
            response = FileResponse(
                qr_code.qr_code.open('rb'),
                as_attachment=True,
                filename=f'assignment_{qr_code.assignment.assignment_id}_qr.png'
            )
            return response
        else:
            messages.error(request, 'QR code file not found.')
            return redirect('assignments:detail', pk=qr_code.assignment.pk)


class RegenerateQRCodeView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Regenerate QR code for assignment."""
    permission_required = 'assignments.change_assignment'
    
    def post(self, request, pk):
        assignment = get_object_or_404(Assignment, pk=pk)
        
        try:
            qr_code = create_assignment_qr_code(assignment, request)
            if qr_code:
                messages.success(request, 'QR code regenerated successfully.')
            else:
                messages.error(request, 'Failed to regenerate QR code.')
        except Exception as e:
            messages.error(request, f'Error regenerating QR code: {str(e)}')
        
        return redirect('assignments:detail', pk=pk)


# ============================================================================
# Search and Filtering
# ============================================================================

class AssignmentSearchView(LoginRequiredMixin, ListView):
    """Advanced assignment search."""
    model = Assignment
    template_name = 'assignments/search.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Assignment.objects.select_related(
            'device', 'assigned_to', 'assigned_location'
        )
        
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(assignment_id__icontains=query) |
                Q(device__device_id__icontains=query) |
                Q(device__name__icontains=query) |
                Q(assigned_to__first_name__icontains=query) |
                Q(assigned_to__last_name__icontains=query) |
                Q(purpose__icontains=query)
            )
        
        return queryset.order_by('-assigned_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['query'] = self.request.GET.get('q', '')
        return context


class AssignmentFilterView(LoginRequiredMixin, ListView):
    """Advanced assignment filtering."""
    model = Assignment
    template_name = 'assignments/filter.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Assignment.objects.select_related(
            'device', 'assigned_to', 'assigned_location'
        )
        
        form = AssignmentFilterForm(self.request.GET)
        if form.is_valid():
            # Apply all filters as in AssignmentListView
            # (Implementation similar to get_queryset in AssignmentListView)
            pass
        
        return queryset.order_by('-assigned_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = AssignmentFilterForm(self.request.GET)
        return context


# ============================================================================
# Export and Reports
# ============================================================================

class AssignmentExportView(LoginRequiredMixin, TemplateView):
    """Assignment export options interface."""
    template_name = 'assignments/export.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get statistics for export preview
        context.update({
            'total_assignments': Assignment.objects.count(),
            'active_assignments': Assignment.objects.filter(status='ACTIVE').count(),
            'returned_assignments': Assignment.objects.filter(status='RETURNED').count(),
            'overdue_assignments': Assignment.objects.filter(
                status='ACTIVE',
                expected_return_date__lt=timezone.now().date()
            ).count(),
            'temporary_assignments': Assignment.objects.filter(assignment_type='TEMPORARY').count(),
            'permanent_assignments': Assignment.objects.filter(assignment_type='PERMANENT').count(),
            'employees': User.objects.filter(is_active=True, is_active_employee=True),
            'locations': Location.objects.filter(is_active=True),
        })
        
        return context


class AssignmentCSVExportView(LoginRequiredMixin, View):
    """Export assignments to CSV with filtering."""
    
    def get(self, request):
        return self._export_csv(request)
    
    def post(self, request):
        return self._export_csv(request)
    
    def _export_csv(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="assignments_{date.today()}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Assignment ID', 'Device ID', 'Device Name', 'Device Brand', 'Device Model',
            'Employee Name', 'Employee ID', 'Location', 'Assignment Type', 'Status',
            'Assigned Date', 'Expected Return Date', 'Actual Return Date',
            'Purpose', 'Days Assigned', 'Condition at Assignment', 'Condition at Return',
            'Assigned By'
        ])
        
        # Get filtered queryset
        queryset = self._get_filtered_queryset(request)
        
        for assignment in queryset:
            writer.writerow([
                assignment.assignment_id,
                assignment.device.device_id if assignment.device else '',
                assignment.device.name if assignment.device else '',
                assignment.device.brand if assignment.device else '',
                assignment.device.model if assignment.device else '',
                assignment.assigned_to.get_full_name() if assignment.assigned_to else '',
                getattr(assignment.assigned_to, 'employee_id', '') if assignment.assigned_to else '',
                assignment.assigned_location.name if assignment.assigned_location else '',
                assignment.get_assignment_type_display(),
                assignment.get_status_display(),
                assignment.assigned_date,
                assignment.expected_return_date or '',
                assignment.actual_return_date or '',
                assignment.purpose,
                assignment.days_assigned(),
                assignment.get_condition_at_assignment_display(),
                assignment.get_condition_at_return_display() if assignment.condition_at_return else '',
                assignment.assigned_by.get_full_name() if assignment.assigned_by else ''
            ])
        
        return response



class AssignmentExcelExportView(LoginRequiredMixin, View):
    """Export assignments to Excel with multiple sheets."""
    
    def get(self, request):
        return self._export_excel(request)
    
    def post(self, request):
        return self._export_excel(request)
    
    def _export_excel(self, request):
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create main assignments sheet
        ws_main = wb.create_sheet(title="All Assignments")
        
        # Headers with styling
        headers = [
            'Assignment ID', 'Device ID', 'Device Name', 'Employee Name',
            'Location', 'Assignment Type', 'Status', 'Assigned Date',
            'Expected Return Date', 'Actual Return Date', 'Purpose', 'Days Assigned'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Get filtered data
        queryset = self._get_filtered_queryset(request)
        
        for row, assignment in enumerate(queryset, 2):
            ws_main.cell(row=row, column=1, value=assignment.assignment_id)
            ws_main.cell(row=row, column=2, value=assignment.device.device_id if assignment.device else '')
            ws_main.cell(row=row, column=3, value=assignment.device.name if assignment.device else '')
            ws_main.cell(row=row, column=4, value=assignment.assigned_to.get_full_name() if assignment.assigned_to else '')
            ws_main.cell(row=row, column=5, value=assignment.assigned_location.name if assignment.assigned_location else '')
            ws_main.cell(row=row, column=6, value=assignment.get_assignment_type_display())
            ws_main.cell(row=row, column=7, value=assignment.get_status_display())
            ws_main.cell(row=row, column=8, value=assignment.assigned_date)
            ws_main.cell(row=row, column=9, value=assignment.expected_return_date)
            ws_main.cell(row=row, column=10, value=assignment.actual_return_date)
            ws_main.cell(row=row, column=11, value=assignment.purpose)
            ws_main.cell(row=row, column=12, value=assignment.days_assigned())
        
        # Auto-adjust column widths
        for column in ws_main.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_main.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Create summary sheet
        ws_summary = wb.create_sheet(title="Summary", index=0)
        
        # Summary statistics
        total_assignments = queryset.count()
        active_count = queryset.filter(status='ACTIVE').count()
        returned_count = queryset.filter(status='RETURNED').count()
        overdue_count = queryset.filter(
            status='ACTIVE',
            expected_return_date__lt=timezone.now().date()
        ).count()
        
        summary_data = [
            ['Assignment Statistics', ''],
            ['Total Assignments', total_assignments],
            ['Active Assignments', active_count],
            ['Returned Assignments', returned_count],
            ['Overdue Assignments', overdue_count],
            ['', ''],
            ['Export Date', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Generated by', request.user.get_full_name()],
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=value)
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="assignments_{date.today()}.xlsx"'
        wb.save(response)
        return response

class AssignmentPDFExportView(LoginRequiredMixin, View):
    """Export assignments to PDF report."""
    
    def get(self, request):
        return self._export_pdf(request)
    
    def post(self, request):
        return self._export_pdf(request)
    
    def _export_pdf(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="assignments_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("Assignment Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Report info
        info_data = [
            ['Report Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Generated by:', request.user.get_full_name()],
            ['System:', 'PIMS - Bangladesh Parliament Secretariat'],
        ]
        
        info_table = Table(info_data)
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 30))
        
        # Get filtered data
        queryset = self._get_filtered_queryset(request)
        
        # Summary statistics
        summary_para = Paragraph(f"<b>Summary:</b> Total of {queryset.count()} assignments", styles['Normal'])
        elements.append(summary_para)
        elements.append(Spacer(1, 20))
        
        # Assignment table
        table_data = [['Assignment ID', 'Device', 'Employee', 'Status', 'Assigned Date']]
        
        for assignment in queryset[:50]:  # Limit for PDF readability
            table_data.append([
                assignment.assignment_id,
                f"{assignment.device.device_id}" if assignment.device else 'N/A',
                assignment.assigned_to.get_full_name() if assignment.assigned_to else 'N/A',
                assignment.get_status_display(),
                assignment.assigned_date.strftime('%Y-%m-%d')
            ])
        
        if queryset.count() > 50:
            table_data.append(['...', '...', '...', '...', f'And {queryset.count() - 50} more'])
        
        assignment_table = Table(table_data)
        assignment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(assignment_table)
        doc.build(elements)
        
        return response

class AssignmentJSONExportView(LoginRequiredMixin, View):
    """Export assignments to JSON format."""
    
    def get(self, request):
        return self._export_json(request)
    
    def post(self, request):
        return self._export_json(request)
    
    def _export_json(self, request):
        # Get filtered queryset
        queryset = self._get_filtered_queryset(request)
        
        assignments_data = []
        for assignment in queryset:
            assignments_data.append({
                'assignment_id': assignment.assignment_id,
                'device': {
                    'device_id': assignment.device.device_id if assignment.device else None,
                    'name': assignment.device.name if assignment.device else None,
                    'brand': assignment.device.brand if assignment.device else None,
                    'model': assignment.device.model if assignment.device else None,
                } if assignment.device else None,
                'assigned_to': {
                    'id': assignment.assigned_to.id if assignment.assigned_to else None,
                    'username': assignment.assigned_to.username if assignment.assigned_to else None,
                    'full_name': assignment.assigned_to.get_full_name() if assignment.assigned_to else None,
                } if assignment.assigned_to else None,
                'assigned_location': {
                    'id': assignment.assigned_location.id if assignment.assigned_location else None,
                    'name': assignment.assigned_location.name if assignment.assigned_location else None,
                } if assignment.assigned_location else None,
                'assignment_type': assignment.assignment_type,
                'status': assignment.status,
                'assigned_date': assignment.assigned_date.isoformat(),
                'expected_return_date': assignment.expected_return_date.isoformat() if assignment.expected_return_date else None,
                'actual_return_date': assignment.actual_return_date.isoformat() if assignment.actual_return_date else None,
                'purpose': assignment.purpose,
                'days_assigned': assignment.days_assigned(),
                'condition_at_assignment': assignment.condition_at_assignment,
                'condition_at_return': assignment.condition_at_return,
                'created_at': assignment.created_at.isoformat(),
                'updated_at': assignment.updated_at.isoformat(),
            })
        
        export_data = {
            'export_info': {
                'generated_at': timezone.now().isoformat(),
                'generated_by': request.user.username,
                'total_records': len(assignments_data),
                'system': 'PIMS - Bangladesh Parliament Secretariat'
            },
            'assignments': assignments_data
        }
        
        response = HttpResponse(
            json.dumps(export_data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="assignments_{date.today()}.json"'
        
        return response
    
    def _get_filtered_queryset(self, request):
        """Get filtered queryset based on request parameters."""
        queryset = Assignment.objects.select_related(
            'device', 'assigned_to', 'assigned_location', 'assigned_by'
        ).order_by('-assigned_date')
        
        # Apply filters from POST or GET data
        data = request.POST if request.method == 'POST' else request.GET
        
        # Status filter
        statuses = data.getlist('status')
        if statuses:
            queryset = queryset.filter(status__in=statuses)
        
        # Assignment type filter
        assignment_types = data.getlist('assignment_type')
        if assignment_types:
            queryset = queryset.filter(assignment_type__in=assignment_types)
        
        # Date range filters
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        if date_from:
            queryset = queryset.filter(assigned_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(assigned_date__lte=date_to)
        
        # Employee filter
        employee = data.get('employee')
        if employee:
            queryset = queryset.filter(assigned_to_id=employee)
        
        # Location filter
        location = data.get('location')
        if location:
            queryset = queryset.filter(assigned_location_id=location)
        
        return queryset

# Mixin for common export functionality
class AssignmentExportMixin:
    """Mixin to provide common export functionality."""
    
    def _get_filtered_queryset(self, request):
        """Get filtered queryset based on request parameters."""
        queryset = Assignment.objects.select_related(
            'device', 'assigned_to', 'assigned_location', 'assigned_by'
        ).order_by('-assigned_date')
        
        # Apply filters from POST or GET data
        data = request.POST if request.method == 'POST' else request.GET
        
        # Status filter
        statuses = data.getlist('status')
        if statuses:
            queryset = queryset.filter(status__in=statuses)
        
        # Assignment type filter
        assignment_types = data.getlist('assignment_type')
        if assignment_types:
            queryset = queryset.filter(assignment_type__in=assignment_types)
        
        # Date range filters
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        if date_from:
            queryset = queryset.filter(assigned_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(assigned_date__lte=date_to)
        
        # Employee filter
        employee = data.get('employee')
        if employee:
            queryset = queryset.filter(assigned_to_id=employee)
        
        # Location filter
        location = data.get('location')
        if location:
            queryset = queryset.filter(assigned_location_id=location)
        
        return queryset

class AssignmentReportsView(LoginRequiredMixin, TemplateView):
    """Assignment reports and analytics dashboard."""
    template_name = 'assignments/reports.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Generate comprehensive statistics for reports
        context.update({
            'total_assignments': Assignment.objects.count(),
            'active_assignments': Assignment.objects.filter(status='ACTIVE').count(),
            'returned_assignments': Assignment.objects.filter(status='RETURNED').count(),
            'overdue_assignments': Assignment.objects.filter(
                status='ACTIVE',
                expected_return_date__lt=timezone.now().date()
            ).count(),
            'assignments_this_month': Assignment.objects.filter(
                assigned_date__month=timezone.now().month,
                assigned_date__year=timezone.now().year
            ).count(),
            'top_devices': Assignment.objects.values('device__name', 'device__device_id')
                .annotate(assignment_count=Count('id'))
                .order_by('-assignment_count')[:5],
            'top_employees': Assignment.objects.values('assigned_to__first_name', 'assigned_to__last_name')
                .annotate(assignment_count=Count('id'))
                .order_by('-assignment_count')[:5],
        })
        
        return context